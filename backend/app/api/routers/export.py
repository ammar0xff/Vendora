"""Export endpoints — Excel download for accounting tables."""
from fastapi import APIRouter, Depends, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from app.db.base import get_db
from app.dependencies import get_current_user
from fastapi.responses import StreamingResponse
import io

router = APIRouter(prefix="/export", tags=["export"])


async def _to_excel(rows: list[dict], sheet_name: str = "Sheet1") -> bytes:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(bold=True, color="ffffff", size=11)
    if not rows:
        wb.save(buf := io.BytesIO())
        return buf.getvalue()
    headers = list(rows[0].keys())
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            val = row[h]
            if isinstance(val, float):
                ws.cell(row=ri, column=ci, value=val).number_format = '#,##0.00'
            else:
                ws.cell(row=ri, column=ci, value=str(val) if val is not None else '')
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 20
    wb.save(buf := io.BytesIO())
    return buf.getvalue()


@router.get("/products")
async def export_products(db: AsyncSession = Depends(get_db), _=Depends(get_current_user)):
    rows = (await db.execute(text("""
        SELECT p.name, p.company, p.unit, p.retail_price, p.wholesale_price, p.cost_price,
               c.name as category, s.name as subcategory, p.stock_status
        FROM products p
        LEFT JOIN subcategories s ON s.id = p.subcategory_id
        LEFT JOIN categories c ON c.id = s.category_id
        ORDER BY p.name
    """))).mappings().fetchall()
    data = [dict(r) for r in rows]
    excel = await _to_excel(data, "المنتجات")
    return StreamingResponse(io.BytesIO(excel), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=products.xlsx"})


@router.get("/sales")
async def export_sales(from_date: str = Query(None), to_date: str = Query(None), db: AsyncSession = Depends(get_db), _=Depends(get_current_user)):
    params = {}
    whr = ["s.status IN ('confirmed','returned')"]
    if from_date:
        whr.append("DATE(s.created_at) >= :fd"); params["fd"] = from_date
    if to_date:
        whr.append("DATE(s.created_at) <= :td"); params["td"] = to_date
    rows = (await db.execute(text(f"""
        SELECT s.invoice_number, s.created_at::date as date, c.name as customer,
               s.total, s.discount_amount, s.net_total, s.paid_amount, s.is_credit,
               s.payment_method, u.full_name as cashier, s.status, w.name as warehouse
        FROM sales s
        LEFT JOIN customers c ON c.id = s.customer_id
        LEFT JOIN users u ON u.id = s.cashier_id
        LEFT JOIN warehouses w ON w.id = s.warehouse_id
        WHERE {' AND '.join(whr)}
        ORDER BY s.created_at DESC
    """), params)).mappings().fetchall()
    data = [dict(r) for r in rows]
    excel = await _to_excel(data, "المبيعات")
    return StreamingResponse(io.BytesIO(excel), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=sales.xlsx"})


@router.get("/stock")
async def export_stock(db: AsyncSession = Depends(get_db), _=Depends(get_current_user)):
    rows = (await db.execute(text("""
        SELECT p.name, p.company, w.name as warehouse, COALESCE(s.qty,0) as qty, p.unit,
               p.retail_price, p.cost_price, (COALESCE(s.qty,0) * p.cost_price) as stock_value
        FROM products p
        CROSS JOIN warehouses w
        LEFT JOIN LATERAL (
            SELECT SUM(CASE WHEN sm.movement_type IN ('purchase','transfer_in','return_in','opening_stock','adjustment_in') THEN sm.qty
                           WHEN sm.movement_type IN ('sale','transfer_out','damage','adjustment_out') THEN -sm.qty
                       END) as qty
            FROM stock_movements sm WHERE sm.product_id=p.id AND sm.warehouse_id=w.id
        ) s ON true
        WHERE p.stock_status='tracked'
        ORDER BY p.name, w.name
    """))).mappings().fetchall()
    data = [dict(r) for r in rows]
    excel = await _to_excel(data, "المخزون")
    return StreamingResponse(io.BytesIO(excel), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=stock.xlsx"})


@router.get("/customers")
async def export_customers(db: AsyncSession = Depends(get_db), _=Depends(get_current_user)):
    rows = (await db.execute(text("""
        SELECT c.name, c.phone, c.balance_due, c.credit_limit
        FROM customers c
        ORDER BY c.name
    """))).mappings().fetchall()
    data = [dict(r) for r in rows]
    excel = await _to_excel(data, "العملاء")
    return StreamingResponse(io.BytesIO(excel), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=customers.xlsx"})


@router.get("/suppliers")
async def export_suppliers(db: AsyncSession = Depends(get_db), _=Depends(get_current_user)):
    rows = (await db.execute(text("""
        SELECT s.name, s.phone, s.balance
        FROM suppliers s
        ORDER BY s.name
    """))).mappings().fetchall()
    data = [dict(r) for r in rows]
    excel = await _to_excel(data, "الموردون")
    return StreamingResponse(io.BytesIO(excel), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=suppliers.xlsx"})


@router.get("/expenses")
async def export_expenses(from_date: str = Query(None), to_date: str = Query(None), db: AsyncSession = Depends(get_db), _=Depends(get_current_user)):
    params = {}
    whr = ["1=1"]
    if from_date:
        whr.append("DATE(e.date) >= :fd"); params["fd"] = from_date
    if to_date:
        whr.append("DATE(e.date) <= :td"); params["td"] = to_date
    rows = (await db.execute(text(f"""
        SELECT e.date, e.description, e.amount, ev.name as vendor, e.status,
               e.payment_method, e.is_recurring
        FROM expenses e
        LEFT JOIN expense_vendors ev ON ev.id = e.vendor_id
        WHERE {' AND '.join(whr)}
        ORDER BY e.date DESC
    """), params)).mappings().fetchall()
    data = [dict(r) for r in rows]
    excel = await _to_excel(data, "المصروفات")
    return StreamingResponse(io.BytesIO(excel), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=expenses.xlsx"})


@router.get("/purchases")
async def export_purchases(from_date: str = Query(None), to_date: str = Query(None), db: AsyncSession = Depends(get_db), _=Depends(get_current_user)):
    params = {}
    whr = ["1=1"]
    if from_date:
        whr.append("DATE(p.created_at) >= :fd"); params["fd"] = from_date
    if to_date:
        whr.append("DATE(p.created_at) <= :td"); params["td"] = to_date
    rows = (await db.execute(text(f"""
        SELECT p.po_number, p.created_at::date as date, s.name as supplier,
               p.total_cost, p.status, w.name as warehouse
        FROM purchases p
        LEFT JOIN suppliers s ON s.id = p.supplier_id
        LEFT JOIN warehouses w ON w.id = p.warehouse_id
        WHERE {' AND '.join(whr)}
        ORDER BY p.created_at DESC
    """), params)).mappings().fetchall()
    data = [dict(r) for r in rows]
    excel = await _to_excel(data, "المشتريات")
    return StreamingResponse(io.BytesIO(excel), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=purchases.xlsx"})
