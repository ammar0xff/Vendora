import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\Neisco_Comer_Price_List_2026.xlsx', data_only=True)

# Check UPVC Ball Valves sheet for these unmatched sizes
ws = wb['UPVC Ball Valves']
print("=== UPVC Ball Valves ===")
for row in ws.iter_rows(min_row=4, values_only=True):
    code, size, pack, box, price = row if len(row) >= 5 else (row[0] if row else None, None, None, None, None)
    if code and size:
        code_str = str(code).strip()
        size_str = str(size).strip()
        if size_str in ['110*4"', '32*1"', '32MM']:
            print(f'{code_str:15s} | {size_str:15s} | price={price} | pack={pack} | box={box}')
        # Check for any entry with 110 or similar
        if '110' in size_str:
            print(f'{code_str:15s} | {size_str:15s} | price={price}')

# Check Sheet2 for Adaptor Series
ws = wb['Adaptor Series Plain-Threaded']
print("\n=== Adaptor Series ===")
for row in ws.iter_rows(min_row=4, values_only=True):
    code, size, pack, box, price = row if len(row) >= 5 else (row[0] if row else None, None, None, None, None)
    if code and size:
        code_str = str(code).strip()
        size_str = str(size).strip()
        if size_str in ['20*3/4"', '32*3/4"']:
            print(f'{code_str:15s} | {size_str:15s} | price={price}')

# Check Imperial Plain Fittings for the unmatched
ws = wb['Imperial Plain Fittings BS']
print("\n=== Imperial Plain Fittings BS ===")
for row in ws.iter_rows(min_row=4, values_only=True):
    code, size, pack, box, price = row if len(row) >= 5 else (row[0] if row else None, None, None, None, None)
    if code and size:
        code_str = str(code).strip()
        size_str = str(size).strip()
        if size_str in ['2"', '1 1/2"*1/2"', '20*1/2"', '25*3/4"', '32*1"', '3"']:
            print(f'{code_str:15s} | {size_str:15s} | price={price}')

# Check Saddles & Bushings
ws = wb['Saddles & Bushings Threaded']
print("\n=== Saddles & Bushings Threaded ===")
for row in ws.iter_rows(min_row=4, values_only=True):
    code, size, pack, box, price = row if len(row) >= 5 else (row[0] if row else None, None, None, None, None)
    if code and size:
        code_str = str(code).strip()
        size_str = str(size).strip()
        if size_str in ['160*1.5"', '2"', '3/4"*1"'] or ('160' in size_str):
            print(f'{code_str:15s} | {size_str:15s} | price={price}')
