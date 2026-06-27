"""
Fix رف 7 quantities — the import script took max of qty_cols [3, 5],
which picked the price (col 5) instead of the actual quantity (col 3).
"""
import asyncio
import uuid
from decimal import Decimal
from openpyxl import load_workbook

from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine, async_sessionmaker
from sqlalchemy import select

import app.models.user
import app.models.warehouse
from app.models.product import Product
from app.models.stock import StockMovement, MovementType

DATABASE_URL = "postgresql+asyncpg://postgres:postgres@db:5432/inventory_db"
BASE = r'/app/LISTS/رف 7.xlsx'

SKIP_NAMES = {
    'اسم الصنف', 'الصنف', 'العدد', 'البيان', 'رقم الرف', 'رقم الرف ا',
    'اجمالى', 'الاجمالى', 'المخزن', 'نوع الرف', 'الاجمالي',
    'الوحدة', 'سعر البيع',
}


def is_noise(val):
    if not val:
        return True
    v = val.strip()
    if not v:
        return True
    if v in SKIP_NAMES:
        return True
    if len(v) <= 2 and not any('\u0600' <= c <= '\u06FF' for c in v):
        return True
    if len(set(v)) <= 2 and len(v) > 5:
        return True
    return False


def extract_raf7():
    """Read رف 7 with correct mapping: name=col2, qty=col3, price=col5"""
    wb = load_workbook(BASE, read_only=True, data_only=True)
    ws = wb.active
    products = []
    for row in ws.iter_rows(values_only=True):
        row = list(row)
        # name in col 2
        if len(row) < 4:
            continue
        val = row[2]
        if val is None or not isinstance(val, str):
            continue
        if is_noise(val):
            continue
        name = val.strip()
        if len(name) <= 2:
            continue

        # qty in col 3
        qty = 0
        if row[3] is not None:
            try:
                qty = float(str(row[3]).replace(',', '').replace(' ', ''))
                if qty == int(qty):
                    qty = int(qty)
            except:
                qty = 0

        # price in col 5
        price = 0
        if len(row) > 5 and row[5] is not None:
            try:
                price = float(str(row[5]).replace(',', '').replace(' ', ''))
                if price == int(price):
                    price = int(price)
            except:
                price = 0

        if qty > 0 or price > 0:
            products.append({'name': name, 'qty': qty, 'price': price})

    wb.close()
    print(f'Extracted {len(products)} products from رف 7')
    return products


async def main():
    print('=' * 60)
    print('FIXING رف 7 QUANTITIES')
    print('=' * 60)

    data = extract_raf7()

    engine = create_async_engine(DATABASE_URL)
    SessionLocal = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with SessionLocal() as db:
        fixed_qty = 0
        fixed_price = 0
        not_found = 0

        for item in data:
            # Find matching products (there could be multiple with same name)
            result = await db.execute(select(Product).where(Product.name == item['name']))
            prods = result.scalars().all()

            if not prods:
                not_found += 1
                continue

            for prod in prods:
                # Find the opening_stock movement for this product
                mvt_result = await db.execute(
                    select(StockMovement).where(
                        StockMovement.product_id == prod.id,
                        StockMovement.movement_type == MovementType.opening_stock,
                    )
                )
                mvts = mvt_result.scalars().all()
                for mvt in mvts:
                    # Fix qty
                    correct_qty = Decimal(str(item['qty']))
                    if mvt.qty != correct_qty:
                        mvt.qty = correct_qty
                        fixed_qty += 1

                    # Fix unit_price if we have it
                    if item['price'] > 0:
                        correct_price = Decimal(str(item['price']))
                        if mvt.unit_price != correct_price:
                            mvt.unit_price = correct_price
                            fixed_price += 1

                # Also fix retail_price on product
                if item['price'] > 0:
                    if prod.retail_price == 0 or prod.retail_price < Decimal(str(item['price'])):
                        prod.retail_price = Decimal(str(item['price']))

        await db.commit()
        print(f'  Qty fixed:     {fixed_qty}')
        print(f'  Price fixed:   {fixed_price}')
        print(f'  Not found:     {not_found}')

    await engine.dispose()
    print('\nDONE')


if __name__ == '__main__':
    asyncio.run(main())
