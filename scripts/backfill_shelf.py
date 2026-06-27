"""
Backfill shelf_number for products from Excel inventory files.
Shelf numbers are in column 0 of most files (رقم الرف column).
"""
import asyncio
import os
import glob
from openpyxl import load_workbook

from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine, async_sessionmaker
from sqlalchemy import select

import app.models.user
import app.models.warehouse
from app.models.product import Product

DATABASE_URL = "postgresql+asyncpg://postgres:postgres@db:5432/inventory_db"
BASE = r'/app/LISTS'

# Per-file config: (name_col, shelf_col)
FILE_CONFIG = {
    0:  (1, 0),          # جرد الرفوف تحت الصندلة
    1:  (1, 0),          # جرد مخزن البولى
    2:  (2, None),       # جرد المخزن الكبير بولي — no shelf col
    3:  (1, 0),          # جرد رف 4
    4:  (0, None),       # جرد الصندلة — 2-col format, no shelf
    5:  (1, 0),          # جرد مخزن الحديد
    6:  (1, 0),          # رف 1 (2)
    7:  (1, 0),          # رف 10
    8:  (1, 0),          # رف 11
    9:  (1, 0),          # رف 12
    10: (1, 0),          # رف 13
    11: (1, 0),          # رف 14
    12: (1, 0),          # رف 2
    13: (1, 0),          # رف 3
    14: (1, 0),          # رف 5
    15: (1, 0),          # رف 6
    16: (2, None),       # رف 7 — shelf codes in col 1, but they're sequential, not useful
    17: (1, 0),          # رف 8
    18: (1, 0),          # رف 9
    19: (1, None),       # مخزن نواكل — no shelf column
    20: (1, 0),          # مخزن داخلى بولى 15
    21: (1, None),       # مخزن داخلي0 — shelf is actually column 0 ('رقم الرف ا')
}

SKIP_NAMES = {
    'اسم الصنف', 'الصنف', 'العدد', 'البيان', 'رقم الرف', 'رقم الرف ا',
    'اجمالى', 'الاجمالى', 'المخزن', 'نوع الرف', 'الاجمالي',
    'الوحدة', 'سعر البيع',
}


def is_noise(val):
    if not val:
        return True
    v = val.strip()
    if not v:
        return True
    if v in SKIP_NAMES:
        return True
    if len(v) <= 2 and not any('\u0600' <= c <= '\u06FF' for c in v):
        return True
    if len(set(v)) <= 2 and len(v) > 5:
        return True
    return False


def extract_shelf_numbers():
    files = sorted(glob.glob(os.path.join(BASE, '*.xlsx')))
    shelf_data = []  # [(name, shelf), ...]

    for idx, f in enumerate(files):
        cfg = FILE_CONFIG.get(idx)
        if cfg is None:
            continue
        nc, sc = cfg
        if sc is None:
            continue

        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active

        count = 0
        for row in ws.iter_rows(values_only=True):
            row = list(row)
            if nc >= len(row) or sc >= len(row):
                continue

            val = row[nc]
            if val is None or not isinstance(val, str):
                continue
            if is_noise(val):
                continue
            name = val.strip()
            if len(name) > 10 and 'مخزن' in name and name.count(' ') > 2:
                continue

            shelf = row[sc]
            if shelf is not None:
                shelf_str = str(shelf).strip()
                if shelf_str and len(shelf_str) < 20:
                    shelf_data.append((name, shelf_str))
                    count += 1

        wb.close()
        fname = os.path.basename(f)
        print(f'  [{idx:2d}] {fname}: {count} shelf entries')

    print(f'\nTotal shelf entries: {len(shelf_data)}')
    return shelf_data


async def main():
    print('=' * 60)
    print('BACKFILLING SHELF NUMBERS')
    print('=' * 60)
    shelf_data = extract_shelf_numbers()

    print('\n' + '=' * 60)
    print('UPDATING DATABASE')
    print('=' * 60)
    engine = create_async_engine(DATABASE_URL)
    SessionLocal = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with SessionLocal() as db:
        updated = 0
        not_found = 0
        for name, shelf in shelf_data:
            result = await db.execute(select(Product).where(Product.name == name))
            prods = result.scalars().all()
            if prods:
                for p in prods:
                    if p.shelf_number is None or p.shelf_number == '':
                        p.shelf_number = shelf
                updated += 1
            else:
                not_found += 1

        await db.commit()
        print(f'  Updated:   {updated}')
        print(f'  Not found: {not_found}')

    await engine.dispose()
    print('\nDONE')


if __name__ == '__main__':
    asyncio.run(main())
