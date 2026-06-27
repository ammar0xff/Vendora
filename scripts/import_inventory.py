"""
Full import from the 13 رف Excel files:
- Create مخازن > جرد if needed
- Create new products / update existing ones
- Set shelf_number, prices, and opening stock
"""
import asyncio, os, glob
from decimal import Decimal
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine, async_sessionmaker
from sqlalchemy import select, text

import app.models.user, app.models.warehouse
from app.models.product import Product, Category, Subcategory
from app.models.stock import StockMovement

DATABASE_URL = "postgresql+asyncpg://postgres:postgres@db:5432/inventory_db"
BASE = r'/app/LISTS/جرد المخازن'

# (name_col, qty_col, retail_col, wholesale_col, shelf_col, header_rows)
FILE_CONFIG = {
    'رف 1 (2).xlsx':  (1, 2, 3, 4, 0, 2),
    'رف 10.xlsx':     (1, 2, 3, 4, 0, 2),
    'رف 11.xlsx':     (1, 2, 3, None, 0, 2),
    'رف 12.xlsx':     (1, 2, None, None, 0, 0),
    'رف 13.xlsx':     (1, 2, None, None, 0, 0),
    'رف 14.xlsx':     (1, 2, 3, 4, 0, 1),
    'رف 2.xlsx':      (1, 2, 3, 4, 0, 2),
    'رف 3.xlsx':      (1, 2, 3, 4, 0, 2),
    'رف 5.xlsx':      (1, 2, None, None, 0, 1),
    'رف 6.xlsx':      (1, 2, 3, 4, 0, 2),
    'رف 7.xlsx':      (2, 3, 4, 5, 1, 2),
    'رف 8.xlsx':      (1, 2, 3, None, 0, 2),
    'رف 9.xlsx':      (1, 2, 4, 4, 0, 2),
}

SKIP_NAMES = {
    'اسم الصنف', 'الصنف', 'العدد', 'البيان', 'رقم الرف', 'رقم الرف ا',
    'اجمالى', 'الاجمالى', 'المخزن', 'نوع الرف', 'الاجمالي',
    'الوحدة', 'سعر البيع', 'العدد', 'السعر', 'الجملة',
}

def is_noise(val):
    if not val or not isinstance(val, str):
        return True
    v = val.strip()
    if not v or v in SKIP_NAMES:
        return True
    if len(v) <= 2 and not any('\u0600' <= c <= '\u06FF' for c in v):
        return True
    if len(set(v)) <= 2 and len(v) > 5:
        return True
    return False

def clean_shelf(val):
    if val is None:
        return None
    v = str(val).strip()
    if not v or len(v) > 20:
        return None
    return v

def parse_qty(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val) if val > 0 else None
    v = str(val).strip()
    if not v:
        return None
    try:
        q = float(v.replace(',', '.'))
        return int(q) if q > 0 else None
    except:
        return None

def parse_price(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(val)) if val > 0 else None
    v = str(val).strip().replace(',', '')
    if not v:
        return None
    try:
        p = Decimal(v)
        return p if p > 0 else None
    except:
        return None

async def main():
    files = sorted(glob.glob(os.path.join(BASE, '*.xlsx')))
    print(f'Found {len(files)} files\n')

    engine = create_async_engine(DATABASE_URL)
    SessionLocal = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with SessionLocal() as db:
        r = await db.execute(select(Category).where(Category.name == 'مخازن'))
        cat = r.scalar_one_or_none()
        if not cat:
            cat = Category(name='مخازن')
            db.add(cat)
            await db.flush()
            print('Created category: مخازن')
        r = await db.execute(select(Subcategory).where(Subcategory.category_id == cat.id, Subcategory.name == 'جرد'))
        sub = r.scalar_one_or_none()
        if not sub:
            sub = Subcategory(category_id=cat.id, name='جرد')
            db.add(sub)
            await db.flush()
            print('Created subcategory: جرد')
        sub_id = sub.id

        r = await db.execute(text("SELECT id FROM warehouses WHERE name = 'معرض المؤمن'"))
        wh_main = r.scalar()
        r = await db.execute(text("SELECT id FROM warehouses WHERE name = 'مخزن الحديد'"))
        wh_iron = r.scalar()
        print(f'Warehouses: main={wh_main} iron={wh_iron}\n')

        new_p = existing_p = shelf_up = price_up = movements = 0

        for f in files:
            fname = os.path.basename(f)
            cfg = FILE_CONFIG.get(fname)
            if cfg is None:
                print(f'  {fname} — NO CONFIG, skipped')
                continue
            nc, qc, rc, wc, sc, hdr = cfg

            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            file_count = 0

            for row in ws.iter_rows(values_only=True):
                row = list(row)
                if nc >= len(row):
                    continue
                if hdr > 0:
                    hdr -= 1
                    continue

                val = row[nc]
                if is_noise(val):
                    continue
                name = val.strip()
                if len(name) < 2:
                    continue

                qty = parse_qty(row[qc] if qc < len(row) else None)
                retail = parse_price(row[rc] if rc is not None and rc < len(row) else None)
                wholesale = parse_price(row[wc] if wc is not None and wc < len(row) else None)
                shelf = clean_shelf(row[sc] if sc < len(row) else None)

                r = await db.execute(select(Product).where(Product.name == name))
                prods = r.scalars().all()
                prod = prods[0] if prods else None

                if prod:
                    existing_p += 1
                    if shelf:
                        prod.shelf_number = shelf
                        shelf_up += 1
                    if retail and (prod.retail_price is None or prod.retail_price <= 0):
                        prod.retail_price = retail
                        price_up += 1
                    if wholesale and (prod.wholesale_price is None or prod.wholesale_price <= 0):
                        prod.wholesale_price = wholesale
                        if not (retail and (prod.retail_price is None or prod.retail_price <= 0)):
                            price_up += 1
                else:
                    new_p += 1
                    prod = Product(
                        subcategory_id=sub_id,
                        name=name,
                        shelf_number=shelf,
                        unit='عدد',
                        retail_price=retail or Decimal('0'),
                        wholesale_price=wholesale or Decimal('0'),
                        cost_price=Decimal('0'),
                        stock_status='untracked',
                    )
                    db.add(prod)
                    await db.flush()
                    if shelf:
                        shelf_up += 1

                if qty and qty > 0 and prod.stock_status == 'untracked':
                    wh_id = wh_iron if 'رف 7' in fname else wh_main
                    existing_mv = await db.execute(
                        select(StockMovement).where(
                            StockMovement.product_id == prod.id,
                            StockMovement.warehouse_id == wh_id,
                            StockMovement.movement_type == 'opening_stock',
                        )
                    )
                    if not existing_mv.scalar_one_or_none():
                        db.add(StockMovement(
                            product_id=prod.id,
                            warehouse_id=wh_id,
                            movement_type='opening_stock',
                            qty=qty,
                            unit_cost=Decimal('0'),
                            note='جرد',
                        ))
                        movements += 1
                        prod.stock_status = 'tracked'

                file_count += 1

            wb.close()
            print(f'  {fname}: {file_count} products')

        await db.commit()

        print(f'\n{"="*50}')
        print(f'  New products:       {new_p}')
        print(f'  Existing matched:   {existing_p}')
        print(f'  Shelf set/updated:  {shelf_up}')
        print(f'  Prices updated:     {price_up}')
        print(f'  Stock movements:    {movements}')
        print(f'{"="*50}')

    await engine.dispose()

if __name__ == '__main__':
    asyncio.run(main())
