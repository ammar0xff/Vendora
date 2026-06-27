"""
Update retail_price for products from Excel inventory files.
"""
import uuid
import os
import glob
import asyncio
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook

from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine, async_sessionmaker
from sqlalchemy import select

import app.models.user
import app.models.warehouse
from app.models.product import Product

DATABASE_URL = "postgresql+asyncpg://postgres:postgres@db:5432/inventory_db"
BASE = r'/app/LISTS'

# Per-file config: (name_col, [qty_cols], price_col)
# Files with prices:
FILE_CONFIG = {
    0:  (1, [2], None),           # جرد الرفوف تحت الصندلة - no price
    1:  (1, [2], None),           # جرد مخزن البولى - no price
    2:  (2, [3], None),           # جرد المخزن الكبير بولي - no price
    3:  (1, [2], 4),              # جرد رف 4
    4:  (0, [1], None),           # جرد الصندلة - no price
    5:  (1, [2], None),           # جرد مخزن الحديد - no price
    6:  (1, [2], 4),              # رف 1 (2)
    7:  (1, [2], 4),              # رف 10
    8:  (1, [2], 3),              # رف 11 (4 cols: shelf, name, qty, price)
    9:  (1, [2], None),           # رف 12 - no price
    10: (1, [2], None),           # رف 13 - no price
    11: (1, [2], 3),              # رف 14 - has price col 3
    12: (1, [2], 4),              # رف 2 (price in col 4 = الاجمالي)
    13: (1, [2], None),           # رف 3 - no actual price data
    14: (1, [2], None),           # رف 5 - no price
    15: (1, [2], 4),              # رف 6
    16: (2, [3, 5], 5),           # رف 7 (name in col 2, price in col 5)
    17: (1, [2], 3),              # رف 8
    18: (1, [2], 4),              # رف 9
    19: (1, [2, 3], 3),           # مخزن نواكل (price in col 3)
    20: (1, [2], None),           # مخزن داخلى بولى 15
    21: (1, [2], None),           # مخزن داخلي0 - no price
}

SKIP_NAMES = {
    'اسم الصنف', 'الصنف', 'العدد', 'البيان', 'رقم الرف', 'رقم الرف ا',
    'اجمالى', 'الاجمالى', 'المخزن', 'نوع الرف', 'الاجمالي',
    'الوحدة', 'سعر البيع',
}


def is_noise(val):
    if not val:
        return True
    v = val.strip()
    if not v:
        return True
    if v in SKIP_NAMES:
        return True
    if len(v) <= 2 and not any('\u0600' <= c <= '\u06FF' for c in v):
        return True
    if len(set(v)) <= 2 and len(v) > 5:
        return True
    return False


def extract_prices():
    files = sorted(glob.glob(os.path.join(BASE, '*.xlsx')))
    price_data = []  # [(name, price), ...]

    for idx, f in enumerate(files):
        cfg = FILE_CONFIG.get(idx)
        if cfg is None:
            continue
        nc, _, pc = cfg
        if pc is None:
            continue  # No price column

        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active

        count = 0
        for row in ws.iter_rows(values_only=True):
            row = list(row)
            if nc >= len(row) or pc >= len(row):
                continue

            val = row[nc]
            if val is None or not isinstance(val, str):
                continue
            if is_noise(val):
                continue
            name = val.strip()

            # Skip section headers
            if len(name) > 10 and 'مخزن' in name and name.count(' ') > 2:
                continue

            # Get price
            price_val = row[pc]
            price = None
            if price_val is not None:
                try:
                    p = float(str(price_val).replace(',', '').replace(' ', ''))
                    if p > 0:
                        price = Decimal(str(p)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                except:
                    pass

            if price is not None:
                price_data.append((name, price))
                count += 1

        wb.close()
        fname = os.path.basename(f)
        print(f'  [{idx:2d}] {fname}: {count} prices')

    print(f'\nTotal prices extracted: {len(price_data)}')
    return price_data


async def main():
    print('=' * 60)
    print('EXTRACTING PRICES FROM EXCEL')
    print('=' * 60)
    price_data = extract_prices()

    print('\n' + '=' * 60)
    print('UPDATING DATABASE')
    print('=' * 60)
    engine = create_async_engine(DATABASE_URL)
    SessionLocal = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with SessionLocal() as db:
        updated = 0
        not_found = 0
        skipped = 0
        for name, price in price_data:
            result = await db.execute(select(Product).where(Product.name == name))
            prods = result.scalars().all()
            if prods:
                for prod in prods:
                    if prod.retail_price == 0 or prod.retail_price < price:
                        prod.retail_price = price
                updated += 1
            else:
                not_found += 1

        if updated > 0:
            await db.commit()
            print(f'  Updated: {updated}')
            print(f'  Skipped (current ≥ new): {skipped}')
            print(f'  Not found in DB: {not_found}')
        else:
            print('  No updates needed')

    await engine.dispose()
    print('\nDONE')


if __name__ == '__main__':
    asyncio.run(main())
