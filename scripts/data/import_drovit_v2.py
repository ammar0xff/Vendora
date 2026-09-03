import openpyxl, subprocess, json, uuid

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_اسعار_دورافيت_2026_عربي.xlsx', data_only=True)
ws = wb.active

rows = []
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)):
    vals = [str(v).strip() if v is not None else '' for v in row]
    if vals[2] and vals[3]:
        rows.append(vals)

print(f"Excel rows: {len(rows)}")

# Comprehensive series map
SERIES_MAP = {
    'بي 3 كومفورتس': 'P3 COMFORTS',
    'ستارك 3': 'STARCK 3',
    'دي نيو': 'D-NEO',
    'دورا ستايل': 'DURASTYLE',
    'دارلينج نيو': 'DARLING',
    'دارلينج': 'DARLING',
    'هابي دي': 'HAPPY D.',
    'هابي دي 2': 'HAPPY D.',
    'دورافيت نمبر 1': 'DURAVIT NO. 1',
    'دي كود': 'D-CODE',
    'إيكو': 'ECHO',
    'دورابلاس': 'DURAPLUS',
    'إميليا': 'EMILIA',
    'جولف': 'GOLF',
    'ديون': 'DUNE',
    'متنوعة': 'OTHERS',
    'إكسسوارات سيراميك': 'ACCESSORIES CERAMIC',
    'إكسسوارات إيزي': 'ACCESSORIES EASY',
    'إكسسوارات كروم': 'ACCESSORIES CHROME',
    'إكس لارج': 'X-LARGE',
    'فوستر': 'FOSTER',
    'كيثو': 'KETHO',
    'كارو': 'CARO',
    'إل كيوب': 'L-CUBE',
    'فيتريوم': 'VITRIUM',
    'بيورا فيدا': 'PURAVIDA',
    'ستارك 1': 'STARCK 1',
    'فيرو': 'VERO',
}

# Get existing subcategories
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-F', '\t',
     '-c', "SELECT row_to_json(t)::text FROM (SELECT id, name FROM subcategories) t"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
existing_subcats = {}
for line in r.stdout.strip().split('\n'):
    line = line.strip()
    if line.startswith('{'):
        try:
            sc = json.loads(line)
            existing_subcats[sc['name']] = sc['id']
        except:
            pass

print(f"Existing subcats: {len(existing_subcats)}")

# Ensure all needed subcategories exist
cat_ids = {}
for arabic, eng in SERIES_MAP.items():
    if eng in existing_subcats:
        cat_ids[arabic] = existing_subcats[eng]
    else:
        new_id = str(uuid.uuid4())
        safe = eng.replace("'", "''")
        subprocess.run([
            'docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
            '-c', f"INSERT INTO subcategories (id, name) VALUES ('{new_id}', E'{safe}')"
        ], capture_output=True, text=True)
        cat_ids[arabic] = new_id
        existing_subcats[eng] = new_id
        print(f"  Created: {eng}")

# Also ensure OTHERS exists for fallback
if 'OTHERS' not in existing_subcats:
    oid = str(uuid.uuid4())
    subprocess.run([
        'docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
        '-c', f"INSERT INTO subcategories (id, name) VALUES ('{oid}', 'OTHERS')"
    ], capture_output=True, text=True)
    existing_subcats['OTHERS'] = oid
    print("  Created: OTHERS")

# Check for unmapped series
unmapped = set()
for vals in rows:
    series_ar = vals[1]
    if series_ar and series_ar not in SERIES_MAP:
        unmapped.add(series_ar)

if unmapped:
    print(f"\nWARNING: Unmapped series ({len(unmapped)}):")
    for s in sorted(unmapped):
        print(f"  '{s}' -> hex: {[hex(ord(c)) for c in s]}")

# Generate SQL
sql = []
sql.append("DELETE FROM products WHERE company = 'دروفيت';")
unit = 'قطعة'
now = "NOW()"
inserted = 0
skipped = 0

for vals in rows:
    series_ar = vals[1]
    name = vals[2]
    code = vals[3]
    size = vals[4]
    color = vals[5]
    price_str = vals[6]
    
    subcat_id = cat_ids.get(series_ar, existing_subcats.get('OTHERS'))
    if not subcat_id:
        skipped += 1
        continue
    
    price_val = 0
    try:
        price_val = float(price_str.replace(',', ''))
    except:
        pass
    
    prod_id = str(uuid.uuid4())
    safe_name = name.replace("'", "''")
    safe_size = size.replace("'", "''") if size else ''
    safe_color = color.replace("'", "''") if color else ''
    
    sql.append(
        f"INSERT INTO products (id, subcategory_id, name, barcode, unit, "
        f"retail_price, wholesale_price, cost_price, company, size, type, material, "
        f"is_active, created_at, updated_at, stock_status) VALUES ("
        f"'{prod_id}', '{subcat_id}', E'{safe_name}', '{code}', '{unit}', "
        f"{price_val}, {price_val}, {price_val}, 'دروفيت', E'{safe_size}', '', '', "
        f"true, {now}, {now}, 'untracked');"
    )
    inserted += 1

print(f"\nInserting: {inserted}, Skipped: {skipped}")

sql_path = r'C:\eg-co-erp\reimport_drovit_v2.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql))

subprocess.run(['docker', 'cp', sql_path, 'eg-co-erp-db-1:/tmp/reimport_drovit_v2.sql'], capture_output=True)
r2 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db', '-f', '/tmp/reimport_drovit_v2.sql'],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)

error_count = r2.stderr.count('ERROR') if r2.stderr else 0
print(f"Errors: {error_count}")
if r2.stderr and error_count > 0:
    for line in r2.stderr.split('\n'):
        if 'ERROR' in line:
            print(f"  {line[:150]}")

r3 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A',
     '-c', "SELECT count(*) FROM products WHERE company = 'دروفيت'"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(f"دروفيت in DB: {r3.stdout.strip()}")
