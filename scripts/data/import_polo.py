import openpyxl, subprocess, uuid

def clean(val):
    if val is None: return ''
    return str(val).strip()

def fmt_price(val):
    try:
        return f"{float(str(val).replace(',', '')):.2f}"
    except:
        return '0'

def esc(val):
    return val.replace("'", "''")

now = "NOW()"
CAT_NAME = 'بولو'

sql_parts = []

# Create category if not exists, get its ID
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-c', f"SELECT id FROM categories WHERE name = '{CAT_NAME}'"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
cat_id = r.stdout.strip()

if cat_id:
    # Delete old data
    sql_parts.append(f"DELETE FROM products WHERE company = '{CAT_NAME}';")
    sql_parts.append(f"DELETE FROM subcategories WHERE category_id = '{cat_id}';")
else:
    # Create new category - generate UUID in Python to use immediately
    cat_id = str(uuid.uuid4())
    sql_parts.append(
        f"INSERT INTO categories (id, name) VALUES ('{cat_id}', '{CAT_NAME}');"
    )

print(f"Category: {CAT_NAME} -> {cat_id}")

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\Price_List_April_2026.xlsx', data_only=True)
# Skip Summary sheet (index 0)
sheet_names = wb.sheetnames[1:]

total = 0
for sn in sheet_names:
    ws = wb[sn]
    subcat_id = str(uuid.uuid4())
    safe_sn = esc(sn)
    sql_parts.append(
        f"INSERT INTO subcategories (id, category_id, name, created_at) "
        f"VALUES ('{subcat_id}', '{cat_id}', E'{safe_sn}', {now});"
    )
    
    count = 0
    for r in range(4, ws.max_row + 1):
        code_or_letter = clean(ws.cell(r, 1).value)
        name = clean(ws.cell(r, 2).value)
        size = clean(ws.cell(r, 3).value)
        price = ws.cell(r, 4).value
        
        # Skip total row
        if code_or_letter == 'Total Items:' or not name:
            continue
        
        safe_name = esc(name)
        safe_size = esc(size)
        pv = fmt_price(price)
        prod_id = str(uuid.uuid4())
        
        sql_parts.append(
            f"INSERT INTO products (id, subcategory_id, name, barcode, unit, "
            f"retail_price, wholesale_price, cost_price, company, size, type, material, "
            f"is_active, created_at, updated_at, stock_status) VALUES ("
            f"'{prod_id}', '{subcat_id}', E'{safe_name}', '', 'قطعة', "
            f"{pv}, {pv}, {pv}, '{CAT_NAME}', E'{safe_size}', '', '', "
            f"true, {now}, {now}, 'untracked');"
        )
        count += 1
        total += 1
    
    print(f"  {sn}: {count} products")

print(f"Total: {total}")

sql_path = r'C:\eg-co-erp\import_polo.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_parts))

subprocess.run(['docker', 'cp', sql_path, 'eg-co-erp-db-1:/tmp/import_polo.sql'], capture_output=True)
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-f', '/tmp/import_polo.sql'],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
errors = r.stderr.count('ERROR') if r.stderr else 0
print(f"Errors: {errors}")
if r.stderr and errors > 0:
    for l in r.stderr.split('\n')[:15]:
        if 'ERROR' in l:
            print(f"  {l[:200]}")

# Verify
r2 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-c', f"SELECT company, count(*) FROM products WHERE company = '{CAT_NAME}' GROUP BY company;"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(r2.stdout)

r3 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-c', f"SELECT s.name, count(p.id) FROM subcategories s LEFT JOIN products p ON p.subcategory_id = s.id WHERE s.category_id = '{cat_id}' GROUP BY s.name ORDER BY s.name;"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(r3.stdout)
