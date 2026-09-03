import openpyxl

out = open(r'C:\eg-co-erp\ideal_excel_structure.txt', 'w', encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\Ideal_Standard_2026_Price_List.xlsx', data_only=True)
out.write(f"Sheets ({len(wb.sheetnames)}):\n")
for i, sn in enumerate(wb.sheetnames):
    out.write(f"  {i}: {repr(sn)}\n")

out.write("\n" + "="*60 + "\n")

# Explore each sheet
for sn in wb.sheetnames:
    ws = wb[sn]
    out.write(f"\n=== Sheet: {repr(sn)} (rows={ws.max_row}, cols={ws.max_column}) ===\n")
    
    # Print rows until we find the data
    for r in range(1, min(ws.max_row + 1, 50)):
        vals = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                vals.append(f"{c}:{repr(str(cell.value)[:60])}")
        if vals:
            out.write(f"  Row {r}: {vals}\n")

out.close()
print("Done")
