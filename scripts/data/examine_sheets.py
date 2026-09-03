import openpyxl, sys

out = open(r'C:\eg-co-erp\sheets_out.txt', 'w', encoding='utf-8')
wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_اسعار_دورافيت_2026_عربي.xlsx', data_only=True)
for i, sn in enumerate(wb.sheetnames):
    ws = wb[sn]
    out.write(f"Sheet {i}: {repr(sn)} (rows={ws.max_row}, cols={ws.max_column})\n")
    # Print header row
    header = []
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        val = cell.value
        header.append(repr(val)[:60] if val is not None else 'None')
    out.write(f"  Header: {header}\n")
    # Print first data row
    if ws.max_row >= 2:
        row2 = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=c)
            val = cell.value
            row2.append(f'{c}:{type(val).__name__}={repr(val)[:80]}' if val is not None else f'{c}:None')
    out.write(f"  Row 2: {row2}\n")
out.close()
