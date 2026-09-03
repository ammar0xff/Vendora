import openpyxl, subprocess, re

# Read Excel
wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_أسعار_شيلد_يناير_2026.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]

# Build list of (subcat, prod, size, price_str)
excel_rows = []
for r in range(6, ws.max_row + 1):
    subcat = str(ws.cell(r, 1).value or '').strip()
    prod = str(ws.cell(r, 2).value or '').strip()
    size = str(ws.cell(r, 3).value or '').strip()
    price = ws.cell(r, 5).value
    if not prod:
        continue
    price_str = str(int(price)) if price is not None and price == int(price) else str(price) if price else ''
    excel_rows.append((subcat, prod, size, price_str))

# Get all products from DB
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-F||', '-c',
     "SELECT p.id, p.retail_price::text, trim(s.name) FROM products p JOIN subcategories s ON p.subcategory_id = s.id WHERE p.company = 'شيلد' ORDER BY p.id"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)

updates = []
not_found = []
for line in [l for l in r.stdout.strip().split('\n') if l.strip()]:
    parts = line.split('||')
    if len(parts) < 3:
        continue
    prod_id = parts[0].strip()
    db_price = parts[1].strip().rstrip('.0')
    db_subcat = parts[2].strip()
    
    # Find matching Excel row by (subcat, price)
    matched = False
    for subcat, prod, size, price_str in excel_rows:
        if subcat == db_subcat and price_str == db_price:
            # Build clean name
            clean = f'{prod} - {size}' if size and size != '-' else prod
            clean = clean.replace("'", "''")
            updates.append(f"UPDATE products SET name = E'{clean}' WHERE id = '{prod_id}';")
            matched = True
            break
    
    if not matched:
        # Try matching by price only (some subcats may differ in encoding)
        for subcat, prod, size, price_str in excel_rows:
            if price_str == db_price:
                clean = f'{prod} - {size}' if size and size != '-' else prod
                clean = clean.replace("'", "''")
                updates.append(f"UPDATE products SET name = E'{clean}' WHERE id = '{prod_id}';")
                matched = True
                break
    
    if not matched:
        not_found.append(f'{prod_id[:8]} | price={db_price} | subcat={db_subcat}')

print(f'Generated {len(updates)} updates')
if not_found:
    print(f'Not matched ({len(not_found)}):')
    for nf in not_found[:5]:
        print(f'  {nf}')

sql_path = r'C:\eg-co-erp\rebuild_shield_v2.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(updates))

subprocess.run(['docker', 'cp', sql_path, 'eg-co-erp-db-1:/tmp/rebuild_shield_v2.sql'], capture_output=True)
r2 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-f', '/tmp/rebuild_shield_v2.sql'],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
errors = r2.stderr.count('ERROR') if r2.stderr else 0
print(f'Errors: {errors}')

# Check results
r3 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-c', "SELECT name FROM products WHERE company = 'شيلد' ORDER BY name;"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(r3.stdout)

# Check duplicates
r4 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-c', "SELECT count(*) FROM (SELECT name FROM products WHERE company = 'شيلد' GROUP BY name HAVING count(*) > 1) d;"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(f'Duplicates: {r4.stdout.strip()}')
