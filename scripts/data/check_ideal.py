import openpyxl
wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\Ideal_Standard_2026_Price_List.xlsx', data_only=True)
print('Sheets:', wb.sheetnames[:5], '...' if len(wb.sheetnames) > 5 else '')
ws = wb[wb.sheetnames[0]]
print(f'{wb.sheetnames[0]}: rows={ws.max_row}, cols={ws.max_column}')
for r in range(1, 6):
    vals = [str(ws.cell(r, c).value or '')[:40] for c in range(1, min(ws.max_column + 1, 10))]
    print(f'Row {r}: {" | ".join(vals)}')
