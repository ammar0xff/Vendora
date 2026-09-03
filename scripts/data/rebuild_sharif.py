import openpyxl, subprocess
from collections import defaultdict

# === الشريف ===
print('=== Processing الشريف ===')

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\done\الشريف.xlsx', data_only=True)
ws = wb['All_Products']

# Build Excel map: (price_str, subcat_hex) -> list of items
excel_map = defaultdict(list)
for r_idx in range(2, ws.max_row + 1):
    subcat = str(ws.cell(r_idx, 1).value or '').strip()
    naw = str(ws.cell(r_idx, 2).value or '').strip()
    item = str(ws.cell(r_idx, 3).value or '').strip()
    size = str(ws.cell(r_idx, 4).value or '').strip()
    extra = str(ws.cell(r_idx, 5).value or '').strip()
    price = ws.cell(r_idx, 6).value
    if price is None or not item:
        continue
    p = float(str(price))
    price_str = str(int(p)) if p == int(p) else f'{p:.2f}'
    key = (price_str, subcat.encode('utf-8').hex())
    excel_map[key].append({'naw': naw, 'item': item, 'size': size, 'extra': extra, 'subcat': subcat})

print(f'Excel rows: {ws.max_row - 1}')
print(f'Unique keys: {len(excel_map)}')

# Get DB products - query subcat name as-is from DB
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-F||', '-c',
     "SELECT p.id, p.retail_price::text, encode(convert_to(s.name, 'UTF8'), 'hex') "
     "FROM products p JOIN subcategories s ON p.subcategory_id = s.id "
     "WHERE p.company = 'الشريف' ORDER BY s.name, p.retail_price, p.id"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)

db_lines = [l for l in r.stdout.strip().split('\n') if l.strip()]
print(f'DB products: {len(db_lines)}')

key_counters = defaultdict(int)
updates = []
unmatched = []

for line in db_lines:
    parts = line.split('||')
    if len(parts) < 3: continue
    prod_id = parts[0].strip()
    raw_price = parts[1].strip()
    subcat_hex = parts[2].strip()
    
    # Normalize price same way as Excel
    p = float(raw_price)
    price_str = str(int(p)) if p == int(p) else f'{p:.2f}'
    
    key = (price_str, subcat_hex)
    items = excel_map.get(key, [])
    idx = key_counters[key]
    
    if idx < len(items):
        row = items[idx]
        key_counters[key] += 1
        
        naw = row['naw']
        item = row['item']
        size = row['size']
        extra = row['extra']
        subcat = row['subcat']
        
        # Build name
        if size:
            if naw and naw != item:
                name = f'{naw} {item} - {size}'
            else:
                name = f'{item} - {size}'
        else:
            if naw and naw != item:
                name = f'{naw} {item}'
            else:
                name = item
        
        if extra and extra != '-':
            name += f' ({extra})'
        
        updates.append(f"UPDATE products SET name = E'{name.replace(chr(39), chr(39)+chr(39))}' WHERE id = '{prod_id}';")
    else:
        unmatched.append(f'{prod_id[:12]} price={price_str} key_count={len(items)}')

if unmatched:
    print(f'Unmatched ({len(unmatched)}):')
    for u in unmatched[:5]:
        print(f'  {u}')

print(f'Generated {len(updates)} updates')

if updates:
    sql_path = r'C:\eg-co-erp\rebuild_sharif.sql'
    with open(sql_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(updates))
    subprocess.run(['docker', 'cp', sql_path, 'eg-co-erp-db-1:/tmp/rebuild_sharif.sql'], capture_output=True)
    r2 = subprocess.run(
        ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
         '-f', '/tmp/rebuild_sharif.sql'],
        capture_output=True, text=True, encoding='utf-8', errors='replace'
    )
    errors = r2.stderr.count('ERROR') if r2.stderr else 0
    if errors:
        lines = r2.stderr.split('\n')
        for l in lines[:5]:
            if 'ERROR' in l:
                print(f'  ERROR: {l[:200]}')
    print(f'Apply errors: {errors}')

# Verify
r3 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-c', "SELECT name FROM products WHERE company = 'الشريف' ORDER BY name LIMIT 30;"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(r3.stdout)

r4 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-c', "SELECT count(*) FROM (SELECT name FROM products WHERE company = 'الشريف' GROUP BY name HAVING count(*) > 1) d;"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(f'Duplicate names: {r4.stdout.strip()}')
