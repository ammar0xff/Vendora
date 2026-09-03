import openpyxl, subprocess, uuid

def clean(val):
    if val is None: return ''
    return str(val).strip()

def fmt_price(val):
    try:
        return f"{float(str(val).replace(',', '')):.2f}"
    except:
        return '0'

def esc(val):
    return val.replace("'", "''")

now = "NOW()"
COMER_CAT_ID = '53b01a68-011b-4b04-9265-aa843a4ffee4'
sql_parts = []

# Delete old كومر
sql_parts.append("DELETE FROM products WHERE company = 'كومر';")
sql_parts.append(f"DELETE FROM subcategories WHERE category_id = '{COMER_CAT_ID}';")

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\done\Neisco_Comer_Price_List_2026.xlsx', data_only=True)

comer_count = 0
for sn in wb.sheetnames:
    ws = wb[sn]
    subcat_id = str(uuid.uuid4())
    # Use sheet name directly (no trimming)
    safe_sn = esc(sn)
    sql_parts.append(
        f"INSERT INTO subcategories (id, category_id, name, created_at) "
        f"VALUES ('{subcat_id}', '{COMER_CAT_ID}', E'{safe_sn}', {now});"
    )
    for r in range(4, ws.max_row + 1):
        code = clean(ws.cell(r, 1).value)
        size = clean(ws.cell(r, 2).value)
        price = ws.cell(r, 5).value
        if not code:
            continue
        safe_code = esc(code)
        safe_size = esc(size)
        pv = fmt_price(price)
        prod_id = str(uuid.uuid4())
        prod_name = f"{safe_code} {safe_size}" if safe_size else safe_code
        sql_parts.append(
            f"INSERT INTO products (id, subcategory_id, name, barcode, unit, "
            f"retail_price, wholesale_price, cost_price, company, size, type, material, "
            f"is_active, created_at, updated_at, stock_status) VALUES ("
            f"'{prod_id}', '{subcat_id}', E'{prod_name}', '{safe_code}', 'قطعة', "
            f"{pv}, {pv}, {pv}, 'كومر', E'{safe_size}', '', '', "
            f"true, {now}, {now}, 'untracked');"
        )
        comer_count += 1

print(f"Generated {comer_count} products for كومر")

sql_path = r'C:\eg-co-erp\fix_comer_import.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_parts))

subprocess.run(['docker', 'cp', sql_path, 'eg-co-erp-db-1:/tmp/fix_comer_import.sql'], capture_output=True)
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-f', '/tmp/fix_comer_import.sql'],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
errors = r.stderr.count('ERROR') if r.stderr else 0
print(f"Errors: {errors}")
if r.stderr and errors > 0:
    for l in r.stderr.split('\n')[:15]:
        if 'ERROR' in l:
            print(f"  {l[:200]}")

print("\n=== Verification ===")
r2 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-c', "SELECT s.name, count(p.id) as prods FROM categories c JOIN subcategories s ON s.category_id = c.id LEFT JOIN products p ON p.subcategory_id = s.id WHERE c.name = 'كومر' GROUP BY s.name ORDER BY s.name"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(r2.stdout)
