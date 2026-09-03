import openpyxl, subprocess

fp = r'C:\eg-co-erp\LISTS\ادخال اسعار.xlsx'
wb = openpyxl.load_workbook(fp)
print('Sheets:', wb.sheetnames)

for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'\n=== {sn} === rows={ws.max_row} cols={ws.max_column}')
    for r in range(1, ws.max_row + 1):
        vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            vals.append(f'{c}:{repr(v)[:50]}')
        print(f'  Row {r}: {", ".join(vals)}')
