import openpyxl, subprocess

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_أسعار_شيلد_يناير_2026.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]

# Read all Excel data into list of dicts
excel_rows = []
for r in range(6, ws.max_row + 1):
    subcat = str(ws.cell(r, 1).value or '').strip()
    prod = str(ws.cell(r, 2).value or '').strip()
    size = str(ws.cell(r, 3).value or '').strip()
    price = str(int(ws.cell(r, 5).value)) if ws.cell(r, 5).value is not None and ws.cell(r, 5).value == int(ws.cell(r, 5).value) else str(ws.cell(r, 5).value) if ws.cell(r, 5).value else ''
    pack = ws.cell(r, 4).value or ''
    excel_rows.append({'subcat': subcat, 'prod': prod, 'size': size, 'price': price, 'pack': str(pack).strip()})

# Print all unique entries for debugging
from collections import defaultdict
by_subcat = defaultdict(list)
for row in excel_rows:
    by_subcat[row['subcat']].append(row)

for subcat, rows in sorted(by_subcat.items()):
    print(f'=== {repr(subcat)} ({len(rows)} items) ===')
    for row in rows:
        clean_name = f"{row['prod']} - {row['size']}" if row['size'] and row['size'] != '-' else row['prod']
        print(f"  price={row['price']:>5s}  pack={row['pack']:>8s}  name={clean_name}")
