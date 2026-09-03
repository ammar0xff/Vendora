import openpyxl, subprocess

# Re-read شيلد data
wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_أسعار_شيلد_يناير_2026.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]

# Build name -> size map from the Excel data
from collections import defaultdict
name_map = defaultdict(list)
for r in range(6, ws.max_row + 1):
    cat = str(ws.cell(r, 1).value or '').strip()
    prod = str(ws.cell(r, 2).value or '').strip()
    size = str(ws.cell(r, 3).value or '').strip()
    price = ws.cell(r, 5).value
    if not prod:
        continue
    # Build clean name: product type + size
    if size and size != '-':
        clean_name = f'{prod} - {size}'
    else:
        clean_name = prod
    name_map[clean_name].append((cat, price))

# Get current products from DB
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-F|', '-c',
     "SELECT p.id, p.name, p.retail_price, s.name FROM products p JOIN subcategories s ON p.subcategory_id = s.id WHERE p.company = 'شيلد' ORDER BY p.name"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)

# Build new names based on Excel data
lines = r.stdout.strip().split('\n')
updates = []
for line in lines:
    if not line.strip():
        continue
    parts = line.split('|')
    if len(parts) < 3:
        continue
    prod_id = parts[0].strip()
    old_name = parts[1].strip()
    db_price = parts[2].strip()
    db_subcat = parts[3].strip() if len(parts) > 3 else ''
    
    # Find matching entry in name_map by price (most reliable way to match)
    found = False
    for clean_name, entries in name_map.items():
        for entry_cat, entry_price in entries:
            if entry_price is not None and str(entry_price) == db_price:
                new_name = clean_name.replace("'", "''")
                updates.append(f"UPDATE products SET name = E'{new_name}' WHERE id = '{prod_id}';")
                found = True
                break
        if found:
            break
    
    if not found:
        # Fallback: strip parenthetical content that looks like specs
        # Keep short aliases like (معالج UV), (جديد) but remove long specs
        import re
        cleaned = re.sub(r'\s*\([^)]*نحاس[^)]*\)', '', old_name)
        cleaned = re.sub(r'\s*\([^)]*جسم[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*جوان[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*قلب[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*يد[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*بوصلة[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*كونتر[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*بلية[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*مقبض[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*سطح[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*صامولة[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*عمود[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*جشمه[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*زنك[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*لاكور[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*بلاستيك[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*لوجو[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*ستانلس[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*مسمار[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*سوسته[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*باب[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*طبق[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*فتيل[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*نيكل[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*معدني[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*PVC[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*PTFE[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*NBR[^)]*\)', '', cleaned)
        cleaned = re.sub(r'\s*\([^)]*PPR[^)]*\)', '', cleaned)
        # Also handle nested parens: strip remaining long paren content
        cleaned = re.sub(r'\s*\([^)]{20,}\)', '', cleaned)
        # Clean up
        cleaned = re.sub(r'\s+', ' ', cleaned).strip()
        cleaned = cleaned.replace("'", "''")
        if cleaned and cleaned != old_name:
            updates.append(f"UPDATE products SET name = E'{cleaned}' WHERE id = '{prod_id}';")

print(f'Generated {len(updates)} updates')

sql_path = r'C:\eg-co-erp\strip_shield_specs.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(updates))

subprocess.run(['docker', 'cp', sql_path, 'eg-co-erp-db-1:/tmp/strip_shield_specs.sql'], capture_output=True)
r2 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-f', '/tmp/strip_shield_specs.sql'],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
errors = r2.stderr.count('ERROR') if r2.stderr else 0
print(f'Errors: {errors}')
if errors > 0:
    for l in r2.stderr.split('\n')[:10]:
        if 'ERROR' in l:
            print(f'  {l[:200]}')

# Verify
r3 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-c', "SELECT name FROM products WHERE company = 'شيلد' ORDER BY name;"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(r3.stdout)
