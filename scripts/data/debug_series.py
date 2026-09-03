import openpyxl

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_اسعار_دورافيت_2026_عربي.xlsx', data_only=True)
ws = wb.active

# Get unique series names and their Arabic names properly
seen = {}
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)):
    vals = [str(v).strip() if v is not None else '' for v in row]
    if vals[1] and vals[1] not in seen:
        # Print the actual characters
        s = vals[1]
        codes = [hex(ord(c)) for c in s]
        seen[s] = codes

for s, codes in seen.items():
    print(f"'{s}' -> {codes}")
