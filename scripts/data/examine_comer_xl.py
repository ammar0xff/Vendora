import openpyxl

out = open(r'C:\eg-co-erp\comer_xl_structure.txt', 'w', encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\done\Neisco_Comer_Price_List_2026.xlsx', data_only=True)
out.write(f"Sheets ({len(wb.sheetnames)}):\n")
for i, sn in enumerate(wb.sheetnames):
    out.write(f"  {i}: {repr(sn)}\n")

out.write("\n" + "="*60 + "\n")

for sn in wb.sheetnames:
    ws = wb[sn]
    out.write(f"\n=== {repr(sn)} (rows={ws.max_row}, cols={ws.max_column}) ===\n")
    for r in range(1, min(ws.max_row + 1, 25)):
        vals = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                vals.append(f"{c}:{repr(str(cell.value)[:60])}")
        if vals:
            out.write(f"  Row {r}: {vals}\n")

out.close()
print("Done")
