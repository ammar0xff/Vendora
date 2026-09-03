import openpyxl, subprocess, re

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_أسعار_شيلد_يناير_2026.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]

# Read Excel rows in order - each gets a sequence number per subcategory
excel_rows = []
for r in range(6, ws.max_row + 1):
    subcat = str(ws.cell(r, 1).value or '').strip()
    prod = str(ws.cell(r, 2).value or '').strip()
    size = str(ws.cell(r, 3).value or '').strip()
    price = ws.cell(r, 5).value
    specs = str(ws.cell(r, 6).value or '').strip()
    price_str = str(int(price)) if isinstance(price, (int, float)) and price == int(price) else str(price or '')
    excel_rows.append({'subcat': subcat, 'prod': prod, 'size': size, 'price': price_str, 'specs': specs})

# Get DB products in order (by subcategory, then by retail_price, then by id/uuid order)
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-F||', '-c',
     "SELECT p.id, p.name, p.retail_price::text, trim(s.name) FROM products p JOIN subcategories s ON p.subcategory_id = s.id WHERE p.company = 'شيلد' ORDER BY s.name, p.retail_price, p.id"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)

db_products = []
for line in [l for l in r.stdout.strip().split('\n') if l.strip()]:
    parts = line.split('||')
    if len(parts) < 4: continue
    db_products.append({'id': parts[0].strip(), 'name': parts[1].strip(), 'price': parts[2].strip().rstrip('.0'), 'subcat': parts[3].strip()})

# Match by position within (subcat, price) groups
from collections import defaultdict

# Group both by (subcat, price)
excel_groups = defaultdict(list)
for row in excel_rows:
    key = (row['subcat'], row['price'])
    excel_groups[key].append(row)

db_groups = defaultdict(list)
for row in db_products:
    key = (row['subcat'], row['price'])
    db_groups[key].append(row)

# Build name using spec extraction
generic = {'جسم نحاس', 'جسم نحاس (لوجو حفر)', 'جسم نحاس (متعدد الاستخدام)', 'سطح عالي الجودة', 'سطح خارجي عالي الجودة',
           'سطح عالي الجودة - طلاء نيكل كروم', 'طلاء نيكل كروم', 'جوان NBR عالي الجودة', 'جوان NBR'}

def extract_key_materials(specs):
    if not specs or specs == '-':
        return ''
    parts = re.split(r'[،,]', specs)
    parts = [p.strip() for p in parts]
    meaningful = [p for p in parts if p not in generic]
    if meaningful:
        return ' - '.join(meaningful[:2])
    return ' - '.join(parts[:2])

def build_name(subcat, prod, size, key_mat):
    if size and size != '-':
        if key_mat:
            return f'{subcat} {prod} ({key_mat}) - {size}'
        return f'{subcat} {prod} - {size}'
    else:
        if key_mat:
            return f'{subcat} {prod} ({key_mat})'
        return f'{subcat} {prod}'

# Match and generate updates
updates = []
unmatched = []
for key, excel_list in excel_groups.items():
    db_list = db_groups.get(key, [])
    
    for i, excel_row in enumerate(excel_list):
        if i < len(db_list):
            db_row = db_list[i]
            key_mat = extract_key_materials(excel_row['specs'])
            new_name = build_name(excel_row['subcat'], excel_row['prod'], excel_row['size'], key_mat)
            # Escape single quotes
            new_name_esc = new_name.replace("'", "''")
            updates.append(f"UPDATE products SET name = E'{new_name_esc}' WHERE id = '{db_row['id']}';")
        else:
            unmatched.append(f"Excel row {i}: {excel_row['subcat']} / {excel_row['prod']} / {excel_row['price']}")

print(f'Generated {len(updates)} updates')
if unmatched:
    print(f'Unmatched Excel rows: {len(unmatched)}')
    for u in unmatched[:3]:
        print(f'  {u}')

# Check for extra DB products
for key, db_list in db_groups.items():
    excel_list = excel_groups.get(key, [])
    if len(db_list) > len(excel_list):
        print(f'WARNING: {key} has {len(db_list)} DB products but only {len(excel_list)} Excel rows')
        for extra in db_list[len(excel_list):]:
            print(f'  Extra: {extra["id"][:12]} ... {extra["name"]}')

# Apply updates
sql = '\n'.join(updates)
sql_path = r'C:\eg-co-erp\rebuild_shield_descriptive.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write(sql)

subprocess.run(['docker', 'cp', sql_path, 'eg-co-erp-db-1:/tmp/rebuild_shield_descriptive.sql'], capture_output=True)
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-f', '/tmp/rebuild_shield_descriptive.sql'],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
errors = r.stderr.count('ERROR') if r.stderr else 0
print(f'Apply errors: {errors}')
