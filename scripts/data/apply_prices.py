import openpyxl, subprocess

fp = r'C:\eg-co-erp\LISTS\ادخال اسعار.xlsx'

wb = openpyxl.load_workbook(fp)

all_updates = []
processed = 0

for sn in wb.sheetnames:
    ws = wb[sn]
    if ws.max_row < 2:
        continue
    
    rows_del = []
    for r in range(ws.max_row, 1, -1):
        price = ws.cell(r, 5).value
        pid = ws.cell(r, 6).value
        if price is not None and pid:
            try:
                pv = float(str(price))
            except (ValueError, TypeError):
                continue
            all_updates.append(f"UPDATE products SET retail_price = {pv} WHERE id = '{str(pid).strip()}';")
            rows_del.append(r)
            processed += 1
    for r in sorted(rows_del, reverse=True):
        ws.delete_rows(r)

wb.save(fp)
print(f'Processed {processed} price updates')

if not all_updates:
    exit(0)

# Apply
sql = 'BEGIN;\n' + '\n'.join(all_updates) + '\nCOMMIT;'
with open(r'C:\eg-co-erp\update_prices_entry.sql', 'w', encoding='utf-8') as f:
    f.write(sql)

subprocess.run(['docker', 'cp', r'C:\eg-co-erp\update_prices_entry.sql', 'eg-co-erp-db-1:/tmp/update_prices_entry.sql'], capture_output=True)
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db', '-f', '/tmp/update_prices_entry.sql'],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(r.stdout[:500] if r.stdout else '')
if r.stderr and 'ERROR' in r.stderr:
    for l in r.stderr.split('\n'):
        if 'ERROR' in l:
            print(l[:200])

# Count products with prices
r2 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-c', "SELECT count(*) FROM products WHERE retail_price > 0;"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(f'Products with price > 0: {r2.stdout.strip()}')
