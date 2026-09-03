import openpyxl

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\done\الشريف.xlsx', data_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'=== {sn} === rows={ws.max_row} cols={ws.max_column}')
    if sn == 'All_Products':
        for r in range(1, min(ws.max_row + 1, 6)):
            vals = [str(ws.cell(r, c).value or '')[:50] for c in range(1, min(ws.max_column + 1, 8))]
            print(f'  Row {r}: {" | ".join(vals)}')
        print('  ...')
        # Sample data rows  
        for r in [50, 100, 150, 200, 300, 400, 500, 600]:
            if r <= ws.max_row:
                vals = [str(ws.cell(r, c).value or '')[:50] for c in range(1, min(ws.max_column + 1, 8))]
                print(f'  Row {r}: {" | ".join(vals)}')
