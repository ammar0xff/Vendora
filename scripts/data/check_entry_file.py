import openpyxl
import os

path = r'C:\eg-co-erp\LISTS'
files = [f for f in os.listdir(path) if 'ادخال' in f or 'سعار' in f]
print(f'Found: {files}')

for fn in files:
    fp = os.path.join(path, fn)
    print(f'\n=== {fn} ===')
    print(f'Full path: {fp}')
    wb = openpyxl.load_workbook(fp, data_only=True)
    print(f'Sheets: {wb.sheetnames}')
    ws = wb[wb.sheetnames[0]]
    print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
    for r in range(1, min(ws.max_row + 1, 6)):
        vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            vals.append(str(v or '')[:40])
        print(f'  Row {r}: {" | ".join(vals)}')
    # Random data rows
    for r in range(10, min(ws.max_row + 1, 16)):
        vals = []
        for c in range(1, min(ws.max_column + 1, 7)):
            v = ws.cell(r, c).value
            vals.append(str(v or '')[:40])
        print(f'  Row {r}: {" | ".join(vals)}')
    print(f'  ... total {ws.max_row} rows')
