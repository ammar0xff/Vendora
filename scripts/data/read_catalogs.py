import openpyxl, json, sys

sys.stdout.reconfigure(encoding='utf-8')

# ========== الشريف.xlsx ==========
path = r'C:\eg-co-erp\LISTS\الشريف.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)

print('=== الشريف.xlsx ===')
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'\nSheet: {sn} ({ws.max_row} rows x {ws.max_column} cols)')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
        print(f'  {i}: {list(row)}')

ws = wb['All_Products']
types = set()
names = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1]: types.add(str(row[1]).strip())
print(f'\nUnique Types ({len(types)}):')
for t in sorted(types):
    print(f'  - [{t}]')

# Sample 5 products per type
from collections import defaultdict
samples = defaultdict(list)
for row in ws.iter_rows(min_row=2, values_only=True):
    t = str(row[1]).strip() if row[1] else ''
    if len(samples[t]) < 3:
        samples[t].append(list(row))
for t, rows in sorted(samples.items()):
    print(f'\n--- {t} (sample) ---')
    for r in rows:
        print(f'    {r}')

# ========== ايجيك.xlsx ==========
path2 = r'C:\eg-co-erp\LISTS\ايجيك.xlsx'
wb2 = openpyxl.load_workbook(path2, data_only=True)
ws2 = wb2['Sheet1']

print('\n\n=== ايجيك.xlsx ===')
print(f'Sheet1: {ws2.max_row} rows x {ws2.max_column} cols')
print(f'Headers: {[cell.value for cell in next(ws2.iter_rows(min_row=1, max_row=1))]}')

# Print first 15 rows
for i, row in enumerate(ws2.iter_rows(min_row=2, max_row=16, values_only=True), 2):
    print(f'  {i}: {list(row)}')

# Unique brands
brands = set()
materials = set()
for row in ws2.iter_rows(min_row=2, values_only=True):
    if row[2]: brands.add(str(row[2]).strip())
    if row[1]: 
        m = str(row[1]).strip()
        # Extract prefix/type before comma or first word
        materials.add(m)
print(f'\nUnique Brands ({len(brands)}):')
for b in sorted(brands):
    print(f'  - [{b}]')

print(f'\nTotal unique materials: {len(materials)}')
# Print some samples
for i, m in enumerate(sorted(materials)[:30]):
    print(f'  {i+1}. {m}')
