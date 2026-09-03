import openpyxl, subprocess, json, uuid

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_اسعار_دورافيت_2026_عربي.xlsx', data_only=True)
ws = wb.active

rows = []
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)):
    vals = [str(v).strip() if v is not None else '' for v in row]
    if vals[2] and vals[3]:
        rows.append(vals)

print(f"Excel rows: {len(rows)}")

SERIES_MAP = {
    'بي 3 كومفورتس': 'P3 COMFORTS',
    'ستارك 3': 'STARCK 3',
    'دي نيو': 'D-NEO',
    'دورا ستايل': 'DURASTYLE',
    'دارلينج نيو': 'DARLING',
    'دارلينج': 'DARLING',
    'هابي دي': 'HAPPY D.',
    'هابي دي 2': 'HAPPY D.',
    'دورافيت نمبر 1': 'DURAVIT NO. 1',
    'دي كود': 'D-CODE',
    'إيكو': 'ECHO',
    'دورابلاس': 'DURAPLUS',
    'إميليا': 'EMILIA',
    'جولف': 'GOLF',
    'ديون': 'DUNE',
    'متنوعة': 'OTHERS',
    'إكسسوارات سيراميك': 'ACCESSORIES CERAMIC',
    'إكسسوارات إيزي': 'ACCESSORIES EASY',
    'إكسسوارات كروم': 'ACCESSORIES CHROME',
    'إكس لارج': 'X-LARGE',
    'فوستر': 'FOSTER',
    'كيثو': 'KETHO',
    'كارو': 'CARO',
    'إل كيوب': 'L-CUBE',
    'فيتريوم': 'VITRIUM',
    'بيورا فيدا': 'PURAVIDA',
    'ستارك 1': 'STARCK 1',
    'فيرو': 'VERO',
    'الطابق الثاني': 'SECOND FLOOR',
    'عمود الطابق الثاني': 'SECOND FLOOR COLUMN',
    'مجموعة تثبيت': 'MOUNTING SET',
}

# Get existing subcategories
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-F', '\t',
     '-c', "SELECT row_to_json(t)::text FROM (SELECT id, name FROM subcategories) t"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
existing_subcats = {}
for line in r.stdout.strip().split('\n'):
    line = line.strip()
    if line.startswith('{'):
        try:
            sc = json.loads(line)
            existing_subcats[sc['name']] = sc['id']
        except:
            pass

print(f"Existing subcats: {len(existing_subcats)}")

# Create ALL needed subcategories first
cat_arabic_to_uuid = {}
sql_create = []
sql_create.append("DELETE FROM products WHERE company = 'دروفيت';")

for arabic, eng in SERIES_MAP.items():
    if eng in existing_subcats:
        cat_arabic_to_uuid[arabic] = existing_subcats[eng]
    else:
        new_id = str(uuid.uuid4())
        safe = eng.replace("'", "''")
        sql_create.append(f"INSERT INTO subcategories (id, name) VALUES ('{new_id}', E'{safe}');")
        cat_arabic_to_uuid[arabic] = new_id
        existing_subcats[eng] = new_id

# Get OTHERS fallback
fallback_id = existing_subcats.get('OTHERS')
if not fallback_id:
    new_id = str(uuid.uuid4())
    sql_create.append(f"INSERT INTO subcategories (id, name) VALUES ('{new_id}', 'OTHERS');")
    fallback_id = new_id

print(f"Subcategories to create: {len(sql_create)-1}")

# Run the subcategory creation SQL
create_sql = '\n'.join(sql_create)
tmp_path = r'C:\eg-co-erp\drovit_setup.sql'
with open(tmp_path, 'w', encoding='utf-8') as f:
    f.write(create_sql)

subprocess.run(['docker', 'cp', tmp_path, 'eg-co-erp-db-1:/tmp/drovit_setup.sql'], capture_output=True)
r1 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db', '-f', '/tmp/drovit_setup.sql'],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
errors = r1.stderr.count('ERROR') if r1.stderr else 0
print(f"Setup errors: {errors}")
if r1.stderr and errors > 0:
    for l in r1.stderr.split('\n')[:5]:
        if 'ERROR' in l:
            print(f"  {l[:150]}")

# Re-fetch to get authoritative IDs
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-F', '\t',
     '-c', "SELECT row_to_json(t)::text FROM (SELECT id, name FROM subcategories) t"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
existing_subcats = {}
for line in r.stdout.strip().split('\n'):
    line = line.strip()
    if line.startswith('{'):
        try:
            sc = json.loads(line)
            existing_subcats[sc['name']] = sc['id']
        except:
            pass

# Remap using actual DB IDs
for arabic, eng in SERIES_MAP.items():
    if eng in existing_subcats:
        cat_arabic_to_uuid[arabic] = existing_subcats[eng]
fallback_id = existing_subcats.get('OTHERS')

# Now do the product inserts
sql_insert = []
unit = 'قطعة'
now = "NOW()"
insert_count = 0

for vals in rows:
    series_ar = vals[1]
    name = vals[2]
    code = vals[3]
    size = vals[4]
    color = vals[5]
    price_str = vals[6]
    
    subcat_id = cat_arabic_to_uuid.get(series_ar, fallback_id)
    if not subcat_id:
        print(f"NO SUBCAT for: {series_ar}")
        continue
    
    price_val = 0
    try:
        price_val = float(price_str.replace(',', ''))
    except:
        pass
    
    prod_id = str(uuid.uuid4())
    safe_name = name.replace("'", "''")
    safe_size = size.replace("'", "''") if size else ''
    
    sql_insert.append(
        f"INSERT INTO products (id, subcategory_id, name, barcode, unit, "
        f"retail_price, wholesale_price, cost_price, company, size, type, material, "
        f"is_active, created_at, updated_at, stock_status) VALUES ("
        f"'{prod_id}', '{subcat_id}', E'{safe_name}', '{code}', '{unit}', "
        f"{price_val}, {price_val}, {price_val}, 'دروفيت', E'{safe_size}', '', '', "
        f"true, {now}, {now}, 'untracked');"
    )
    insert_count += 1

print(f"Products to insert: {insert_count}")

sql_path = r'C:\eg-co-erp\reimport_drovit_v4.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_insert))

subprocess.run(['docker', 'cp', sql_path, 'eg-co-erp-db-1:/tmp/reimport_drovit_v4.sql'], capture_output=True)
r2 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db', '-f', '/tmp/reimport_drovit_v4.sql'],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
errors = r2.stderr.count('ERROR') if r2.stderr else 0
print(f"Insert errors: {errors}")
if r2.stderr and errors > 0:
    for l in r2.stderr.split('\n')[:5]:
        if 'ERROR' in l:
            print(f"  {l[:150]}")

r3 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A',
     '-c', "SELECT count(*) FROM products WHERE company = 'دروفيت'"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(f"دروفيت in DB: {r3.stdout.strip()}")
