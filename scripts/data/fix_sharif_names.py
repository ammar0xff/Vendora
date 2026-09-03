import openpyxl, sys
from decimal import Decimal

sys.stdout.reconfigure(encoding='utf-8')

path = r'C:\eg-co-erp\LISTS\done\الشريف.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['All_Products']

# Build clean names from Excel & save as UPDATE SQL
lines = []

def sql(s):
    lines.append(s)

def sv(v):
    return str(v) if v is not None else ''

# Map الفئة to subcategory name
# The subcategory names in the DB match الفئة values like مواسير, وصلات, etc., 
# but with the system prefix
# We need to find the right subcategory_id

# First, let's figure out which subcategory each product belongs to
# by looking at الفئة column and the system prefix

subcat_map = {}  # (company, الفئة) -> subcategory name in DB

for row in ws.iter_rows(min_row=2, values_only=True):
    system, cat, prod_name, size, thickness, price = row
    if not system:
        continue
    system = sv(system).strip()
    cat = sv(cat).strip()
    
    # The DB subcategory was likely set to the Arabic value from الفئة
    # or something like the original system name
    # Let me just look at what subcategories الشريف products use
    pass

# Since I can't be sure of the exact subcategory mapping from here,
# I'll just UPDATE by matching on the original constructed name

# Build: original name that was in DB
# Original pattern: "{A} - {C} - {D}" (+ "سمك {E}" for pipes/moases)
# After الرصف->الصرف: same but with صرف

# Clean name pattern: "{C} - {D}" (+ " - سمك {E}" if thickness)
for row in ws.iter_rows(min_row=2, values_only=True):
    system, cat, prod_name, size, thickness, price = row
    if not system:
        continue
    system = sv(system).strip()
    cat = sv(cat).strip()
    prod_name = sv(prod_name).strip()
    size = sv(size).strip()
    thickness = sv(thickness).strip()
    price_v = float(price) if price else None
    
    # Construct the old DB name (before my corruption)
    # Original: {A} - {C} - {D}  
    # After الرصف->الصرف: {system_with_صرف} - {C_with_صرف} - {D}
    old_name_prefix = system.replace('الرصف', 'الصرف')
    old_prod_name = prod_name.replace('الرصف', 'الصرف')
    
    # Reconstruct what the DB name WAS before my SUBSTRING corruption
    if cat == 'مواسير' and thickness:
        old_db_name = f'{old_name_prefix} - {old_prod_name} - {size} - سمك {thickness}'
        clean_name = f'{old_prod_name} - {size} - سمك {thickness}'
    else:
        old_db_name = f'{old_name_prefix} - {old_prod_name} - {size}'
        clean_name = f'{old_prod_name} - {size}'
    
    # Escape single quotes
    old_name_esc = old_db_name.replace("'", "''")
    clean_name_esc = clean_name.replace("'", "''")
    
    # Only UPDATE UPVC-N products (which got corrupted)
    if 'UPVC-N' in system:
        sql(f"UPDATE products SET name = '{clean_name_esc}' WHERE company = 'الشريف' AND name = '{old_name_esc}';")
    # For other systems, also clean if they have prefixes
    elif 'PPR نظام' in system and False:  # disabled - user didn't ask
        sql(f"UPDATE products SET name = '{clean_name_esc}' WHERE company = 'الشريف' AND name = '{old_name_esc}';")

output = r'C:\eg-co-erp\fix_sharif_names.sql'
with open(output, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

total_sql = len(lines)
print(f"Wrote {total_sql} UPDATE statements")
