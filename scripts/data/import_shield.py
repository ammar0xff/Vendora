import openpyxl, subprocess, uuid

def clean(val):
    if val is None: return ''
    return str(val).strip()

def fmt_price(val):
    try:
        return f"{float(str(val).replace(',', '')):.2f}"
    except:
        return '0'

def esc(val):
    return str(val).replace("'", "''")

now = "NOW()"
CAT_NAME = 'شيلد'

sql_parts = []

r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-c', f"SELECT id FROM categories WHERE name = '{CAT_NAME}'"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
cat_id = r.stdout.strip()

if cat_id:
    sql_parts.append(f"DELETE FROM products WHERE company = '{CAT_NAME}';")
    sql_parts.append(f"DELETE FROM subcategories WHERE category_id = '{cat_id}';")
else:
    cat_id = str(uuid.uuid4())
    sql_parts.append(f"INSERT INTO categories (id, name) VALUES ('{cat_id}', '{CAT_NAME}');")

print(f"Category: {CAT_NAME} -> {cat_id}")

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_أسعار_شيلد_يناير_2026.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]

# Group by column 1 (category)
from collections import OrderedDict
groups = OrderedDict()
for r in range(6, ws.max_row + 1):
    cat = clean(ws.cell(r, 1).value)
    prod = clean(ws.cell(r, 2).value)
    size = clean(ws.cell(r, 3).value)
    pack = clean(ws.cell(r, 4).value)
    price = ws.cell(r, 5).value
    specs = clean(ws.cell(r, 6).value)
    if not prod:
        continue
    if cat not in groups:
        groups[cat] = []
    groups[cat].append((prod, size, pack, price, specs))

print(f"Categories ({len(groups)}):")
total = 0
for cat, items in groups.items():
    print(f"  {cat}: {len(items)} items")
    subcat_id = str(uuid.uuid4())
    safe_cat = esc(cat)
    sql_parts.append(
        f"INSERT INTO subcategories (id, category_id, name, created_at) "
        f"VALUES ('{subcat_id}', '{cat_id}', E'{safe_cat}', {now});"
    )
    for prod, size, pack, price, specs in items:
        # Build name: product + size + specs 
        parts = [prod]
        if size and size != '-':
            parts.append(size)
        name = ' - '.join(parts)
        if specs and specs != '-':
            name = f"{name} ({specs[:80]})" if len(specs) < 80 else f"{name} ({specs[:77]}...)"
        
        safe_name = esc(name)
        safe_size = esc(size)
        pv = fmt_price(price)
        prod_id = str(uuid.uuid4())
        
        sql_parts.append(
            f"INSERT INTO products (id, subcategory_id, name, barcode, unit, "
            f"retail_price, wholesale_price, cost_price, company, size, type, material, "
            f"is_active, created_at, updated_at, stock_status) VALUES ("
            f"'{prod_id}', '{subcat_id}', E'{safe_name}', '', 'قطعة', "
            f"{pv}, {pv}, {pv}, '{CAT_NAME}', E'{safe_size}', '', '', "
            f"true, {now}, {now}, 'untracked');"
        )
        total += 1

print(f"Total: {total}")

sql_path = r'C:\eg-co-erp\import_shield.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_parts))

subprocess.run(['docker', 'cp', sql_path, 'eg-co-erp-db-1:/tmp/import_shield.sql'], capture_output=True)
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-f', '/tmp/import_shield.sql'],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
errors = r.stderr.count('ERROR') if r.stderr else 0
print(f"Errors: {errors}")
if r.stderr and errors > 0:
    for l in r.stderr.split('\n')[:15]:
        if 'ERROR' in l:
            print(f"  {l[:200]}")

r2 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-c', f"SELECT company, count(*) FROM products WHERE company = '{CAT_NAME}' GROUP BY company;"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(r2.stdout)

r3 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-c', f"SELECT s.name, count(p.id) FROM subcategories s LEFT JOIN products p ON p.subcategory_id = s.id WHERE s.category_id = '{cat_id}' GROUP BY s.name ORDER BY s.name;"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(r3.stdout)
