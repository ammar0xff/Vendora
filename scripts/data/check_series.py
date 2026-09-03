import openpyxl
wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_اسعار_دورافيت_2026_عربي.xlsx', data_only=True)
ws = wb.active

series_set = set()
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)):
    vals = [str(v).strip() if v is not None else '' for v in row]
    if vals[1]:
        series_set.add(vals[1])

print(f"Series names found: {len(series_set)}")
for s in sorted(series_set):
    print(f"  '{s}'")
