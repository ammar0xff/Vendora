import openpyxl

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_أسعار_شيلد_يناير_2026.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]

# Focus on rows 6-8 (محبس زاوية with price 125)
for r in [6, 7, 8]:
    subcat = ws.cell(r, 1).value or ''
    prod = ws.cell(r, 2).value or ''
    size = ws.cell(r, 3).value or ''
    pack = ws.cell(r, 4).value or ''
    price = ws.cell(r, 5).value or ''
    
    print(f'Row {r}:')
    print(f'  subcat: {repr(subcat)}')
    print(f'  prod:   {repr(prod)}')
    print(f'  size:   {repr(size)}')
    print(f'  pack:   {repr(pack)}')
    print(f'  price:  {repr(price)}')
    print(f'  prod hex: {" ".join(hex(ord(c)) for c in str(prod))}')
    print()
