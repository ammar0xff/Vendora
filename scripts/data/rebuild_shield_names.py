import openpyxl, subprocess

# Read Excel data
wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_أسعار_شيلد_يناير_2026.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]

# Build lookup: subcategory -> price -> clean_name
excel_data = {}  # (subcat, price_str) -> clean_name
for r in range(6, ws.max_row + 1):
    subcat = str(ws.cell(r, 1).value or '').strip()
    prod = str(ws.cell(r, 2).value or '').strip()
    size = str(ws.cell(r, 3).value or '').strip()
    price = ws.cell(r, 5).value
    if not prod:
        continue
    # Clean name: product type + size
    if size and size != '-':
        clean = f'{prod} - {size}'
    else:
        clean = prod
    price_str = str(price) if price is not None else ''
    excel_data[(subcat, price_str)] = clean

# Get current products from DB
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-F|', '-c',
     "SELECT p.id, p.retail_price, s.name FROM products p JOIN subcategories s ON p.subcategory_id = s.id WHERE p.company = 'شيلد'"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)

updates = []
lines = [l for l in r.stdout.strip().split('\n') if l.strip()]
for line in lines:
    parts = line.split('|')
    if len(parts) < 3:
        continue
    prod_id = parts[0].strip()
    db_price = parts[1].strip()
    db_subcat = parts[2].strip()
    
    key = (db_subcat, db_price)
    clean_name = excel_data.get(key)
    
    if clean_name:
        price_suffix = ''
        # If duplicate with same subcat+price, add price to name
        same_keys = [k for k in excel_data if k[0] == db_subcat and k[1] == db_price]
        if len(same_keys) > 1:
            # Check if duplicate by name too
            names_at_key = [excel_data[k] for k in same_keys]
            if names_at_key.count(clean_name) > 1:
                price_suffix = f' [{db_price} EGP]'
        
        final_name = clean_name + price_suffix
        final_name = final_name.replace("'", "''")
        updates.append(f"UPDATE products SET name = E'{final_name}' WHERE id = '{prod_id}';")
    else:
        # Fallback: just strip anything that looks like specs
        print(f'No match: subcat={db_subcat}, price={db_price} -> id={prod_id[:8]}')

print(f'Generated {len(updates)} updates')

sql_path = r'C:\eg-co-erp\rebuild_shield_names.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(updates))

subprocess.run(['docker', 'cp', sql_path, 'eg-co-erp-db-1:/tmp/rebuild_shield_names.sql'], capture_output=True)
r2 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-f', '/tmp/rebuild_shield_names.sql'],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
errors = r2.stderr.count('ERROR') if r2.stderr else 0
print(f'Errors: {errors}')
if errors > 0:
    for l in r2.stderr.split('\n')[:10]:
        if 'ERROR' in l:
            print(f'  {l[:200]}')

# Verify
r3 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-c', "SELECT name FROM products WHERE company = 'شيلد' ORDER BY name;"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(r3.stdout)

# Check duplicates
r4 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-c', "SELECT count(*) FROM (SELECT name FROM products WHERE company = 'شيلد' GROUP BY name HAVING count(*) > 1) d;"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(f'Duplicates: {r4.stdout.strip()}')
