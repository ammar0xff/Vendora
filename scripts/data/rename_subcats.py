import openpyxl, subprocess, json

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_اسعار_دورافيت_2026_عربي.xlsx', data_only=True)
ws = wb.active

# Get unique series Arabic names from Excel
series_arabic = set()
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)):
    vals = [str(v).strip() if v is not None else '' for v in row]
    if vals[1]:
        series_arabic.add(vals[1])

# English translations for each Arabic series name
SERIES_MAP = {
    'بي 3 كومفورتس': 'P3 COMFORTS',
    'ستارك 3': 'STARCK 3',
    'ستارك 1': 'STARCK 1',
    'دي نيو': 'D-NEO',
    'دورا ستايل': 'DURASTYLE',
    'دارلينج نيو': 'DARLING',
    'دارلينج': 'DARLING',
    'هابي دي': 'HAPPY D.',
    'هابي دي 2': 'HAPPY D.',
    'دورافيت نمبر 1': 'DURAVIT NO. 1',
    'دي كود': 'D-CODE',
    'إيكو': 'ECHO',
    'دورابلاس': 'DURAPLUS',
    'إميليا': 'EMILIA',
    'جولف': 'GOLF',
    'ديون': 'DUNE',
    'متنوعة': 'OTHERS',
    'إكسسوارات سيراميك': 'ACCESSORIES CERAMIC',
    'إكسسوارات إيزي': 'ACCESSORIES EASY',
    'إكسسوارات كروم': 'ACCESSORIES CHROME',
    'إكس لارج': 'X-LARGE',
    'فوستر': 'FOSTER',
    'كيثو': 'KETHO',
    'كارو': 'CARO',
    'إل كيوب': 'L-CUBE',
    'فيتريوم': 'VITRIUM',
    'بيورا فيدا': 'PURAVIDA',
    'فيرو': 'VERO',
    'الطابق الثاني': 'SECOND FLOOR',
    'عمود الطابق الثاني': 'SECOND FLOOR COLUMN',
    'مجموعة تثبيت': 'MOUNTING SET',
}

# Reverse: English -> Arabic
eng_to_arabic = {}
for ar, en in SERIES_MAP.items():
    if en in eng_to_arabic:
        # Keep the longer/more specific Arabic name
        existing = eng_to_arabic[en]
        if len(ar) > len(existing):
            eng_to_arabic[en] = ar
    else:
        eng_to_arabic[en] = ar

# Update subcategory names
updates = []
for en, ar in eng_to_arabic.items():
    safe = ar.replace("'", "''")
    updates.append(f"UPDATE subcategories SET name = E'{safe}' WHERE name = '{en}' AND category_id = '6d3ebf46-0d63-45a7-b271-74ce542d732b';")

sql = '\n'.join(updates)
print(f"Updating {len(updates)} subcategories")

sql_path = r'C:\eg-co-erp\rename_drovit_subcats.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write(sql)

subprocess.run(['docker', 'cp', sql_path, 'eg-co-erp-db-1:/tmp/rename_drovit_subcats.sql'], capture_output=True)
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db', '-f', '/tmp/rename_drovit_subcats.sql'],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
errors = r.stderr.count('ERROR') if r.stderr else 0
print(f"Errors: {errors}")

# Verify
r2 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-c', "SELECT name FROM subcategories WHERE category_id = '6d3ebf46-0d63-45a7-b271-74ce542d732b' ORDER BY name"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
# Write to file
with open('C:\\eg-co-erp\\drovit_subcats.txt', 'w', encoding='utf-8') as f:
    f.write(r2.stdout)
print("Done - results in C:\\eg-co-erp\\drovit_subcats.txt")
