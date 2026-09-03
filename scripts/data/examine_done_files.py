import openpyxl

out = open(r'C:\eg-co-erp\done_files_structure.txt', 'w', encoding='utf-8')

files = [
    r'C:\eg-co-erp\LISTS\done\الشريف.xlsx',
    r'C:\eg-co-erp\LISTS\done\ايجيك.xlsx',
    r'C:\eg-co-erp\LISTS\done\Neisco_Comer_Price_List_2026.xlsx',
    r'C:\eg-co-erp\LISTS\done\قائمة أسعار روك بولي 2024 - جداول منظمة.xlsx',
    r'C:\eg-co-erp\LISTS\done\قائمة_أسعار_روك_110_المطورة.xlsx',
    r'C:\eg-co-erp\LISTS\done\قائمة_أسعار_روك_١١٤.xlsx',
]

for fp in files:
    out.write(f"\n{'='*60}\n{fp}\n{'='*60}\n")
    try:
        wb = openpyxl.load_workbook(fp, data_only=True)
    except Exception as e:
        out.write(f"Error: {e}\n")
        continue
    out.write(f"Sheets ({len(wb.sheetnames)}):\n")
    for i, sn in enumerate(wb.sheetnames):
        ws = wb[sn]
        out.write(f"  {i}: {repr(sn)} (rows={ws.max_row}, cols={ws.max_column})\n")
        # Show first 3 rows
        for r in range(1, min(ws.max_row + 1, 4)):
            vals = []
            for c in range(1, min(ws.max_column + 1, 8)):
                cell = ws.cell(row=r, column=c)
                if cell.value is not None:
                    vals.append(f"{c}:{repr(str(cell.value)[:50])}")
            if vals:
                out.write(f"    Row {r}: {vals}\n")
out.close()
print("Done")
