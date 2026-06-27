import os, glob
from openpyxl import load_workbook

BASE = r'C:\eg-co-erp\LISTS\جرد المخازن'
files = sorted(glob.glob(os.path.join(BASE, '*.xlsx')))

for idx, f in enumerate(files):
    wb = load_workbook(f, read_only=True, data_only=True)
    ws = wb.active
    fname = os.path.basename(f)
    print(f'[{idx:2d}] {fname}')
    print(f'      Shape: {ws.max_row} rows x {ws.max_column} cols')
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:
            print(f'      Row {i}: {list(row)}')
        else:
            break
    # Print a few data rows too
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if 3 <= i < 8:
            print(f'      Row {i}: {list(row)}')
    wb.close()
    print()
