import os, glob, json
from openpyxl import load_workbook

BASE = r'C:\eg-co-erp\LISTS\جرد المخازن'
files = sorted(glob.glob(os.path.join(BASE, '*.xlsx')))

for f in files:
    wb = load_workbook(f, read_only=True, data_only=True)
    ws = wb.active
    fname = os.path.basename(f)
    print(f'=== {fname} ===')
    
    # Print header rows (first 5 non-empty rows)
    headers_found = 0
    for row in ws.iter_rows(values_only=True):
        row = list(row)
        # Print any row that has text headers
        has_arabic = False
        for v in row:
            if v is not None and isinstance(v, str) and any('\u0600' <= c <= '\u06FF' for c in v.strip()):
                has_arabic = True
                break
        if has_arabic and headers_found < 3:
            print(f'  {row}')
            headers_found += 1
    
    # Print first 3 data rows (skip header rows)
    data_count = 0
    for row in ws.iter_rows(values_only=True):
        row = list(row)
        # Find product name column
        name_idx = None
        for i, v in enumerate(row):
            if v is not None and isinstance(v, str) and len(v.strip()) > 3:
                if any('\u0600' <= c <= '\u06FF' for c in v.strip()):
                    if not any(kw in v for kw in ['اسم', 'الصنف', 'العدد', 'رقم', 'اجمالى', 'بيان']):
                        name_idx = i
                        break
        if name_idx is not None and data_count < 3:
            print(f'  Row with name@{name_idx}: {row}')
            data_count += 1
    
    wb.close()
    print()
