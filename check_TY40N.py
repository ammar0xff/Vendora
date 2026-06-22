import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\Neisco_Comer_Price_List_2026.xlsx', data_only=True)

ws = wb['Stubs Flanges Unions Metric']
print("=== Checking TY40N0250 row ===")
for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
    code, size, pack, box, price = row if len(row) >= 5 else (row[0] if row else None, None, None, None, None)
    if not code:
        continue
    code_str = str(code).strip()
    if 'TY40' in code_str:
        print(f'Row {i}: code={code_str!r}, size={size!r}, price={price!r}, pack={pack!r}, box={box!r}')
        size_str = str(size).strip() if size else 'None'
        price_str = str(price).strip() if price else 'None'
        print(f'  size_str={size_str!r}, price_str={price_str!r}')

# Also let me print ALL rows in that sheet for reference
print("\n=== ALL rows in Stubs Flanges Unions Metric ===")
for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
    if not row[0]:
        continue
    code, size, pack, box, price = row if len(row) >= 5 else (row[0] if row else None, None, None, None, None)
    code_str = str(code).strip()
    size_str = str(size).strip() if size else '?'
    price_val = price if price else 0
    print(f'Row {i}: {code_str:15s} | {size_str:15s} | price={price_val}')
