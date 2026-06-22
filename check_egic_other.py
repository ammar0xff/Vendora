import openpyxl, sys, os
sys.stdout.reconfigure(encoding='utf-8')

d = r'C:\eg-co-erp\LISTS\done'
for f in os.listdir(d):
    if not f.endswith('.xlsx') or 'شريف' in repr(f)[:50]:
        continue
    path = os.path.join(d, f)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Sheet1']

print("=== المنتجات المصنفة تحت 'أخرى' (Other) ===")
for row in ws.iter_rows(min_row=2, values_only=True):
    code, material, brand, price = row
    if not material:
        continue
    material_str = str(material).strip()
    words = material_str.split()
    first = words[0] if words else ''

    # Determine if it falls under 'أخرى'
    if first not in ('مواسير', 'كوع', 'تي', 'وصلة', 'مسلوب', 'جلبة', 'طبة', 'صليبة',
                     'مشترك', 'محبس', 'صفاية', 'علاية', 'غطاء', 'بردة', 'بيبة',
                     'ياردة', 'غراء', 'هواية', 'رقبة', 'مخرج', 'جاليتراب', 'قفيز',
                     'واي', 'طلمبة', 'سيفون', 'مجمع', 'غرفة', 'S', 'درجة', 'P.P',
                     'بطارية') and \
       not first.startswith('S-') and not first.startswith('Sمم') and not first.startswith('S-مم') and \
       not first.startswith('S-°') and not first.startswith('PP-مم') and \
       not first.startswith('ML') and not first.startswith('UVمم') and not first.startswith('PP'):
        brand_str = str(brand).strip() if brand else ''
        price_str = str(price) if price else 'N/A'
        print(f'  {material_str:70s} | brand={brand_str} | price={price_str} | code={code}')

# Also check for weird size patterns
print("\n=== فحص أنماط المقاسات (Size patterns) ===")
for row in ws.iter_rows(min_row=2, values_only=True):
    code, material, brand, price = row
    if not material:
        continue
    material_str = str(material).strip()
    words = material_str.split()
    
    # Find size - usually last word or word containing digits + مم
    last_word = words[-1] if words else ''
    has_mm = any('مم' in w or 'M' in w.upper() for w in words)
    print(f'  {material_str:60s} | last word: {last_word:15s} | has مم: {has_mm}')
    if 'مم' not in material_str:
        print(f'    *** NO مم in name!')
