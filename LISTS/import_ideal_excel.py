import openpyxl, subprocess, uuid

# Check ايديال category ID
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-c', "SELECT id FROM categories WHERE name = 'ايديال'"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
IDEAL_CAT_ID = r.stdout.strip()
print(f"ايديال category ID: {IDEAL_CAT_ID}")

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\Ideal_Standard_2026_Price_List.xlsx', data_only=True)
print(f"Sheets: {len(wb.sheetnames)}")

sql_parts = []
# Delete old ايديال products and subcategories
sql_parts.append(f"DELETE FROM products WHERE subcategory_id IN (SELECT id FROM subcategories WHERE category_id = '{IDEAL_CAT_ID}');")
sql_parts.append(f"DELETE FROM subcategories WHERE category_id = '{IDEAL_CAT_ID}';")

now = "NOW()"
total = 0

for sn in wb.sheetnames:
    ws = wb[sn]
    subcat_id = str(uuid.uuid4())
    safe_name = sn.replace("'", "''")
    sql_parts.append(
        f"INSERT INTO subcategories (id, category_id, name, created_at) "
        f"VALUES ('{subcat_id}', '{IDEAL_CAT_ID}', E'{safe_name}', {now});"
    )
    
    # Determine column structure from header
    header = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        header.append(str(h).strip() if h else '')
    
    # Columns: code=0, name_ar=1, name_en=2, price_white=3, price_beige=4, weight=5
    # Some have 5 cols (no beige price) or 5 cols (price + weight)
    has_beige = any('برجامون' in h for h in header)
    has_weight = any('الوزن' in h for h in header)
    
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        code = vals[0]
        name = vals[1]
        if not code or not name:
            continue
        code = str(code).strip()
        name = str(name).strip()
        
        # Price: column 3 (0-indexed) = white, 4 = beige if exists
        price_val = 0
        if len(vals) > 3 and vals[3] is not None:
            try:
                price_val = float(str(vals[3]).replace(',', ''))
            except:
                pass
        
        prod_id = str(uuid.uuid4())
        safe_prod = name.replace("'", "''")
        safe_code = code.replace("'", "''")
        
        sql_parts.append(
            f"INSERT INTO products (id, subcategory_id, name, barcode, unit, "
            f"retail_price, wholesale_price, cost_price, company, size, is_active, "
            f"created_at, updated_at, stock_status) VALUES ("
            f"'{prod_id}', '{subcat_id}', E'{safe_prod}', '{safe_code}', 'قطعة', "
            f"{price_val}, {price_val}, {price_val}, 'ايديال', '', "
            f"true, {now}, {now}, 'untracked');"
        )
        total += 1

print(f"Total products: {total}")

sql_path = r'C:\eg-co-erp\import_ideal_excel.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_parts))

subprocess.run(['docker', 'cp', sql_path, 'eg-co-erp-db-1:/tmp/import_ideal_excel.sql'], capture_output=True)
r2 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db', '-f', '/tmp/import_ideal_excel.sql'],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
errors = r2.stderr.count('ERROR') if r2.stderr else 0
print(f"Errors: {errors}")
if r2.stderr and errors > 0:
    for l in r2.stderr.split('\n')[:15]:
        if 'ERROR' in l:
            print(f"  {l[:200]}")

r3 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-c', "SELECT count(*) FROM products WHERE company = 'ايديال'"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(f"ايديال in DB: {r3.stdout.strip()}")

r4 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-c', "SELECT s.name, count(p.id) FROM subcategories s LEFT JOIN products p ON p.subcategory_id = s.id WHERE s.category_id = '" + IDEAL_CAT_ID + "' GROUP BY s.name ORDER BY s.name"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
with open('C:\\eg-co-erp\\ideal_by_sheet.txt', 'w', encoding='utf-8') as f:
    f.write(r4.stdout)
print("Written to C:\\eg-co-erp\\ideal_by_sheet.txt")
