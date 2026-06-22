import openpyxl

out = open(r'C:\eg-co-erp\comer_all_data.txt', 'w', encoding='utf-8')
wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\done\Neisco_Comer_Price_List_2026.xlsx', data_only=True)

total_rows = 0
for sn in wb.sheetnames:
    ws = wb[sn]
    out.write(f"\n=== {sn} (rows={ws.max_row}) ===\n")
    for r in range(1, ws.max_row + 1):
        vals = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is not None:
                vals.append(repr(str(v)[:80]))
        if vals:
            total_rows += 1
            out.write(f"  R{r}: {vals}\n")

out.write(f"\n\nTotal non-empty rows: {total_rows}\n")
out.close()
print(f"Done. Total rows: {total_rows}")
