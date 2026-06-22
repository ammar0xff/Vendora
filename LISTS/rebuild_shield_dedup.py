import openpyxl, subprocess, re
from collections import defaultdict

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_أسعار_شيلد_يناير_2026.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]

# Excel rows in order - use list per key
excel_map = defaultdict(list)
for r_idx in range(6, ws.max_row + 1):
    subcat = str(ws.cell(r_idx, 1).value or '').strip()
    subcat_bytes = subcat.encode('utf-8').hex()
    prod = str(ws.cell(r_idx, 2).value or '').strip()
    size = str(ws.cell(r_idx, 3).value or '').strip()
    price = ws.cell(r_idx, 5).value
    specs = str(ws.cell(r_idx, 6).value or '').strip()
    price_str = str(int(price)) if isinstance(price, (int, float)) and price == int(price) else str(price or '')
    key = (price_str, subcat_bytes)
    excel_map[key].append({'prod': prod, 'size': size, 'specs': specs, 'subcat': subcat})

# DB products in order
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-F||', '-c',
     "SELECT p.id, p.retail_price::text, encode(convert_to(s.name, 'UTF8'), 'hex') "
     "FROM products p JOIN subcategories s ON p.subcategory_id = s.id "
     "WHERE p.company = 'شيلد' ORDER BY s.name, p.retail_price, p.id"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)

# Counter per key to handle duplicates
key_counters = defaultdict(int)
updates = []

generic = {'جسم نحاس', 'جسم نحاس (لوجو حفر)', 'جسم نحاس (متعدد الاستخدام)', 'سطح عالي الجودة', 'سطح خارجي عالي الجودة',
           'سطح عالي الجودة - طلاء نيكل كروم', 'طلاء نيكل كروم', 'جوان NBR عالي الجودة', 'جوان NBR'}

for line in [l for l in r.stdout.strip().split('\n') if l.strip()]:
    parts = line.split('||')
    if len(parts) < 3: continue
    prod_id = parts[0].strip()
    price = str(int(float(parts[1].strip())))
    subcat_hex = parts[2].strip()
    
    key = (price, subcat_hex)
    excel_list = excel_map.get(key, [])
    idx = key_counters[key]
    
    if idx < len(excel_list):
        row = excel_list[idx]
        key_counters[key] += 1
        
        prod = row['prod']
        size = row['size']
        specs = row['specs']
        subcat = row['subcat']
        
        # Extract key materials
        if specs and specs != '-':
            spec_parts = re.split(r'[،,]', specs)
            spec_parts = [s.strip() for s in spec_parts]
            meaningful = [s for s in spec_parts if s not in generic]
            key_mat = ' - '.join(meaningful[:2]) if meaningful else ' - '.join(spec_parts[:2])
        else:
            key_mat = ''
        
        if size and size != '-':
            if key_mat:
                new_name = f'{subcat} {prod} ({key_mat}) - {size}'
            else:
                new_name = f'{subcat} {prod} - {size}'
        else:
            if key_mat:
                new_name = f'{subcat} {prod} ({key_mat})'
            else:
                new_name = f'{subcat} {prod}'
        
        new_name = new_name.replace("'", "''")
        updates.append(f"UPDATE products SET name = E'{new_name}' WHERE id = '{prod_id}';")
    else:
        print(f'EXTRA DB: id={prod_id[:12]} price={price}')

print(f'Generated {len(updates)} updates')

# Apply
if updates:
    sql_path = r'C:\eg-co-erp\rebuild_shield_final2.sql'
    with open(sql_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(updates))
    subprocess.run(['docker', 'cp', sql_path, 'eg-co-erp-db-1:/tmp/rebuild_shield_final2.sql'], capture_output=True)
    r2 = subprocess.run(
        ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
         '-f', '/tmp/rebuild_shield_final2.sql'],
        capture_output=True, text=True, encoding='utf-8', errors='replace'
    )
    errors = r2.stderr.count('ERROR') if r2.stderr else 0
    print(f'Apply errors: {errors}')

# Verify
r3 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-c', "SELECT name FROM products WHERE company = 'شيلد' ORDER BY name;"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(r3.stdout)

# Check duplicates
r4 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-c', "SELECT count(*) FROM (SELECT name FROM products WHERE company = 'شيلد' GROUP BY name HAVING count(*) > 1) d;"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(f'Duplicates: {r4.stdout.strip()}')
