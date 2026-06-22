import openpyxl, subprocess, os

# Write output to file for clean encoding
out_path = r'C:\eg-co-erp\LISTS\shield_comparison.txt'

# Read current DB names
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-F||', '-c',
     "SELECT p.name, p.retail_price::text, trim(s.name) FROM products p JOIN subcategories s ON p.subcategory_id = s.id WHERE p.company = 'شيلد' ORDER BY s.name, p.retail_price"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)

db_data = {}
for line in [l for l in r.stdout.strip().split('\n') if l.strip()]:
    parts = line.split('||')
    if len(parts) < 3: continue
    name, price, subcat = parts[0].strip(), parts[1].strip().rstrip('.0'), parts[2].strip()
    key = (subcat, price)
    db_data[key] = name

# Read Excel
wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_أسعار_شيلد_يناير_2026.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]

lines = []
for r in range(6, ws.max_row + 1):
    subcat = str(ws.cell(r, 1).value or '').strip()
    prod = str(ws.cell(r, 2).value or '').strip()
    size = str(ws.cell(r, 3).value or '').strip()
    pack = str(ws.cell(r, 4).value or '').strip()
    price = ws.cell(r, 5).value
    specs = str(ws.cell(r, 6).value or '').strip()
    
    price_str = str(int(price)) if price is not None and isinstance(price, (int, float)) and price == int(price) else str(price) if price else ''
    key = (subcat, price_str)
    db_name = db_data.get(key, '?')
    
    lines.append(f'=== {subcat} ===')
    lines.append(f'  Current: {db_name}')
    lines.append(f'  Product: {prod}')
    lines.append(f'  Size:    {size}')
    lines.append(f'  Pack:    {pack}')
    lines.append(f'  Price:   {price}')
    lines.append(f'  Specs:   {specs[:200]}')
    lines.append('')

with open(out_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f'Output written to {out_path}')
print(f'File size: {os.path.getsize(out_path)} bytes')

