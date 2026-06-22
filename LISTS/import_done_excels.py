import openpyxl, subprocess, uuid

def clean(val):
    if val is None: return ''
    return str(val).strip()

def fmt_price(val):
    try:
        return f"{float(str(val).replace(',', '')):.2f}"
    except:
        return '0'

def esc(val):
    return val.replace("'", "''")

now = "NOW()"
sql_parts = []

# ============================================================
# 1. الشريف
# ============================================================
print("=== الشريف ===")
SHARIF_CAT_ID = None
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-c', "SELECT id FROM categories WHERE name = 'الشريف'"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
SHARIF_CAT_ID = r.stdout.strip()
print(f"  Category: {SHARIF_CAT_ID}")

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\done\الشريف.xlsx', data_only=True)
ws = wb['All_Products']

# Collect sections -> products
sections = {}
for r in range(2, ws.max_row + 1):
    sec = clean(ws.cell(r, 1).value)
    code = clean(ws.cell(r, 2).value)
    name = clean(ws.cell(r, 3).value)
    size = clean(ws.cell(r, 4).value)
    thickness = clean(ws.cell(r, 5).value)
    price = ws.cell(r, 6).value
    if not name and not code:
        continue
    if sec not in sections:
        sections[sec] = []
    sections[sec].append((code, name, size, thickness, price))

print(f"  Sections found: {list(sections.keys())}")
total_sharif = sum(len(v) for v in sections.values())
print(f"  Total products: {total_sharif}")

# Delete old
sql_parts.append("DELETE FROM products WHERE company = 'الشريف';")
sql_parts.append(f"DELETE FROM subcategories WHERE category_id = '{SHARIF_CAT_ID}';")

sharif_count = 0
for sec_name, prods in sections.items():
    subcat_id = str(uuid.uuid4())
    safe_name = esc(sec_name)
    sql_parts.append(
        f"INSERT INTO subcategories (id, category_id, name, created_at) "
        f"VALUES ('{subcat_id}', '{SHARIF_CAT_ID}', E'{safe_name}', {now});"
    )
    for code, name, size, thickness, price in prods:
        safe_code = esc(code)
        safe_name_prod = esc(name)
        safe_size = esc(size)
        safe_thick = esc(thickness)
        full_size = f"{safe_size} - {safe_thick}" if safe_thick else safe_size
        pv = fmt_price(price)
        prod_id = str(uuid.uuid4())
        sql_parts.append(
            f"INSERT INTO products (id, subcategory_id, name, barcode, unit, "
            f"retail_price, wholesale_price, cost_price, company, size, type, material, "
            f"is_active, created_at, updated_at, stock_status) VALUES ("
            f"'{prod_id}', '{subcat_id}', E'{safe_name_prod}', '{safe_code}', 'قطعة', "
            f"{pv}, {pv}, {pv}, 'الشريف', E'{full_size}', '', '', "
            f"true, {now}, {now}, 'untracked');"
        )
        sharif_count += 1
print(f"  Generated {sharif_count} products")

# ============================================================
# 2. ايجيك
# ============================================================
print("\n=== ايجيك ===")
EGIC_CAT_ID = None
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-c', "SELECT id FROM categories WHERE name = 'ايجيك'"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
EGIC_CAT_ID = r.stdout.strip()
print(f"  Category: {EGIC_CAT_ID}")

wb2 = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\done\ايجيك.xlsx', data_only=True)
ws2 = wb2['Sheet1']

# Group by brand
brand_prods = {}
for r in range(2, ws2.max_row + 1):
    code = clean(ws2.cell(r, 1).value)
    name = clean(ws2.cell(r, 2).value)
    brand = clean(ws2.cell(r, 3).value) or 'اخرى'
    price = ws2.cell(r, 4).value
    if not code and not name:
        continue
    if brand not in brand_prods:
        brand_prods[brand] = []
    brand_prods[brand].append((code, name, price))

print(f"  Brands found: {list(brand_prods.keys())}")
total_egic = sum(len(v) for v in brand_prods.values())
print(f"  Total products: {total_egic}")

# First, get IDs of current subcategories under ايجيك
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-c', f"SELECT id FROM subcategories WHERE category_id = '{EGIC_CAT_ID}'"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
old_subcat_ids = [x.strip() for x in r.stdout.strip().split('\n') if x.strip()]

# Delete old products linked to these subcategories
for sid in old_subcat_ids:
    sql_parts.append(f"DELETE FROM products WHERE subcategory_id = '{sid}';")
# Delete old subcategories
sql_parts.append(f"DELETE FROM subcategories WHERE category_id = '{EGIC_CAT_ID}';")

egic_count = 0
for brand, prods in brand_prods.items():
    subcat_id = str(uuid.uuid4())
    safe_brand = esc(brand)
    sql_parts.append(
        f"INSERT INTO subcategories (id, category_id, name, created_at) "
        f"VALUES ('{subcat_id}', '{EGIC_CAT_ID}', E'{safe_brand}', {now});"
    )
    for code, name, price in prods:
        safe_code = esc(code)
        safe_name_prod = esc(name)
        pv = fmt_price(price)
        prod_id = str(uuid.uuid4())
        # Also include brand info in the name to distinguish
        display_name = f"[{brand}] {safe_name_prod}"
        sql_parts.append(
            f"INSERT INTO products (id, subcategory_id, name, barcode, unit, "
            f"retail_price, wholesale_price, cost_price, company, size, type, material, "
            f"is_active, created_at, updated_at, stock_status) VALUES ("
            f"'{prod_id}', '{subcat_id}', E'{display_name}', '{safe_code}', 'قطعة', "
            f"{pv}, {pv}, {pv}, 'ايجيك', '', '', '', "
            f"true, {now}, {now}, 'untracked');"
        )
        egic_count += 1
print(f"  Generated {egic_count} products")

# ============================================================
# 3. Neisco_Comer (كومر)
# ============================================================
print("\n=== كومر ===")
COMER_CAT_ID = None
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-c', "SELECT id FROM categories WHERE name = 'كومر'"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
COMER_CAT_ID = r.stdout.strip()
print(f"  Category: {COMER_CAT_ID}")

wb3 = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\done\Neisco_Comer_Price_List_2026.xlsx', data_only=True)
sheet_names = wb3.sheetnames

# Extract short name for subcategory from first row of each sheet
def get_short_name(sheet):
    ws = wb3[sheet]
    first_val = ws.cell(1, 1).value
    if first_val:
        s = str(first_val).strip().split('-')[0].strip()
        if len(s) > 80:
            s = s[:80]
        return s
    return sheet

products_per_sheet = []
for sn in sheet_names:
    ws = wb3[sn]
    count = 0
    for r in range(4, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if code and str(code).strip():
            count += 1
    products_per_sheet.append((sn, count))
    print(f"  {sn}: {count} products")

total_comer = sum(x[1] for x in products_per_sheet)
print(f"  Total: {total_comer}")

# Delete old
sql_parts.append("DELETE FROM products WHERE company = 'كومر';")
sql_parts.append(f"DELETE FROM subcategories WHERE category_id = '{COMER_CAT_ID}';")

comer_count = 0
for sn in sheet_names:
    ws = wb3[sn]
    subcat_id = str(uuid.uuid4())
    safe_sn = esc(get_short_name(sn))
    sql_parts.append(
        f"INSERT INTO subcategories (id, category_id, name, created_at) "
        f"VALUES ('{subcat_id}', '{COMER_CAT_ID}', E'{safe_sn}', {now});"
    )
    for r in range(4, ws.max_row + 1):
        code = clean(ws.cell(r, 1).value)
        size = clean(ws.cell(r, 2).value)
        pack = clean(ws.cell(r, 3).value)
        box = clean(ws.cell(r, 4).value)
        price = ws.cell(r, 5).value
        if not code:
            continue
        # Use code as both barcode and part of name
        safe_code = esc(code)
        safe_size = esc(size)
        pv = fmt_price(price)
        prod_id = str(uuid.uuid4())
        # Name = code + size
        prod_name = f"{safe_code} {safe_size}" if safe_size else safe_code
        sql_parts.append(
            f"INSERT INTO products (id, subcategory_id, name, barcode, unit, "
            f"retail_price, wholesale_price, cost_price, company, size, type, material, "
            f"is_active, created_at, updated_at, stock_status) VALUES ("
            f"'{prod_id}', '{subcat_id}', E'{prod_name}', '{safe_code}', 'قطعة', "
            f"{pv}, {pv}, {pv}, 'كومر', E'{safe_size}', '', '', "
            f"true, {now}, {now}, 'untracked');"
        )
        comer_count += 1
print(f"  Generated {comer_count} products")

# ============================================================
# Write and execute SQL
# ============================================================
sql_path = r'C:\eg-co-erp\import_done_excels.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_parts))
print(f"\nSQL written to {sql_path}")
print(f"Total statements: {len(sql_parts)}")

print("Copying SQL to container...")
subprocess.run(['docker', 'cp', sql_path, 'eg-co-erp-db-1:/tmp/import_done_excels.sql'],
               capture_output=True)

print("Executing SQL...")
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-f', '/tmp/import_done_excels.sql'],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
errors = r.stderr.count('ERROR') if r.stderr else 0
print(f"Errors: {errors}")
if r.stderr and errors > 0:
    for l in r.stderr.split('\n')[:15]:
        if 'ERROR' in l:
            print(f"  {l[:200]}")

# Verify
print("\n=== Verification ===")
r2 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-c', "SELECT company, count(*) FROM products WHERE company IN ('الشريف', 'ايجيك', 'كومر') GROUP BY company ORDER BY company"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(r2.stdout)

r3 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-c', "SELECT c.name as company, s.name, count(p.id) as prods FROM categories c JOIN subcategories s ON s.category_id = c.id LEFT JOIN products p ON p.subcategory_id = s.id WHERE c.name IN ('الشريف', 'ايجيك', 'كومر') GROUP BY c.name, s.name ORDER BY c.name, s.name"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(r3.stdout)
