import openpyxl, subprocess, json, uuid

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_اسعار_دورافيت_2026_عربي.xlsx', data_only=True)
ws = wb.active

# Read all rows
rows = []
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)):
    vals = [str(v).strip() if v is not None else '' for v in row]
    if vals[2] and vals[3]:  # must have product name and code
        rows.append(vals)

print(f"Excel rows: {len(rows)}")

# Map series Arabic name to English subcategory name
SERIES_MAP = {
    'بي 3 كومفورتس': 'P3 COMFORTS',
    'ستارك 3': 'STARCK 3',
    'دي نيو': 'D-NEO',
    'دورا ستايل': 'DURASTYLE',
    'دارلينج نيو': 'DARLING',
    'هابي دي 2': 'HAPPY D.',
    'دورافيت نمبر 1': 'DURAVIT NO. 1',
    'دي كود': 'D-CODE',
    'إيكو': 'ECHO',
    'دورابلاس': 'DURAPLUS',
    'إميليا': 'EMILIA',
    'جولف': 'GOLF',
    'ديون': 'DUNE',
    'متنوعة': 'OTHERS',
    'إكسسوارات سيراميك': 'ACCESSORIES CERAMIC',
    'إكسسوارات إيزي': 'ACCESSORIES EASY',
    'إكسسوارات كروم': 'ACCESSORIES CHROME',
}

# Get existing subcategories
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-F', '\t',
     '-c', "SELECT row_to_json(t)::text FROM (SELECT id, name FROM subcategories) t"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
existing_subcats = {}
for line in r.stdout.strip().split('\n'):
    line = line.strip()
    if line.startswith('{'):
        try:
            sc = json.loads(line)
            existing_subcats[sc['name']] = sc['id']
        except:
            pass

print(f"Existing subcats: {len(existing_subcats)}")

# Create missing subcategories
cat_ids = {}
for arabic, eng in SERIES_MAP.items():
    if eng in existing_subcats:
        cat_ids[arabic] = existing_subcats[eng]
    else:
        new_id = str(uuid.uuid4())
        safe_name = eng.replace("'", "''")
        subprocess.run([
            'docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
            '-c', f"INSERT INTO subcategories (id, name) VALUES ('{new_id}', E'{safe_name}')"
        ], capture_output=True, text=True, encoding='utf-8', errors='replace')
        cat_ids[arabic] = new_id
        print(f"  Created subcategory: {eng} -> {new_id}")

print(f"Subcategory mapping ready")

# Generate SQL to delete old دروفيت products and import new ones
sql_lines = []

# 1. Delete all existing دروفيت products
sql_lines.append("DELETE FROM products WHERE company = 'دروفيت';")

# 2. Import from Excel
unit = 'قطعة'
now = "NOW()"

for vals in rows:
    section = vals[0]  # القسم
    series_ar = vals[1]  # السلسلة (Arabic)
    name = vals[2]  # المنتج
    code = vals[3]  # الكود (10-digit)
    size = vals[4]  # المقاس
    color = vals[5]  # اللون
    price_str = vals[6]  # السعر
    notes = vals[7]  # ملاحظات
    
    # Get subcategory_id
    subcat_id = cat_ids.get(series_ar, existing_subcats.get('OTHERS', ''))
    if not subcat_id:
        print(f"  WARNING: No subcategory for {series_ar}, skipping")
        continue
    
    # Generate UUID
    prod_id = str(uuid.uuid4())
    
    # Clean price (may contain commas)
    price = 0
    try:
        price = float(price_str.replace(',', ''))
    except:
        price = 0
    
    # Build name: clean Arabic + code maybe in brackets
    prod_name = name
    
    safe_name = prod_name.replace("'", "''")
    safe_size = size.replace("'", "''") if size else ''
    safe_color = color.replace("'", "''") if color else ''
    safe_notes = notes.replace("'", "''") if notes else ''
    
    sql_lines.append(
        f"INSERT INTO products (id, subcategory_id, name, barcode, unit, "
        f"retail_price, wholesale_price, cost_price, company, size, type, material, "
        f"is_active, created_at, updated_at, stock_status) VALUES ("
        f"'{prod_id}', '{subcat_id}', E'{safe_name}', '{code}', '{unit}', "
        f"{price}, {price}, {price}, 'دروفيت', E'{safe_size}', '', '', "
        f"true, {now}, {now}, 'untracked');"
    )

sql = '\n'.join(sql_lines)
print(f"SQL lines: {len(sql_lines)} (1 DELETE + {len(sql_lines)-1} INSERTs)")

# Write SQL file
sql_path = r'C:\eg-co-erp\reimport_drovit.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write(sql)

# Execute
subprocess.run(['docker', 'cp', sql_path, 'eg-co-erp-db-1:/tmp/reimport_drovit.sql'], capture_output=True)
r2 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db', '-f', '/tmp/reimport_drovit.sql'],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(f"Errors: {r2.stderr.count('ERROR') if r2.stderr else 0}")
if r2.stderr:
    # Show first 3 errors
    for line in r2.stderr.split('\n'):
        if 'ERROR' in line:
            print(f"  {line[:120]}")
            break

# Verify
r3 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A',
     '-c', "SELECT count(*) FROM products WHERE company = 'دروفيت'"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(f"دروفيت products in DB: {r3.stdout.strip()}")

# Sample some names
r4 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-c', "SELECT name, barcode FROM products WHERE company = 'دروفيت' ORDER BY random() LIMIT 10"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print("Samples:")
print(r4.stdout)
