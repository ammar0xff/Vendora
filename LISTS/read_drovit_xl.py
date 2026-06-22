import openpyxl, json, sys

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_اسعار_دورافيت_2026_عربي.xlsx', data_only=True)
ws = wb.active

print(f"Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}")

# Print all rows
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 410), values_only=True)):
    vals = [str(v) if v is not None else '' for v in row]
    if any(v for v in vals):
        print(f"  Row {i+1}: {' | '.join(vals)}")
