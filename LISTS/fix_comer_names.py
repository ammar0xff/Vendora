import openpyxl, sys, re
from decimal import Decimal

sys.stdout.reconfigure(encoding='utf-8')

# Read Comer Excel
path = r'C:\eg-co-erp\LISTS\Neisco_Comer_Price_List_2026.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)

# Complete fitting names mapping (all codes)
FITTING_NAMES = {
    # Metric Plain
    'EL50N': 'كوع 90',
    'EY50N': 'كوع 45',
    'SO10N': 'صولة',
    'TE40N': 'تي 90',
    'TR40N': 'تي نقص',
    'ST20N': 'مسلوب',
    'BR00N': 'فلنشة',
    'UN80N': 'يونيون',
    'CA70N': 'غطاء',
    'RP20N': 'نقاص',
    'RB90N': 'جلبة نقص',
    # Metric Reduced
    'RB92N': 'جلبة نقص محول',
    'RB94N': 'جلبة نقص محول سن',
    # BSP Threaded (وصلات سن)
    'EL51N': 'كوع 90 سن',
    'EY51N': 'كوع 45 سن',
    'SO11N': 'صولة سن',
    'TE41N': 'تي 90 سن',
    'PL71N': 'وصلة سن',
    'CA71N': 'طبة سن',
    # Imperial Plain BS (وصلات لصق انجليزي)
    'EL53N': 'كوع 90 انجليزي',
    'EY53N': 'كوع 45 انجليزي',
    'TE43N': 'تي 90 انجليزي',
    'CA73N': 'طبة انجليزي',
    'SO13N': 'صولة انجليزي',
    # Adaptor Series
    'EL52N': 'كوع 90 محول',
    'EL54N': 'كوع 45 محول',
    'TE42N': 'تي محول',
    'TE44N': 'تي محول سن',
    'TR44N': 'تي نقص محول',
    'AD12N': 'محول',
    'AD14N': 'محول سن',
    'SO12N': 'صولة محول',
    'SO14N': 'صولة محول سن',
    'UN82N': 'يونيون محول',
    # Valves
    'BVSL10N': 'محبس كرة',
    'BVDL10N': 'محبس كرة',
    'FV10N': 'محبس رفرف',
    'BVSL11N': 'محبس كرة سن',
    'BVDL11N': 'محبس كرة سن',
    'CVD10N': 'صمام عدم رجوع',
    'CVDC10N': 'صمام عدم رجوع',
    'CVD11N': 'صمام عدم رجوع سن',
    'CVD13N': 'صمام عدم رجوع سن',
    'BUT10N': 'صمام فراشة',
    # Threaded Reducers
    'RE61N': 'ناقص سن',
    # Saddles
    'SA51N': 'سرج',
    'VS51N': 'سرج',
}

# Sheet -> subcategory mapping
SHEET_TYPES = {
    'Metric Plain Fittings (1)': 'كوع 90',
    'Metric Plain Fittings (2)': 'كوع 45',
    'Sockets & Tees Metric': 'جلبة وتي',
    'Stubs Flanges Unions Metric': 'مسلوب وفلنشة',
    'Reduced Metric Fittings': 'نقاص وتي نقص',
    'BSP Threaded Fittings': 'وصلات سن',
    'Saddles & Bushings Threaded': 'سرج وجلبة سن',
    'Imperial Plain Fittings BS': 'وصلات لصق انجليزي',
    'Adaptor Series Plain-Threaded': 'محولات',
    'UPVC Ball Valves': 'محابس',
    'Check & Butterfly Valves': 'صمامات',
}

lines = []
upd = 0

for sname in wb.sheetnames:
    ws = wb[sname]
    subcat = SHEET_TYPES.get(sname, sname[:20])
    
    for row in ws.iter_rows(min_row=4, values_only=True):
        code, size, pack, box, price = row if len(row) >= 5 else (row[0] if row else None, None, None, None, None)
        if not code or not size:
            continue
        code_str = str(code).strip()
        size_str = str(size).strip()
        if not price or str(price).strip() in ('', '-'):
            continue
        price_val = Decimal(str(price))
        
        # Get fitting name from code prefix
        prefix = None
        for key, name in FITTING_NAMES.items():
            if code_str.startswith(key):
                prefix = name
                break
        if not prefix:
            continue  # skip if still unknown
        
        product_name = f'{prefix} {size_str}'
        product_name_esc = product_name.replace("'", "''")
        subcat_esc = subcat.replace("'", "''")
        size_esc = size_str.replace("'", "''")
        
        lines.append(f"""UPDATE products SET name = '{product_name_esc}'
WHERE company = 'كومر'
  AND subcategory_id = (SELECT id FROM subcategories WHERE name = '{subcat_esc}'
                        AND category_id = (SELECT id FROM categories WHERE name = 'كومر'))
  AND retail_price = {price_val}
  AND size = '{size_esc}'
  AND name LIKE 'كومر%';""")
        upd += 1

output = r'C:\eg-co-erp\fix_comer_names.sql'
with open(output, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f"Wrote {upd} UPDATE statements")
