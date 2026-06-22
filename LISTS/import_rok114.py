import openpyxl
from decimal import Decimal

path = r'C:\eg-co-erp\LISTS\قائمة_أسعار_روك_١١٤.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb[wb.sheetnames[0]]

CATEGORY = 'روك 114'
lines = []
seen_subs = set()

def sql(s):
    lines.append(s)

sql(f"""INSERT INTO categories (id, name) VALUES (gen_random_uuid(), '{CATEGORY}') ON CONFLICT (name) DO NOTHING;""")

def ensure_subcat(name):
    if name not in seen_subs:
        sql(f"""INSERT INTO subcategories (id, category_id, name) SELECT gen_random_uuid(), (SELECT id FROM categories WHERE name = '{CATEGORY}'), '{name}' WHERE NOT EXISTS (SELECT 1 FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = '{name}');""")
        seen_subs.add(name)

def parse_price(v):
    if v is None or v == '-':
        return None
    if isinstance(v, str):
        try:
            return float(v.replace(',', '.'))
        except ValueError:
            return None
    if isinstance(v, (int, float)):
        return float(v)
    return None

def product_val(pname, subcat, price, size_val=None):
    sz = size_val if size_val else 'NULL'
    sql(f"""INSERT INTO products (id, subcategory_id, name, unit, retail_price, wholesale_price, cost_price, is_active, company, size, stock_status) SELECT gen_random_uuid(), (SELECT id FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = '{subcat}'), '{pname}', 'عدد', {Decimal(str(price))}, 0, 0, true, '{CATEGORY}', {sz if sz == 'NULL' else f"'{sz}'"}, 'untracked' WHERE NOT EXISTS (SELECT 1 FROM products WHERE name = '{pname}' AND company = '{CATEGORY}');""")

# ============ Section 1 (R4-R12): كوع عاده, كوع باب, كوع مفتوح, مشترك واي, جلبة لحام ============
ensure_subcat('كوع عاده 114')
ensure_subcat('كوع باب 114')

for r in range(5, 13):
    row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    size_val = str(row[0]).strip() if row[0] else None
    if not size_val:
        continue
    
    # كوع عاده (col B = idx 1)
    p = parse_price(row[1])
    if p:
        product_val(f'كوع عاده {size_val}', 'كوع عاده 114', p, size_val)
    
    # كوع باب (col C = idx 2)
    p = parse_price(row[2])
    if p:
        product_val(f'كوع باب {size_val}', 'كوع باب 114', p, size_val)

# كوع مفتوح (cols D(idx3)=size, E(idx4)=price)
ensure_subcat('كوع مفتوح 114')
for r in range(5, 13):
    row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    sz = str(row[3]).strip() if row[3] else None
    if sz and sz != '-':
        p = parse_price(row[4])
        if p:
            product_val(f'كوع مفتوح {sz}', 'كوع مفتوح 114', p, sz)

# مشترك واي (cols F(idx5)=size, G(idx6)=price)
ensure_subcat('مشترك واي 114')
for r in range(5, 13):
    row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    sz = str(row[5]).strip() if row[5] else None
    if sz and sz != '-':
        p = parse_price(row[6])
        if p:
            product_val(f'مشترك واي {sz}', 'مشترك واي 114', p, sz)

# جلبة لحام (cols H(idx7)=size, I(idx8)=price)
ensure_subcat('جلبة لحام 114')
for r in range(5, 13):
    row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    sz = str(row[7]).strip() if row[7] else None
    if sz and sz != '-':
        p = parse_price(row[8])
        if p:
            product_val(f'جلبة لحام {sz}', 'جلبة لحام 114', p, sz)

# ============ Section 2 (R15-R21): مشترك عاده, مشترك باب ============
ensure_subcat('مشترك عاده 114')
ensure_subcat('مشترك باب 114')

for r in range(16, 22):
    row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    size_val = str(row[0]).strip() if row[0] else None
    if not size_val:
        continue
    p = parse_price(row[1])
    if p:
        product_val(f'مشترك عاده {size_val}', 'مشترك عاده 114', p, size_val)
    p = parse_price(row[2])
    if p:
        product_val(f'مشترك باب {size_val}', 'مشترك باب 114', p, size_val)

# ============ Section 3 (R24-R28): بيبة عالية سوبر, بيبه اسم, بوش ============
ensure_subcat('بيبة عالية سوبر 114')
ensure_subcat('بيبه اسم 114')
ensure_subcat('بوش مسلوب 114')

for r in range(25, 29):
    row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    
    # بيبة عالية سوبر (cols A-B)
    sz = str(row[0]).strip() if row[0] else None
    if sz and sz != '-':
        p = parse_price(row[1])
        if p and isinstance(p, (int, float)):
            product_val(f'بيبة عالية سوبر {sz}', 'بيبة عالية سوبر 114', p, sz)
    
    # بيبه اسم (cols C-D)
    sz = str(row[2]).strip() if row[2] else None
    if sz and sz != '-':
        p = parse_price(row[3])
        if p and isinstance(p, (int, float)):
            product_val(f'بيبه اسم {sz}', 'بيبه اسم 114', p, sz)
    
    # بوش (cols E-F)
    sz = str(row[4]).strip() if row[4] else None
    if sz and sz != '-':
        p = parse_price(row[5])
        if p and isinstance(p, (int, float)):
            product_val(f'بوش {sz}', 'بوش مسلوب 114', p, sz)

# ============ Section 4 (R31-R34): مشترك مسلوب, صليبه, جلبة إصلاح, هواية ============
ensure_subcat('مشترك مسلوب 114')
ensure_subcat('صليبه 114')
ensure_subcat('جلبة إصلاح 114')
ensure_subcat('هواية 114')

for r in range(32, 35):
    row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    
    # مشترك مسلوب (cols A(idx0)=size, B(idx1)=سعر عاده, C(idx2)=باب)
    sz = str(row[0]).strip() if row[0] else None
    if sz and sz != '-':
        p = parse_price(row[1])
        if p:
            product_val(f'مشترك مسلوب عاده {sz}', 'مشترك مسلوب 114', p, sz)
        p = parse_price(row[2])
        if p:
            product_val(f'مشترك مسلوب باب {sz}', 'مشترك مسلوب 114', p, sz)
    
    # صليبه (cols D(idx3)=size, E(idx4)=price)
    sz = str(row[3]).strip() if row[3] else None
    if sz and sz != '-':
        p = parse_price(row[4])
        if p:
            product_val(f'صليبه {sz}', 'صليبه 114', p, sz)
    
    # جلبة إصلاح (cols F(idx5)=size, G(idx6)=price)
    sz = str(row[5]).strip() if row[5] else None
    if sz and sz != '-':
        p = parse_price(row[6])
        if p:
            product_val(f'جلبة إصلاح {sz}', 'جلبة إصلاح 114', p, sz)
    
    # هواية (cols H(idx7)=size, I(idx8)=price)
    sz = str(row[7]).strip() if row[7] else None
    if sz and sz != '-':
        p = parse_price(row[8])
        if p:
            product_val(f'هواية {sz}', 'هواية 114', p, sz)

# ============ Section 5 (R37-R48): أصناف متنوعة ============
ensure_subcat('أصناف متنوعة 114')

for r in range(38, 49):
    row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    item_name = str(row[0]).strip() if row[0] else None
    item_spec = str(row[1]).strip() if row[1] and str(row[1]).strip() != '-' else ''
    price = parse_price(row[2])
    
    if item_name and price:
        if item_spec:
            pname = f'{item_name} {item_spec}'
        else:
            pname = item_name
        product_val(pname, 'أصناف متنوعة 114', price, item_spec if item_spec else None)

# Footer
sql(f"""
SELECT 'categories' AS tbl, count(*) FROM categories WHERE name = '{CATEGORY}'
UNION ALL SELECT 'subcategories', count(*) FROM subcategories s JOIN categories c ON s.category_id = c.id WHERE c.name = '{CATEGORY}'
UNION ALL SELECT 'products', count(*) FROM products p WHERE p.company = '{CATEGORY}';
""")

output = r'C:\eg-co-erp\import_rok114.sql'
with open(output, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
print(f"Rok 114: SQL written to {output}")
