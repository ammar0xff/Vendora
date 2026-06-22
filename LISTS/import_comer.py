import openpyxl, subprocess, uuid
from decimal import Decimal

def esc(v):
    if v is None:
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"

lines = []
def sql(s):
    lines.append(s)

# Sheet -> Arabic product type mapping
SHEET_TYPES = {
    'Metric Plain Fittings (1)': 'كوع 90',
    'Metric Plain Fittings (2)': 'كوع 45',
    'Sockets & Tees Metric': 'جلبة وتي',
    'Stubs Flanges Unions Metric': 'مسلوب وفلنشة',
    'Reduced Metric Fittings': 'نقاص وتي نقص',
    'BSP Threaded Fittings': 'وصلات سن',
    'Saddles & Bushings Threaded': 'سرج وجلبة سن',
    'Imperial Plain Fittings BS': 'وصلات لصق انجليزي',
    'Adaptor Series Plain-Threaded': 'محولات',
    'UPVC Ball Valves': 'محابس',
    'Check & Butterfly Valves': 'صمامات',
}

# Fitting type keywords for building Arabic names
FITTING_NAMES = {
    'EL50N': 'كوع 90',
    'EY50N': 'كوع 45',
    'EL51N': 'كوع 90 سن',
    'EY51N': 'كوع 45 سن',
    'EL53N': 'كوع 90 انجليزي',
    'EY53N': 'كوع 45 انجليزي',
    'EL52N': 'كوع 90 محول',
    'EL54N': 'كوع 45 محول',
    'SO10N': 'صولة',
    'SO12N': 'صولة محول',
    'SO14N': 'صولة محول سن',
    'TE40N': 'تي 90',
    'TE42N': 'تي محول',
    'TE44N': 'تي محول سن',
    'TR40N': 'تي نقص',
    'TR44N': 'تي نقص محول',
    'ST20N': 'مسلوب',
    'BR00N': 'فلنشة',
    'UN80N': 'يونيون',
    'UN82N': 'يونيون محول',
    'CA70N': 'غطاء',
    'RP20N': 'نقاص',
    'RB90N': 'جلبة نقص',
    'RB92N': 'جلبة نقص محول',
    'RB94N': 'جلبة نقص محول سن',
    'RE61N': 'ناقص',
    'SA51N': 'سرج',
    'VS51N': 'سرج',
    'AD12N': 'محول',
    'AD14N': 'محول سن',
    'BVSL10N': 'محبس كرة',
    'BVDL10N': 'محبس كرة',
    'FV10N': 'محبس رفرف',
    'CVD10N': 'صمام عدم رجوع',
    'CVDC10N': 'صمام عدم رجوع',
    'BUT10N': 'صمام فراشة',
}

path = r'C:\eg-co-erp\LISTS\done\Neisco_Comer_Price_List_2026.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)

total_products = 0

for sname in wb.sheetnames:
    ws = wb[sname]
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    subcat_name = SHEET_TYPES.get(sname, sname[:20])
    
    subcat_sql = f"""INSERT INTO subcategories (id, category_id, name) SELECT gen_random_uuid(), (SELECT id FROM categories WHERE name = 'كومر'), {esc(subcat_name)} WHERE NOT EXISTS (SELECT 1 FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = 'كومر') AND name = {esc(subcat_name)});"""
    sql(subcat_sql)
    
    for row in rows:
        code, size, pack, box, price = row if len(row) >= 5 else (row[0] if row else None, None, None, None, None)
        if not code or not size:
            continue
        code_str = str(code).strip()
        size_str = str(size).strip()
        
        # Build Arabic name: fitting type + size
        prefix = 'كومر'
        for key, name in FITTING_NAMES.items():
            if code_str.startswith(key):
                prefix = name
                break
        
        product_name = f'{prefix} {size_str} [{code_str}]'
        price_val = Decimal(str(price)) if price is not None and str(price).strip() != '' and str(price).strip() != '-' else Decimal('0')
        
        sql(f"""INSERT INTO products (id, subcategory_id, name, unit, retail_price, wholesale_price, cost_price, is_active, company, size, stock_status) SELECT gen_random_uuid(), (SELECT id FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = 'كومر') AND name = {esc(subcat_name)}), {esc(product_name)}, 'عدد', {price_val}, 0, 0, true, 'كومر', {esc(size_str)}, 'untracked' WHERE NOT EXISTS (SELECT 1 FROM products WHERE name = {esc(product_name)} AND company = 'كومر');""")
        total_products += 1

# Wrap in category creation
header = f"""
INSERT INTO categories (id, name) VALUES (gen_random_uuid(), 'كومر') ON CONFLICT (name) DO NOTHING;
"""
lines.insert(0, header)

# Add verification
lines.append(f"""
SELECT 'categories' AS tbl, count(*) FROM categories WHERE name = 'كومر'
UNION ALL SELECT 'subcategories', count(*) FROM subcategories s JOIN categories c ON s.category_id = c.id WHERE c.name = 'كومر'
UNION ALL SELECT 'products', count(*) FROM products p WHERE p.company = 'كومر';
""")

output = r'C:\eg-co-erp\import_comer.sql'
with open(output, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f"Comer: {total_products} products, {len(SHEET_TYPES)} subcategories")
print(f"SQL written to {output}")
