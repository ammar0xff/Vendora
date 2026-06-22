import openpyxl, datetime, re, os
from decimal import Decimal

DONE_DIR = os.path.join(os.path.dirname(__file__), 'done')

def parse_price(val):
    if val is None or val == '-':
        return None
    if isinstance(val, datetime.datetime):
        return round(val.month + val.day / 100, 2)
    if isinstance(val, str):
        try:
            return float(val.replace(',', '.'))
        except ValueError:
            return None
    if isinstance(val, (int, float)):
        return float(val)
    return None

lines = []
def sql(s):
    lines.append(s)

def parse_combined(val):
    if val is None or val == '-' or val == '':
        return None, None
    s = str(val).strip()
    m = re.match(r'^(.+?)\s*\(([\d.]+)\)$', s)
    if m:
        size_part = m.group(1).strip()
        price_val = float(m.group(2))
        return size_part, price_val
    return None, None

# ───── روك بولي ─────
path_poly = os.path.join(DONE_DIR, 'قائمة أسعار روك بولي 2024 - جداول منظمة.xlsx')
wb = openpyxl.load_workbook(path_poly, data_only=True)
ws = wb[wb.sheetnames[0]]

CATEGORY = 'روك بولي'
MARKUP = 1.75  # +75%

def update_poly_product(pname, price):
    if not price:
        return
    rp = Decimal(str(round(price * MARKUP, 2)))
    sql(f"""UPDATE products SET retail_price = {rp} WHERE name = '{pname}' AND company = '{CATEGORY}' AND retail_price = 0;""")
    sql(f"""UPDATE products SET retail_price = {rp} WHERE name = '{pname}' AND company = '{CATEGORY}' AND retail_price != 0 AND retail_price < {rp};""")

# Sections 1-3: جلب لحام, تي لحام, كوع لحام
sections_1_3 = [('جلب لحام', 'جلبة لحام', 1, 2), ('تي لحام', 'تي لحام', 3, 4), ('كوع لحام', 'كوع لحام 45', 5, 6)]
for subcat, prod_name, _, price_col in sections_1_3:
    for r in range(5, 10):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if not row[0]:
            continue
        size_val = str(row[0]).strip()
        price = parse_price(row[price_col])
        update_poly_product(f'{prod_name} {size_val}', price)

# Sections 4-6: جلب أنثى, تي بسن, كوع بسن
sections_4_6 = [('جلب أنثى', 'جلبة أنثى', 1, 2), ('تي بسن', 'تي بسن', 3, 4), ('كوع بسن', 'كوع بسن', 5, 6)]
for subcat, prod_name, _, price_col in sections_4_6:
    for r in range(13, 19):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if not row[0]:
            continue
        size_val = str(row[0]).strip()
        price = parse_price(row[price_col])
        update_poly_product(f'{prod_name} {size_val}', price)

# كوع مفتوح وطبة اختبار (rows 22-24)
for r in range(22, 25):
    row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    if not row[0]:
        continue
    raw_name = str(row[0]).strip()
    price = parse_price(row[2])
    update_poly_product(raw_name, price)

# Sections: تي مسلوب, جلب مسلوب, جلب ذكر (rows 28-37)
sections_red = [('تي مسلوب', 'تي مسلوب', 1, 2), ('جلب مسلوب', 'جلبة مسلوب', 3, 4), ('جلب ذكر', 'جلبة ذكر', 5, 6)]
for subcat, prod_name, _, price_col in sections_red:
    for r in range(28, 38):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if not row[0]:
            continue
        size_val = str(row[0]).strip()
        price = parse_price(row[price_col])
        update_poly_product(f'{prod_name} {size_val}', price)

# كرنك حقن (rows 41-42)
for r in range(41, 43):
    row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    if not row[0]:
        continue
    raw_name = str(row[0]).strip()
    price = parse_price(row[1])
    update_poly_product(raw_name, price)

# طبة كاب, محبس دفع (rows 46-50)
sections_cap = [('طبة كاب', 'طبة كاب', 1, 2), ('محبس دفع', 'محبس دفع', 3, 4)]
for subcat, prod_name, _, price_col in sections_cap:
    for r in range(46, 51):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if not row[0]:
            continue
        size_val = str(row[0]).strip()
        price = parse_price(row[price_col])
        if price and price > 0:
            update_poly_product(f'{prod_name} {size_val}', price)

# ───── روك 110 ─────
path_110 = os.path.join(DONE_DIR, 'قائمة_أسعار_روك_110_المطورة.xlsx')
wb110 = openpyxl.load_workbook(path_110, data_only=True)

CAT110 = 'روك 110'

def update_110_product(pname, price, cat=CAT110):
    if not price:
        return
    rp = Decimal(str(round(price, 2)))
    sql(f"""UPDATE products SET retail_price = {rp} WHERE name = '{pname}' AND company = '{cat}' AND retail_price = 0;""")
    sql(f"""UPDATE products SET retail_price = {rp} WHERE name = '{pname}' AND company = '{cat}' AND retail_price != 0 AND retail_price < {rp};""")

# Sheet 1
ws1 = wb110[wb110.sheetnames[0]]

# الأكواع (R4-R11)
subcat_names_1 = ['كوع عادي 110', 'كوع باب 110', 'كوع مفتوح 110']
col_indices_1 = [1, 2, 3]
for sc_name, col_idx in zip(subcat_names_1, col_indices_1):
    for r in range(5, 12):
        row = list(ws1.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if not row[0]:
            continue
        size_val = str(row[0]).strip()
        price = parse_price(row[col_idx])
        if price:
            pname = f'{sc_name.replace(" 110", "")} {size_val}'
            update_110_product(pname, price)

# المشتركات (R14-R21) — Col A=size, B=مشترك واي (plain), C=مشترك مسلوب (combined)
for r in range(15, 22):
    row = list(ws1.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    if not row[0]:
        continue
    size_val = str(row[0]).strip()
    # مشترك واي - plain price
    price = parse_price(row[1])
    if price:
        update_110_product(f'مشترك واي {size_val}', price)
    # مشترك مسلوب - combined format
    sub_size, sub_price = parse_combined(row[2])
    if sub_price:
        update_110_product(f'مشترك مسلوب {sub_size}', sub_price)

# جلب لحام وبوش (R24-R31) — Col A=size, B=جلبة لحام, C=بوش (combined), D=تابع البوش (combined)
for r in range(25, 32):
    row = list(ws1.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    if not row[0]:
        continue
    size_val = str(row[0]).strip()
    # جلبة لحام - plain price
    price = parse_price(row[1])
    if price:
        update_110_product(f'جلبة لحام {size_val}', price)
    # بوش - combined format
    sub_size, sub_price = parse_combined(row[2])
    if sub_price:
        update_110_product(f'بوش {sub_size}', sub_price)
    # تابع البوش - combined format
    sub_size, sub_price = parse_combined(row[3])
    if sub_price:
        update_110_product(f'تابع بوش {sub_size}', sub_price)

# Sheet 2: أصناف متنوعة
ws2 = wb110[wb110.sheetnames[1]]
for r in range(4, 33):
    row = list(ws2.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    if not row[0]:
        continue
    pname = str(row[0]).strip()
    price = parse_price(row[1])
    if price:
        update_110_product(pname, price)

# ───── روك 114 ─────
path_114 = os.path.join(DONE_DIR, 'قائمة_أسعار_روك_١١٤.xlsx')
wb114 = openpyxl.load_workbook(path_114, data_only=True)
ws114 = wb114[wb114.sheetnames[0]]

CAT114 = 'روك 114'

def update_114_product(pname, price, size_val=None):
    if not price:
        return
    rp = Decimal(str(round(price, 2)))
    sql(f"""UPDATE products SET retail_price = {rp} WHERE name = '{pname}' AND company = '{CAT114}' AND retail_price = 0;""")
    sql(f"""UPDATE products SET retail_price = {rp} WHERE name = '{pname}' AND company = '{CAT114}' AND retail_price != 0 AND retail_price < {rp};""")

# Section 1 (R4-R12): كوع عاده, كوع باب, كوع مفتوح, مشترك واي, جلبة لحام
for r in range(5, 13):
    row = list(ws114.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    size_val = str(row[0]).strip() if row[0] else None
    if not size_val:
        continue
    p = parse_price(row[1])
    if p:
        update_114_product(f'كوع عاده {size_val}', p)
    p = parse_price(row[2])
    if p:
        update_114_product(f'كوع باب {size_val}', p)

for r in range(5, 13):
    row = list(ws114.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    sz = str(row[3]).strip() if row[3] else None
    if sz and sz != '-':
        p = parse_price(row[4])
        if p:
            update_114_product(f'كوع مفتوح {sz}', p)
    sz = str(row[5]).strip() if row[5] else None
    if sz and sz != '-':
        p = parse_price(row[6])
        if p:
            update_114_product(f'مشترك واي {sz}', p)
    sz = str(row[7]).strip() if row[7] else None
    if sz and sz != '-':
        p = parse_price(row[8])
        if p:
            update_114_product(f'جلبة لحام {sz}', p)

# Section 2 (R15-R21): مشترك عاده, مشترك باب
for r in range(16, 22):
    row = list(ws114.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    size_val = str(row[0]).strip() if row[0] else None
    if not size_val:
        continue
    p = parse_price(row[1])
    if p:
        update_114_product(f'مشترك عاده {size_val}', p)
    p = parse_price(row[2])
    if p:
        update_114_product(f'مشترك باب {size_val}', p)

# Section 3 (R24-R28)
for r in range(25, 29):
    row = list(ws114.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    sz = str(row[0]).strip() if row[0] else None
    if sz and sz != '-':
        p = parse_price(row[1])
        if p:
            update_114_product(f'بيبة عالية سوبر {sz}', p)
    sz = str(row[2]).strip() if row[2] else None
    if sz and sz != '-':
        p = parse_price(row[3])
        if p:
            update_114_product(f'بيبه اسم {sz}', p)
    sz = str(row[4]).strip() if row[4] else None
    if sz and sz != '-':
        p = parse_price(row[5])
        if p:
            update_114_product(f'بوش {sz}', p)

# Section 4 (R31-R34)
for r in range(32, 35):
    row = list(ws114.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    sz = str(row[0]).strip() if row[0] else None
    if sz and sz != '-':
        p = parse_price(row[1])
        if p:
            update_114_product(f'مشترك مسلوب عاده {sz}', p)
        p = parse_price(row[2])
        if p:
            update_114_product(f'مشترك مسلوب باب {sz}', p)
    sz = str(row[3]).strip() if row[3] else None
    if sz and sz != '-':
        p = parse_price(row[4])
        if p:
            update_114_product(f'صليبه {sz}', p)
    sz = str(row[5]).strip() if row[5] else None
    if sz and sz != '-':
        p = parse_price(row[6])
        if p:
            update_114_product(f'جلبة إصلاح {sz}', p)
    sz = str(row[7]).strip() if row[7] else None
    if sz and sz != '-':
        p = parse_price(row[8])
        if p:
            update_114_product(f'هواية {sz}', p)

# Section 5 (R37-R48): أصناف متنوعة
for r in range(38, 49):
    row = list(ws114.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    item_name = str(row[0]).strip() if row[0] else None
    item_spec = str(row[1]).strip() if row[1] and str(row[1]).strip() != '-' else ''
    price = parse_price(row[2])
    if item_name and price:
        if item_spec:
            pname = f'{item_name} {item_spec}'
        else:
            pname = item_name
        update_114_product(pname, price)

sql("SELECT 'Done' as status;")

output = os.path.join(os.path.dirname(__file__), 'update_rock_prices.sql')
with open(output, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f"Generated {len(lines)} SQL statements → {output}")
