import openpyxl, subprocess, uuid

DROVIT_CATEGORY_ID = '6d3ebf46-0d63-45a7-b271-74ce542d732b'
wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_اسعار_دورافيت_2026_عربي.xlsx', data_only=True)

# Skip sheet 0 "جميع المنتجات" - use sheets 1-8 as subcategories
page_names = wb.sheetnames[1:]  # All except "جميع المنتجات"
print(f"Subcategory sheets ({len(page_names)}):")
for p in page_names:
    print(f"  {p}")

sql_parts = []
sql_parts.append("DELETE FROM products WHERE company = 'دروفيت';")
# Also delete the old subcategories that were based on series
sql_parts.append(f"DELETE FROM subcategories WHERE category_id = '{DROVIT_CATEGORY_ID}';")

now = "NOW()"
insert_count = 0

for i, sn in enumerate(page_names):
    ws = wb[sn]
    subcat_id = str(uuid.uuid4())
    safe_name = sn.replace("'", "''")
    sql_parts.append(
        f"INSERT INTO subcategories (id, category_id, name, created_at) "
        f"VALUES ('{subcat_id}', '{DROVIT_CATEGORY_ID}', E'{safe_name}', {now});"
    )
    
    # Read rows from this sheet (skip header row 1)
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        # vals[0]=السلسلة, vals[1]=المنتج, vals[2]=رقم الموديل, vals[3]=الأبعاد, vals[4]=اللون, vals[5]=السعر
        name = vals[1]
        code = vals[2]
        if not name or not code:
            continue
        name = str(name).strip()
        code = str(code).strip()
        
        size = str(vals[3]).strip() if vals[3] else ''
        color = str(vals[4]).strip() if vals[4] else ''
        
        price_val = 0
        try:
            pv = vals[5]
            if pv is not None:
                price_val = float(str(pv).replace(',', ''))
        except:
            pass
        
        prod_id = str(uuid.uuid4())
        safe_name_prod = name.replace("'", "''")
        safe_size = size.replace("'", "''")
        
        sql_parts.append(
            f"INSERT INTO products (id, subcategory_id, name, barcode, unit, "
            f"retail_price, wholesale_price, cost_price, company, size, type, material, "
            f"is_active, created_at, updated_at, stock_status) VALUES ("
            f"'{prod_id}', '{subcat_id}', E'{safe_name_prod}', '{code}', 'قطعة', "
            f"{price_val}, {price_val}, {price_val}, 'دروفيت', E'{safe_size}', '', '', "
            f"true, {now}, {now}, 'untracked');"
        )
        insert_count += 1

print(f"Products to insert: {insert_count}")

sql_path = r'C:\eg-co-erp\reimport_by_page.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_parts))

subprocess.run(['docker', 'cp', sql_path, 'eg-co-erp-db-1:/tmp/reimport_by_page.sql'], capture_output=True)
r = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db', '-f', '/tmp/reimport_by_page.sql'],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
errors = r.stderr.count('ERROR') if r.stderr else 0
print(f"Errors: {errors}")
if r.stderr and errors > 0:
    for l in r.stderr.split('\n')[:10]:
        if 'ERROR' in l:
            print(f"  {l[:150]}")

r2 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A', '-c', "SELECT count(*) FROM products WHERE company = 'دروفيت'"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(f"دروفيت in DB: {r2.stdout.strip()}")

r3 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-c', "SELECT s.name, count(p.id) FROM subcategories s LEFT JOIN products p ON p.subcategory_id = s.id WHERE s.category_id = '6d3ebf46-0d63-45a7-b271-74ce542d732b' GROUP BY s.name ORDER BY s.name"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
# Write to file
with open('C:\\eg-co-erp\\drovit_by_page.txt', 'w', encoding='utf-8') as f:
    f.write(r3.stdout)
print("Result written to C:\\eg-co-erp\\drovit_by_page.txt")
