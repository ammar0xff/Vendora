import openpyxl, datetime, re, os, asyncio
from decimal import Decimal
from sqlalchemy import text
from app.db.base import AsyncSessionLocal

DONE_DIR = '/tmp'

def parse_price(val):
    if val is None or val == '-':
        return None
    if isinstance(val, datetime.datetime):
        return round(val.month + val.day / 100, 2)
    if isinstance(val, str):
        try:
            return float(val.replace(',', '.'))
        except ValueError:
            return None
    if isinstance(val, (int, float)):
        return float(val)
    return None

def parse_combined(val):
    if val is None or val == '-' or val == '':
        return None, None
    s = str(val).strip()
    m = re.match(r'^(.+?)\s*\(([\d.]+)\)$', s)
    if m:
        return m.group(1).strip(), float(m.group(2))
    return None, None

total_updated = 0

async def update_price(name, company, price):
    global total_updated
    if not price:
        return
    async with AsyncSessionLocal() as db:
        rp = Decimal(str(round(price, 2)))
        # Only update if new price is higher (or current is 0/cheap)
        result = await db.execute(
            text("UPDATE products SET retail_price = :rp WHERE name = :name AND company = :company AND (retail_price = 0 OR retail_price < :rp)"),
            {"rp": rp, "name": name, "company": company}
        )
        if result.rowcount:
            total_updated += 1

async def process_poly():
    path = os.path.join(DONE_DIR, 'قائمة أسعار روك بولي 2024 - جداول منظمة.xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    COMPANY = 'روك بولي'
    MARKUP = 1.75

    async def upd(name, price):
        if price:
            await update_price(name, COMPANY, price * MARKUP)

    sections_1_3 = [('جلبة لحام', 1, 2), ('تي لحام', 3, 4), ('كوع لحام 45', 5, 6)]
    for prod_name, _, price_col in sections_1_3:
        for r in range(5, 10):
            row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
            if not row[0]:
                continue
            sz = str(row[0]).strip()
            await upd(f'{prod_name} {sz}', parse_price(row[price_col]))

    sections_4_6 = [('جلبة أنثى', 1, 2), ('تي بسن', 3, 4), ('كوع بسن', 5, 6)]
    for prod_name, _, price_col in sections_4_6:
        for r in range(13, 19):
            row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
            if not row[0]:
                continue
            sz = str(row[0]).strip()
            await upd(f'{prod_name} {sz}', parse_price(row[price_col]))

    for r in range(22, 25):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if not row[0]:
            continue
        await upd(str(row[0]).strip(), parse_price(row[2]))

    sections_red = [('تي مسلوب', 1, 2), ('جلبة مسلوب', 3, 4), ('جلبة ذكر', 5, 6)]
    for prod_name, _, price_col in sections_red:
        for r in range(28, 38):
            row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
            if not row[0]:
                continue
            sz = str(row[0]).strip()
            await upd(f'{prod_name} {sz}', parse_price(row[price_col]))

    for r in range(41, 43):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if not row[0]:
            continue
        await upd(str(row[0]).strip(), parse_price(row[1]))

    for prod_name, _, price_col in [('طبة كاب', 1, 2), ('محبس دفع', 3, 4)]:
        for r in range(46, 51):
            row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
            if not row[0]:
                continue
            sz = str(row[0]).strip()
            price = parse_price(row[price_col])
            if price and price > 0:
                await upd(f'{prod_name} {sz}', price)

async def process_110():
    path = os.path.join(DONE_DIR, 'قائمة_أسعار_روك_110_المطورة.xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    ws1 = wb[wb.sheetnames[0]]
    COMPANY = 'روك 110'

    async def upd(name, price):
        await update_price(name, COMPANY, price)

    for sc_name, col_idx in [('كوع عادي', 1), ('كوع باب', 2), ('كوع مفتوح', 3)]:
        for r in range(5, 12):
            row = list(ws1.iter_rows(min_row=r, max_row=r, values_only=True))[0]
            if not row[0]:
                continue
            sz = str(row[0]).strip()
            p = parse_price(row[col_idx])
            if p:
                await upd(f'{sc_name} {sz}', p)

    for r in range(15, 22):
        row = list(ws1.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if not row[0]:
            continue
        sz = str(row[0]).strip()
        p = parse_price(row[1])
        if p:
            await upd(f'مشترك واي {sz}', p)
        sub_sz, sub_p = parse_combined(row[2])
        if sub_p:
            await upd(f'مشترك مسلوب {sub_sz}', sub_p)

    for r in range(25, 32):
        row = list(ws1.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if not row[0]:
            continue
        sz = str(row[0]).strip()
        p = parse_price(row[1])
        if p:
            await upd(f'جلبة لحام {sz}', p)
        sub_sz, sub_p = parse_combined(row[2])
        if sub_p:
            await upd(f'بوش {sub_sz}', sub_p)
        sub_sz, sub_p = parse_combined(row[3])
        if sub_p:
            await upd(f'تابع بوش {sub_sz}', sub_p)

    ws2 = wb[wb.sheetnames[1]]
    for r in range(4, 33):
        row = list(ws2.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if not row[0]:
            continue
        p = parse_price(row[1])
        if p:
            await upd(str(row[0]).strip(), p)

async def process_114():
    path = os.path.join(DONE_DIR, 'قائمة_أسعار_روك_١١٤.xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    COMPANY = 'روك 114'

    async def upd(name, price):
        await update_price(name, COMPANY, price)

    for r in range(5, 13):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        sz = str(row[0]).strip() if row[0] else None
        if not sz:
            continue
        p = parse_price(row[1])
        if p:
            await upd(f'كوع عاده {sz}', p)
        p = parse_price(row[2])
        if p:
            await upd(f'كوع باب {sz}', p)

    for r in range(5, 13):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        for col_sz, col_pr, pname in [(3, 4, 'كوع مفتوح'), (5, 6, 'مشترك واي'), (7, 8, 'جلبة لحام')]:
            sz = str(row[col_sz]).strip() if row[col_sz] else None
            if sz and sz != '-':
                p = parse_price(row[col_pr])
                if p:
                    await upd(f'{pname} {sz}', p)

    for r in range(16, 22):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        sz = str(row[0]).strip() if row[0] else None
        if not sz:
            continue
        p = parse_price(row[1])
        if p:
            await upd(f'مشترك عاده {sz}', p)
        p = parse_price(row[2])
        if p:
            await upd(f'مشترك باب {sz}', p)

    for r in range(25, 29):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        for col_sz, col_pr, pname in [(0, 1, 'بيبة عالية سوبر'), (2, 3, 'بيبه اسم'), (4, 5, 'بوش')]:
            sz = str(row[col_sz]).strip() if row[col_sz] else None
            if sz and sz != '-':
                p = parse_price(row[col_pr])
                if p:
                    await upd(f'{pname} {sz}', p)

    for r in range(32, 35):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        sz = str(row[0]).strip() if row[0] else None
        if sz and sz != '-':
            p = parse_price(row[1])
            if p:
                await upd(f'مشترك مسلوب عاده {sz}', p)
            p = parse_price(row[2])
            if p:
                await upd(f'مشترك مسلوب باب {sz}', p)
        for col_sz, col_pr, pname in [(3, 4, 'صليبه'), (5, 6, 'جلبة إصلاح'), (7, 8, 'هواية')]:
            sz = str(row[col_sz]).strip() if row[col_sz] else None
            if sz and sz != '-':
                p = parse_price(row[col_pr])
                if p:
                    await upd(f'{pname} {sz}', p)

    for r in range(38, 49):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        item_name = str(row[0]).strip() if row[0] else None
        item_spec = str(row[1]).strip() if row[1] and str(row[1]).strip() != '-' else ''
        price = parse_price(row[2])
        if item_name and price:
            pname = f'{item_name} {item_spec}' if item_spec else item_name
            await upd(pname, price)

async def main():
    await process_poly()
    await process_110()
    await process_114()
    print(f'Total updated: {total_updated}')

asyncio.run(main())
