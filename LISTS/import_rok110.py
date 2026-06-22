import openpyxl, re
from decimal import Decimal

path = r'C:\eg-co-erp\LISTS\قائمة_أسعار_روك_110_المطورة.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)

CATEGORY = 'روك 110'
lines = []
seen_subs = set()

def sql(s):
    lines.append(s)

sql(f"""INSERT INTO categories (id, name) VALUES (gen_random_uuid(), '{CATEGORY}') ON CONFLICT (name) DO NOTHING;""")

def ensure_subcat(name):
    if name not in seen_subs:
        sql(f"""INSERT INTO subcategories (id, category_id, name) SELECT gen_random_uuid(), (SELECT id FROM categories WHERE name = '{CATEGORY}'), '{name}' WHERE NOT EXISTS (SELECT 1 FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = '{name}');""")
        seen_subs.add(name)

def parse_price(v):
    if v is None or v == '-':
        return None
    if isinstance(v, str):
        return float(v.replace(',', '.'))
    return float(v)

def parse_combined(val):
    if val is None or val == '-' or val == '':
        return None, None
    s = str(val).strip()
    m = re.match(r'^(.+?)\s*\(([\d.]+)\)$', s)
    if m:
        size_part = m.group(1).strip()
        price_val = float(m.group(2))
        return size_part, price_val
    return None, None

# ============ Sheet 1: الأكواع والمشتركات والجلاب ============
ws1 = wb[wb.sheetnames[0]]

# --- Section 1: الأكواع (R4-R11) ---
# Col A=size, B=كوع عادي, C=كوع باب, D=كوع مفتوح
subcat_names_1 = ['كوع عادي 110', 'كوع باب 110', 'كوع مفتوح 110']
col_indices_1 = [1, 2, 3]  # B, C, D

for sc_name, col_idx in zip(subcat_names_1, col_indices_1):
    ensure_subcat(sc_name)
    for r in range(5, 12):
        row = list(ws1.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if not row[0]:
            continue
        size_val = str(row[0]).strip()
        price = parse_price(row[col_idx])
        if price:
            pname = f'{sc_name.replace(" 110", "")} {size_val}'
            clean_size = size_val.replace("'", '').replace('"', '').strip()
            sql(f"""INSERT INTO products (id, subcategory_id, name, unit, retail_price, wholesale_price, cost_price, is_active, company, size, stock_status) SELECT gen_random_uuid(), (SELECT id FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = '{sc_name}'), '{pname}', 'عدد', {Decimal(str(price))}, 0, 0, true, '{CATEGORY}', '{clean_size}', 'untracked' WHERE NOT EXISTS (SELECT 1 FROM products WHERE name = '{pname}' AND company = '{CATEGORY}');""")

# --- Section 2: المشتركات (R14-R21) ---
# Col A=size, B=مشترك واي, C=مشترك مسلوب (combined)
ensure_subcat('مشترك واي 110')
ensure_subcat('مشترك مسلوب 110')

for r in range(15, 22):
    row = list(ws1.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    if not row[0]:
        continue
    size_val = str(row[0]).strip()
    
    # مشترك واي - plain price in col B
    price = parse_price(row[1])
    if price:
        pname = f'مشترك واي {size_val}'
        sql(f"""INSERT INTO products (id, subcategory_id, name, unit, retail_price, wholesale_price, cost_price, is_active, company, size, stock_status) SELECT gen_random_uuid(), (SELECT id FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = 'مشترك واي 110'), '{pname}', 'عدد', {Decimal(str(price))}, 0, 0, true, '{CATEGORY}', '{size_val}', 'untracked' WHERE NOT EXISTS (SELECT 1 FROM products WHERE name = '{pname}' AND company = '{CATEGORY}');""")
    
    # مشترك مسلوب - combined format "size*size (price)"
    sub_size, sub_price = parse_combined(row[2])
    if sub_price:
        pname = f'مشترك مسلوب {sub_size}'
        cln = sub_size.replace("'", '').replace('"', '').strip()
        sql(f"""INSERT INTO products (id, subcategory_id, name, unit, retail_price, wholesale_price, cost_price, is_active, company, size, stock_status) SELECT gen_random_uuid(), (SELECT id FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = 'مشترك مسلوب 110'), '{pname}', 'عدد', {Decimal(str(sub_price))}, 0, 0, true, '{CATEGORY}', '{cln}', 'untracked' WHERE NOT EXISTS (SELECT 1 FROM products WHERE name = '{pname}' AND company = '{CATEGORY}');""")

# --- Section 3: جلب لحام وبوش (R24-R31) ---
# Col A=size, B=جلبة لحام, C=بوش, D=تابع البوش
ensure_subcat('جلبة لحام 110')
ensure_subcat('بوش 110')
ensure_subcat('تابع بوش 110')

for r in range(25, 32):
    row = list(ws1.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    if not row[0]:
        continue
    size_val = str(row[0]).strip()
    
    # جلبة لحام - plain price
    price = parse_price(row[1])
    if price:
        pname = f'جلبة لحام {size_val}'
        sql(f"""INSERT INTO products (id, subcategory_id, name, unit, retail_price, wholesale_price, cost_price, is_active, company, size, stock_status) SELECT gen_random_uuid(), (SELECT id FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = 'جلبة لحام 110'), '{pname}', 'عدد', {Decimal(str(price))}, 0, 0, true, '{CATEGORY}', '{size_val}', 'untracked' WHERE NOT EXISTS (SELECT 1 FROM products WHERE name = '{pname}' AND company = '{CATEGORY}');""")
    
    # بوش - combined
    sub_size, sub_price = parse_combined(row[2])
    if sub_price:
        pname = f'بوش {sub_size}'
        cln = sub_size.replace("'", '').replace('"', '').strip()
        sql(f"""INSERT INTO products (id, subcategory_id, name, unit, retail_price, wholesale_price, cost_price, is_active, company, size, stock_status) SELECT gen_random_uuid(), (SELECT id FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = 'بوش 110'), '{pname}', 'عدد', {Decimal(str(sub_price))}, 0, 0, true, '{CATEGORY}', '{cln}', 'untracked' WHERE NOT EXISTS (SELECT 1 FROM products WHERE name = '{pname}' AND company = '{CATEGORY}');""")
    
    # تابع البوش - combined
    sub_size, sub_price = parse_combined(row[3])
    if sub_price:
        pname = f'تابع بوش {sub_size}'
        cln = sub_size.replace("'", '').replace('"', '').strip()
        sql(f"""INSERT INTO products (id, subcategory_id, name, unit, retail_price, wholesale_price, cost_price, is_active, company, size, stock_status) SELECT gen_random_uuid(), (SELECT id FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = 'تابع بوش 110'), '{pname}', 'عدد', {Decimal(str(sub_price))}, 0, 0, true, '{CATEGORY}', '{cln}', 'untracked' WHERE NOT EXISTS (SELECT 1 FROM products WHERE name = '{pname}' AND company = '{CATEGORY}');""")

# ============ Sheet 2: أصناف متنوعة وطبات وبسن ============
ws2 = wb[wb.sheetnames[1]]
ensure_subcat('أصناف متنوعة 110')

for r in range(4, 33):
    row = list(ws2.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    if not row[0] or not row[0]:
        continue
    pname = str(row[0]).strip()
    price = parse_price(row[1])
    if price:
        sql(f"""INSERT INTO products (id, subcategory_id, name, unit, retail_price, wholesale_price, cost_price, is_active, company, size, stock_status) SELECT gen_random_uuid(), (SELECT id FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = 'أصناف متنوعة 110'), '{pname}', 'عدد', {Decimal(str(price))}, 0, 0, true, '{CATEGORY}', NULL, 'untracked' WHERE NOT EXISTS (SELECT 1 FROM products WHERE name = '{pname}' AND company = '{CATEGORY}');""")

# Footer
sql(f"""
SELECT 'categories' AS tbl, count(*) FROM categories WHERE name = '{CATEGORY}'
UNION ALL SELECT 'subcategories', count(*) FROM subcategories s JOIN categories c ON s.category_id = c.id WHERE c.name = '{CATEGORY}'
UNION ALL SELECT 'products', count(*) FROM products p WHERE p.company = '{CATEGORY}';
""")

output = r'C:\eg-co-erp\import_rok110.sql'
with open(output, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
print(f"Rok 110: SQL written to {output}")
