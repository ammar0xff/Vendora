import sys
import openpyxl
import uuid
import psycopg2
from decimal import Decimal

sys.stdout.reconfigure(encoding='utf-8')

DB_PARAMS = {
    'host': 'localhost',
    'port': 5432,
    'dbname': 'inventory_db',
    'user': 'postgres',
    'password': 'postgres',
}

conn = psycopg2.connect(**DB_PARAMS)
cur = conn.cursor()

def insert_category(name):
    """Insert a category, return its UUID."""
    cur.execute("SELECT id FROM categories WHERE name = %s", (name,))
    row = cur.fetchone()
    if row:
        return row[0]
    cid = uuid.uuid4()
    cur.execute("INSERT INTO categories (id, name) VALUES (%s, %s)", (cid, name))
    return cid

def insert_subcategory(category_id, name):
    """Insert a subcategory, return its UUID."""
    cur.execute("SELECT id FROM subcategories WHERE category_id = %s AND name = %s", (category_id, name))
    row = cur.fetchone()
    if row:
        return row[0]
    sid = uuid.uuid4()
    cur.execute("INSERT INTO subcategories (id, category_id, name) VALUES (%s, %s, %s)", (sid, category_id, name))
    return sid

def product_exists(name, company, size_val):
    """Check if product already exists."""
    cur.execute(
        "SELECT id FROM products WHERE name = %s AND company = %s AND size = %s",
        (name, company, size_val if size_val else None)
    )
    return cur.fetchone()

def insert_product(subcategory_id, name, size_val, thickness, price, company, material=None):
    """Insert a product."""
    # Build product name including thickness if provided
    full_name = name
    if thickness:
        full_name = f"{name} - سمك {thickness}"
    
    existing = product_exists(full_name, company, size_val)
    if existing:
        return existing[0]
    
    pid = uuid.uuid4()
    cur.execute("""
        INSERT INTO products (id, subcategory_id, name, unit, retail_price, company, size, material, stock_status)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (pid, subcategory_id, full_name, 'عدد', Decimal(str(price)) if price else Decimal('0'),
          company, size_val if size_val else None, material, 'untracked'))
    return pid

print("=== Importing الشريف catalog ===")
company_el_sharif = 'الشريف'
cat_el_sharif_id = insert_category(company_el_sharif)

path = r'C:\eg-co-erp\LISTS\الشريف.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['All_Products']

# Create subcategories & insert products
sharif_subcategories = {}
total = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    system_type, category_name, product_name, size_val, thickness, price = row
    if not category_name or not product_name:
        continue
    
    cat_name = str(category_name).strip()
    if cat_name not in sharif_subcategories:
        subcat_id = insert_subcategory(cat_el_sharif_id, cat_name)
        sharif_subcategories[cat_name] = subcat_id
    else:
        subcat_id = sharif_subcategories[cat_name]
    
    insert_product(
        subcat_id,
        str(product_name).strip(),
        str(size_val).strip() if size_val else None,
        str(thickness).strip() if thickness else None,
        price,
        company_el_sharif
    )
    total += 1

print(f"  Imported {total} products in {len(sharif_subcategories)} subcategories")

print("\n=== Importing ايجيك catalog ===")
company_egytec = 'ايجيك'
cat_egytec_id = insert_category(company_egytec)

path2 = r'C:\eg-co-erp\LISTS\ايجيك.xlsx'
wb2 = openpyxl.load_workbook(path2, data_only=True)
ws2 = wb2['Sheet1']

# Egytec subcategories - group by material type prefix  
# We'll use the first 2-3 words as the subcategory name
egytec_subcategories = {}
total2 = 0
skipped_no_price = 0

for row in ws2.iter_rows(min_row=2, values_only=True):
    code, material, brand, price = row
    if not material:
        continue
    
    material_str = str(material).strip()
    brand_str = str(brand).strip() if brand else ''
    
    # Determine subcategory from material prefix
    words = material_str.split()
    if len(words) >= 3:
        subcat_name = ' '.join(words[:2])
    elif len(words) == 2:
        subcat_name = ' '.join(words[:2])
    else:
        subcat_name = words[0]
    
    # Clean up subcategory name
    subcat_name = subcat_name.strip()
    
    if subcat_name not in egytec_subcategories:
        subcat_id = insert_subcategory(cat_egytec_id, subcat_name)
        egytec_subcategories[subcat_name] = subcat_id
    else:
        subcat_id = egytec_subcategories[subcat_name]
    
    if price is None:
        skipped_no_price += 1
        continue
    
    # Extract size from material name if possible
    size_val = None
    for word in words:
        if any(c in word for c in 'م/"\"\''):
            size_val = word
            break
    
    pid = insert_product(
        subcat_id,
        material_str,
        size_val,
        None,
        price,
        company_egytec,
        material=brand_str
    )
    total2 += 1

print(f"  Imported {total2} products in {len(egytec_subcategories)} subcategories")
print(f"  Skipped {skipped_no_price} products without prices")

conn.commit()
cur.close()
conn.close()

print("\n=== Import completed successfully! ===")
