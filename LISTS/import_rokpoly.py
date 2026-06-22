import openpyxl, datetime
from decimal import Decimal

def parse_price(val):
    if val is None or val == '-':
        return None
    if isinstance(val, datetime.datetime):
        return round(val.month + val.day / 100, 2)
    if isinstance(val, str):
        return float(val.replace(',', '.'))
    if isinstance(val, (int, float)):
        return float(val)
    return None

path = r'C:\eg-co-erp\LISTS\قائمة أسعار روك بولي 2024 - جداول منظمة.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb[wb.sheetnames[0]]

lines = []
def sql(s):
    lines.append(s)

CATEGORY = 'روك بولي'

# Category definition: (subcategory_name, data_start, data_end, col_mapping)
# col_mapping: list of (name_in_arabic, col_for_qty, col_for_price)
# columns: A=size, B=col2, C=col3, D=col4, E=col5, F=col6, G=col7

sections = [
    ('جلب لحام', 5, 9, [('جلبة لحام', 1, 2)]),
    ('تي لحام', 5, 9, [('تي لحام', 3, 4)]),
    ('كوع لحام', 5, 9, [('كوع لحام 45', 5, 6)]),
    ('جلب أنثى', 13, 18, [('جلبة أنثى', 1, 2)]),
    ('تي بسن', 13, 18, [('تي بسن', 3, 4)]),
    ('كوع بسن', 13, 18, [('كوع بسن', 5, 6)]),
    ('كوع مفتوح وطبة اختبار', 22, 24, [('كوع مفتوح', 1, 2), ('طبة اختبار', 1, 2)]),
    ('تي مسلوب', 28, 37, [('تي مسلوب', 1, 2)]),
    ('جلب مسلوب', 28, 37, [('جلبة مسلوب', 3, 4)]),
    ('جلب ذكر', 28, 37, [('جلبة ذكر', 5, 6)]),
    ('كرنك حقن', 41, 42, [('كرنك حقن ثقيل', 0, 1)]),  # col B = price, no qty
    ('طبة كاب', 46, 50, [('طبة كاب', 1, 2)]),
    ('محبس دفع', 46, 50, [('محبس دفع', 3, 4)]),
]

# For section with multiple product types from same rows, we need to handle specially
# Sections 1-3 share rows 5-9, same for other grouped sections

seen_subs = set()
total_products = 0

# Handle sections 1-3: different subcategories from same rows
sections_1_3 = [('جلب لحام', 'جلبة لحام', 1, 2), ('تي لحام', 'تي لحام', 3, 4), ('كوع لحام', 'كوع لحام 45', 5, 6)]
for subcat, prod_name, qty_col, price_col in sections_1_3:
    if subcat not in seen_subs:
        sql(f"""INSERT INTO subcategories (id, category_id, name) SELECT gen_random_uuid(), (SELECT id FROM categories WHERE name = '{CATEGORY}'), '{subcat}' WHERE NOT EXISTS (SELECT 1 FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = '{subcat}');""")
        seen_subs.add(subcat)
    for r in range(5, 10):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if not row[0]:
            continue
        size_val = str(row[0]).strip()
        price = parse_price(row[price_col])
        if price:
            pname = f'{prod_name} {size_val}'
            price_val = Decimal(str(price))
            sql(f"""INSERT INTO products (id, subcategory_id, name, unit, retail_price, wholesale_price, cost_price, is_active, company, size, stock_status) SELECT gen_random_uuid(), (SELECT id FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = '{subcat}'), '{pname}', 'عدد', {price_val}, 0, 0, true, '{CATEGORY}', '{size_val}', 'untracked' WHERE NOT EXISTS (SELECT 1 FROM products WHERE name = '{pname}' AND company = '{CATEGORY}');""")
            total_products += 1

# Sections 4-6: Threaded fittings  
sections_4_6 = [('جلب أنثى', 'جلبة أنثى', 1, 2), ('تي بسن', 'تي بسن', 3, 4), ('كوع بسن', 'كوع بسن', 5, 6)]
for subcat, prod_name, qty_col, price_col in sections_4_6:
    if subcat not in seen_subs:
        sql(f"""INSERT INTO subcategories (id, category_id, name) SELECT gen_random_uuid(), (SELECT id FROM categories WHERE name = '{CATEGORY}'), '{subcat}' WHERE NOT EXISTS (SELECT 1 FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = '{subcat}');""")
        seen_subs.add(subcat)
    for r in range(13, 19):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if not row[0]:
            continue
        size_val = str(row[0]).strip()
        price = parse_price(row[price_col])
        if price:
            pname = f'{prod_name} {size_val}'
            price_val = Decimal(str(price))
            sql(f"""INSERT INTO products (id, subcategory_id, name, unit, retail_price, wholesale_price, cost_price, is_active, company, size, stock_status) SELECT gen_random_uuid(), (SELECT id FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = '{subcat}'), '{pname}', 'عدد', {price_val}, 0, 0, true, '{CATEGORY}', '{size_val}', 'untracked' WHERE NOT EXISTS (SELECT 1 FROM products WHERE name = '{pname}' AND company = '{CATEGORY}');""")
            total_products += 1

# Section: كوع مفتوح وطبة اختبار (rows 22-24)
subcat = 'كوع مفتوح وطبة اختبار'
sql(f"""INSERT INTO subcategories (id, category_id, name) SELECT gen_random_uuid(), (SELECT id FROM categories WHERE name = '{CATEGORY}'), '{subcat}' WHERE NOT EXISTS (SELECT 1 FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = '{subcat}');""")
seen_subs.add(subcat)
for r in range(22, 25):
    row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    if not row[0]:
        continue
    raw_name = str(row[0]).strip()
    price = parse_price(row[2])
    if price:
        pname = f'{raw_name}'
        price_val = Decimal(str(price))
        sql(f"""INSERT INTO products (id, subcategory_id, name, unit, retail_price, wholesale_price, cost_price, is_active, company, size, stock_status) SELECT gen_random_uuid(), (SELECT id FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = '{subcat}'), '{pname}', 'عدد', {price_val}, 0, 0, true, '{CATEGORY}', NULL, 'untracked' WHERE NOT EXISTS (SELECT 1 FROM products WHERE name = '{pname}' AND company = '{CATEGORY}');""")
        total_products += 1

# Sections: تي مسلوب, جلب مسلوب, جلب ذكر (rows 28-37)
sections_red = [('تي مسلوب', 'تي مسلوب', 1, 2), ('جلب مسلوب', 'جلبة مسلوب', 3, 4), ('جلب ذكر', 'جلبة ذكر', 5, 6)]
for subcat, prod_name, qty_col, price_col in sections_red:
    if subcat not in seen_subs:
        sql(f"""INSERT INTO subcategories (id, category_id, name) SELECT gen_random_uuid(), (SELECT id FROM categories WHERE name = '{CATEGORY}'), '{subcat}' WHERE NOT EXISTS (SELECT 1 FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = '{subcat}');""")
        seen_subs.add(subcat)
    for r in range(28, 38):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if not row[0]:
            continue
        size_val = str(row[0]).strip()
        price = parse_price(row[price_col])
        if price:
            pname = f'{prod_name} {size_val}'
            price_val = Decimal(str(price))
            sql(f"""INSERT INTO products (id, subcategory_id, name, unit, retail_price, wholesale_price, cost_price, is_active, company, size, stock_status) SELECT gen_random_uuid(), (SELECT id FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = '{subcat}'), '{pname}', 'عدد', {price_val}, 0, 0, true, '{CATEGORY}', '{size_val}', 'untracked' WHERE NOT EXISTS (SELECT 1 FROM products WHERE name = '{pname}' AND company = '{CATEGORY}');""")
            total_products += 1

# Section: كرنك حقن (rows 41-42)
subcat = 'كرنك حقن'
sql(f"""INSERT INTO subcategories (id, category_id, name) SELECT gen_random_uuid(), (SELECT id FROM categories WHERE name = '{CATEGORY}'), '{subcat}' WHERE NOT EXISTS (SELECT 1 FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = '{subcat}');""")
seen_subs.add(subcat)
for r in range(41, 43):
    row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    if not row[0]:
        continue
    raw_name = str(row[0]).strip()
    price = parse_price(row[1])
    if price:
        pname = f'{raw_name}'
        price_val = Decimal(str(price))
        sql(f"""INSERT INTO products (id, subcategory_id, name, unit, retail_price, wholesale_price, cost_price, is_active, company, size, stock_status) SELECT gen_random_uuid(), (SELECT id FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = '{subcat}'), '{pname}', 'عدد', {price_val}, 0, 0, true, '{CATEGORY}', NULL, 'untracked' WHERE NOT EXISTS (SELECT 1 FROM products WHERE name = '{pname}' AND company = '{CATEGORY}');""")
        total_products += 1

# Sections: طبة كاب, محبس دفع (rows 46-50)
sections_cap = [('طبة كاب', 'طبة كاب', 1, 2), ('محبس دفع', 'محبس دفع', 3, 4)]
for subcat, prod_name, qty_col, price_col in sections_cap:
    if subcat not in seen_subs:
        sql(f"""INSERT INTO subcategories (id, category_id, name) SELECT gen_random_uuid(), (SELECT id FROM categories WHERE name = '{CATEGORY}'), '{subcat}' WHERE NOT EXISTS (SELECT 1 FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = '{subcat}');""")
        seen_subs.add(subcat)
    for r in range(46, 51):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if not row[0]:
            continue
        size_val = str(row[0]).strip()
        price = parse_price(row[price_col])
        if price and price > 0:
            pname = f'{prod_name} {size_val}'
            price_val = Decimal(str(price))
            sql(f"""INSERT INTO products (id, subcategory_id, name, unit, retail_price, wholesale_price, cost_price, is_active, company, size, stock_status) SELECT gen_random_uuid(), (SELECT id FROM subcategories WHERE category_id = (SELECT id FROM categories WHERE name = '{CATEGORY}') AND name = '{subcat}'), '{pname}', 'عدد', {price_val}, 0, 0, true, '{CATEGORY}', '{size_val}', 'untracked' WHERE NOT EXISTS (SELECT 1 FROM products WHERE name = '{pname}' AND company = '{CATEGORY}');""")
            total_products += 1

# Header
header = f"""INSERT INTO categories (id, name) VALUES (gen_random_uuid(), '{CATEGORY}') ON CONFLICT (name) DO NOTHING;"""
lines.insert(0, header)

footer = f"""
SELECT 'categories' AS tbl, count(*) FROM categories WHERE name = '{CATEGORY}'
UNION ALL SELECT 'subcategories', count(*) FROM subcategories s JOIN categories c ON s.category_id = c.id WHERE c.name = '{CATEGORY}'
UNION ALL SELECT 'products', count(*) FROM products p WHERE p.company = '{CATEGORY}';
"""
lines.append(footer)

output = r'C:\eg-co-erp\import_rokpoly.sql'
with open(output, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f"Rok Poly: {total_products} products, {len(seen_subs)} subcategories")
print(f"SQL written to {output}")
