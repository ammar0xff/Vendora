import openpyxl

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_أسعار_شيلد_يناير_2026.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]

# Print header
for c in range(1, 8):
    h = ws.cell(3, c).value
    print(f'Col{c}: {h}')

print()
print('--- All data rows ---')
for r in range(6, ws.max_row + 1):
    subcat = ws.cell(r, 1).value or ''
    prod = ws.cell(r, 2).value or ''
    size = ws.cell(r, 3).value or ''
    pack = ws.cell(r, 4).value or ''
    price = ws.cell(r, 5).value or ''
    specs = ws.cell(r, 6).value or ''
    print(f'Row {r}: [{subcat}] [{prod}] [{size}] [{pack}] [{price}] [{str(specs)[:50]}]')
