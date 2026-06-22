import openpyxl, subprocess, json, uuid

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\قائمة_اسعار_دورافيت_2026_عربي.xlsx', data_only=True)
ws = wb.active

rows = []
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)):
    vals = [str(v).strip() if v is not None else '' for v in row]
    if vals[2] and vals[3]:
        rows.append(vals)

print(f"Excel rows: {len(rows)}")

SERIES_MAP = {
    'بي 3 كومفورتس': 'P3 COMFORTS',
    'ستارك 3': 'STARCK 3',
    'دي نيو': 'D-NEO',
    'دورا ستايل': 'DURASTYLE',
    'دارلينج نيو': 'DARLING',
    'دارلينج': 'DARLING',
    'هابي دي': 'HAPPY D.',
    'هابي دي 2': 'HAPPY D.',
    'دورافيت نمبر 1': 'DURAVIT NO. 1',
    'دي كود': 'D-CODE',
    'إيكو': 'ECHO',
    'دورابلاس': 'DURAPLUS',
    'إميليا': 'EMILIA',
    'جولف': 'GOLF',
    'ديون': 'DUNE',
    'متنوعة': 'OTHERS',
    'إكسسوارات سيراميك': 'ACCESSORIES CERAMIC',
    'إكسسوارات إيزي': 'ACCESSORIES EASY',
    'إكسسوارات كروم': 'ACCESSORIES CHROME',
    'إكس لارج': 'X-LARGE',
    'فوستر': 'FOSTER',
    'كيثو': 'KETHO',
    'كارو': 'CARO',
    'إل كيوب': 'L-CUBE',
    'فيتريوم': 'VITRIUM',
    'بيورا فيدا': 'PURAVIDA',
    'ستارك 1': 'STARCK 1',
    'فيرو': 'VERO',
    'الطابق الثاني': 'SECOND FLOOR',
    'عمود الطابق الثاني': 'SECOND FLOOR COLUMN',
    'مجموعة تثبيت': 'MOUNTING SET',
}

# First, DELETE old records, then ensure ALL subcategories exist
sql_parts = []
sql_parts.append("DELETE FROM products WHERE company = 'دروفيت';")

# Create subcategories with fixed UUIDs (so we can reference them)
cat_guids = {}
for arabic, eng in SERIES_MAP.items():
    if eng not in cat_guids:
        cat_guids[eng] = str(uuid.uuid4())

for arabic, eng in SERIES_MAP.items():
    g = cat_guids[eng]
    safe = eng.replace("'", "''")
    sql_parts.append(
        f"INSERT INTO subcategories (id, name) VALUES ('{g}', E'{safe}') "
        f"ON CONFLICT (name) DO UPDATE SET name = EXCLUDED.name;"
    )

# Map arabic series -> subcategory GUID
arabic_to_guid = {}
for arabic, eng in SERIES_MAP.items():
    arabic_to_guid[arabic] = cat_guids[eng]

# Need a fallback for any unmapped
fallback_guid = str(uuid.uuid4())
sql_parts.append(
    f"INSERT INTO subcategories (id, name) VALUES ('{fallback_guid}', 'OTHERS') "
    f"ON CONFLICT (name) DO NOTHING;"
)
# Also get OTHERS GUID in case it already exists
for arabic, eng in SERIES_MAP.items():
    if eng == 'OTHERS':
        fallback_guid = cat_guids[eng]

unit = 'قطعة'
now = "NOW()"
insert_count = 0

for vals in rows:
    series_ar = vals[1]
    name = vals[2]
    code = vals[3]
    size = vals[4]
    color = vals[5]
    price_str = vals[6]
    
    subcat_id = arabic_to_guid.get(series_ar, fallback_guid)
    
    price_val = 0
    try:
        price_val = float(price_str.replace(',', ''))
    except:
        pass
    
    prod_id = str(uuid.uuid4())
    safe_name = name.replace("'", "''")
    safe_size = size.replace("'", "''") if size else ''
    safe_color = color.replace("'", "''") if color else ''
    
    sql_parts.append(
        f"INSERT INTO products (id, subcategory_id, name, barcode, unit, "
        f"retail_price, wholesale_price, cost_price, company, size, type, material, "
        f"is_active, created_at, updated_at, stock_status) VALUES ("
        f"'{prod_id}', '{subcat_id}', E'{safe_name}', '{code}', '{unit}', "
        f"{price_val}, {price_val}, {price_val}, 'دروفيت', E'{safe_size}', '', '', "
        f"true, {now}, {now}, 'untracked');"
    )
    insert_count += 1

print(f"Products to insert: {insert_count}")
print(f"SQL statements: {len(sql_parts)}")

sql_path = r'C:\eg-co-erp\reimport_drovit_v3.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_parts))

subprocess.run(['docker', 'cp', sql_path, 'eg-co-erp-db-1:/tmp/reimport_drovit_v3.sql'], capture_output=True)
r2 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db', '-f', '/tmp/reimport_drovit_v3.sql'],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)

error_count = r2.stderr.count('ERROR') if r2.stderr else 0
print(f"Errors: {error_count}")
if r2.stderr and error_count > 0:
    errors = [l for l in r2.stderr.split('\n') if 'ERROR' in l]
    for e in errors[:5]:
        print(f"  {e[:150]}")

r3 = subprocess.run(
    ['docker', 'exec', 'eg-co-erp-db-1', 'psql', '-U', 'postgres', '-d', 'inventory_db',
     '-t', '-A',
     '-c', "SELECT count(*) FROM products WHERE company = 'دروفيت'"],
    capture_output=True, text=True, encoding='utf-8', errors='replace'
)
print(f"دروفيت in DB: {r3.stdout.strip()}")
