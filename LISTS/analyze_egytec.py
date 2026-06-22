import openpyxl, sys, re
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8')

path = r'C:\eg-co-erp\LISTS\ايجيك.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Sheet1']

# Extract first 2-3 words of each material to find groupings
prefixes = Counter()
materials_by_prefix = {}

for row in ws.iter_rows(min_row=2, values_only=True):
    code, material, brand, price = row
    if not material:
        continue
    m = str(material).strip()
    # Get first 2 words
    words = m.split()
    if len(words) >= 2:
        prefix = ' '.join(words[:2])
    elif len(words) == 1:
        prefix = words[0]
    else:
        continue
    prefixes[prefix] += 1
    if prefix not in materials_by_prefix:
        materials_by_prefix[prefix] = []
    materials_by_prefix[prefix].append((m, brand, price))

print('=== Egytec - Top Material Prefixes (potential subcategories) ===')
for prefix, count in prefixes.most_common(30):
    print(f'\n--- {prefix} ({count} products) ---')
    for m, brand, price in materials_by_prefix[prefix][:3]:
        print(f'    {m} | {brand} | {price}')

# Also look at the material names that start with common patterns
print('\n\n=== Brand distribution ===')
brand_count = Counter()
for row in ws.iter_rows(min_row=2, values_only=True):
    code, material, brand, price = row
    if brand:
        brand_count[str(brand).strip()] += 1
for b, c in brand_count.most_common():
    print(f'  {b}: {c} products')

# Check for codes structure
print('\n\n=== Code prefix distribution ===')
code_prefix = Counter()
for row in ws.iter_rows(min_row=2, values_only=True):
    code = row[0]
    if code:
        cp = str(code)[:3]
        code_prefix[cp] += 1
for cp, c in code_prefix.most_common(20):
    print(f'  {cp}: {c} products')
