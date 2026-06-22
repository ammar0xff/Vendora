import openpyxl

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\done\الشريف.xlsx', data_only=True)
print('Sheets:', wb.sheetnames)
ws = wb[wb.sheetnames[0]]
print(f'Sheet: {wb.sheetnames[0]}')
print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')

# Print first 10 rows to understand structure
for r in range(1, min(ws.max_row + 1, 12)):
    vals = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r, c).value
        vals.append(str(v or '')[:30])
    print(f'Row {r}: {" | ".join(vals)}')

# Also print a few data rows from the middle
print('\n--- Sample data rows ---')
for r in range(100, 110):
    vals = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r, c).value
        vals.append(str(v or '')[:40])
    print(f'Row {r}: {" | ".join(vals)}')
