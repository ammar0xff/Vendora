import openpyxl, sys, re
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\Neisco_Comer_Price_List_2026.xlsx', data_only=True)

codes = set()
sizes_with_codes = {}
for sname in wb.sheetnames:
    if sname in ('Sheet1', 'Worksheet'): 
        continue
    ws = wb[sname]
    for row in ws.iter_rows(min_row=4, values_only=True):
        code, size, pack, box, price = row if len(row) >= 5 else (row[0] if row else None, None, None, None, None)
        if not code or not size:
            continue
        code_str = str(code).strip()
        size_str = str(size).strip()
        if not price or str(price).strip() in ('', '-'):
            continue
        codes.add(code_str)
        sizes_with_codes[code_str] = size_str

# Find all unique letter-digit prefixes (e.g., BVSL10N, SO11N)
prefixes = set()
for c in sorted(codes):
    # Match pattern like BVSL10N or BVDL10N or SA512 (letters+digits)
    m = re.match(r'^([A-Z]+)(\d+.*)', c)
    if m:
        prefix = m.group(1) + m.group(2)[:3]  # Take only the first few digits
        while len(prefix) > 2 and not prefix[-1].isdigit():
            prefix = prefix[:-1]
        # Get the full alphanumeric start
        m2 = re.match(r'^([A-Z]+\d+[A-Z]?\d*)', c)
        if m2:
            prefixes.add(m2.group(1))

print("All unique codes:")
for c in sorted(codes):
    print(f'  {c:15s} | {sizes_with_codes[c]}')

print(f"\n\nTotal unique codes: {len(codes)}")
print("Codes sorted by prefix pattern:")
for p in sorted(prefixes):
    print(f'  "{p}"')
