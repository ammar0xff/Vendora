import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\Neisco_Comer_Price_List_2026.xlsx', data_only=True)

ws = wb['Saddles & Bushings Threaded']
for row in ws.iter_rows(min_row=4, values_only=True):
    code, size, pack, box, price = row if len(row) >= 5 else (row[0] if row else None, None, None, None, None)
    if not code:
        continue
    code_str = str(code).strip()
    if 'SA512' in code_str or 'VBRZ' in code_str or 'RE21' in code_str:
        if size is None:
            print(f'{code_str}: size=None, price={price}')
        else:
            size_str = str(size).strip()
            size_repr = repr(size)
            price_str = str(price).strip() if price else 'None'
            print(f'{code_str}: size={size_str!r} (type={type(size).__name__}), price={price!r}')
            try:
                price_float = float(price)
                print(f'  price as float: {price_float}')
            except:
                pass

# Also check BVSL12N110L
ws2 = wb['UPVC Ball Valves']
for row in ws2.iter_rows(min_row=4, values_only=True):
    code, size, pack, box, price = row if len(row) >= 5 else (row[0] if row else None, None, None, None, None)
    if not code:
        continue
    code_str = str(code).strip()
    if 'BVSL12' in code_str or 'BVSL15' in code_str or 'BVSL19' in code_str or 'BVDL19' in code_str or 'BVS17' in code_str or 'BVDLC19' in code_str or 'FV10' in code_str:
        if size is None:
            print(f'{code_str}: size=None, price={price}')
            continue
        size_str = str(size).strip()
        print(f'{code_str}: size={size_str!r} (type={type(size).__name__}), price={price!r}')
        try:
            price_float = float(price)
            print(f'  price as float: {price_float}')
        except:
            pass
