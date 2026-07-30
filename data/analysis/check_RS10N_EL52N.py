import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\Neisco_Comer_Price_List_2026.xlsx', data_only=True)

# Check all RS10N entries in the Stubs Flanges Unions Metric sheet
ws = wb['Stubs Flanges Unions Metric']
print("=== RS10N entries ===")
for row in ws.iter_rows(min_row=4, values_only=True):
    code, size, pack, box, price = row if len(row) >= 5 else (row[0] if row else None, None, None, None, None)
    if not code:
        continue
    code_str = str(code).strip()
    if 'RS10N' in code_str:
        size_str = str(size).strip() if size else 'None'
        price_str = str(price) if price else 'None'
        print(f'{code_str:15s} | {size_str:15s} | price={price_str}')

# Also show all unique code prefixes from this sheet
print("\n=== All unique codes in Stubs Flanges Unions Metric ===")
seen = set()
for row in ws.iter_rows(min_row=4, values_only=True):
    code, size, pack, box, price = row if len(row) >= 5 else (row[0] if row else None, None, None, None, None)
    if not code:
        continue
    code_str = str(code).strip()
    size_str = str(size).strip() if size else '?'
    price_str = str(price)[:8] if price else '?'
    prefix = code_str[:6]
    if prefix not in seen:
        seen.add(prefix)
        print(f'{code_str:15s} | {size_str:15s} | price={price_str}')

# Also check Adaptor sheet for EL52N and compute all prices
ws2 = wb['Adaptor Series Plain-Threaded']
print("\n=== EL52N entries in Adaptor Series ===")
for row in ws2.iter_rows(min_row=4, values_only=True):
    code, size, pack, box, price = row if len(row) >= 5 else (row[0] if row else None, None, None, None, None)
    if not code:
        continue
    code_str = str(code).strip()
    if 'EL52' in code_str:
        size_str = str(size).strip() if size else 'None'
        price_str = str(price) if price else 'None'
        print(f'{code_str:15s} | {size_str:15s} | price={price_str}')
