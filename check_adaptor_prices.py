import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\eg-co-erp\LISTS\Neisco_Comer_Price_List_2026.xlsx', data_only=True)

# Check ALL entries in Adaptor Series with sizes like 50* or 63*
ws = wb['Adaptor Series Plain-Threaded']
print("=== All Adaptor entries with 50* or 63* sizes ===")
for row in ws.iter_rows(min_row=4, values_only=True):
    code, size, pack, box, price = row if len(row) >= 5 else (row[0] if row else None, None, None, None, None)
    if not code or not size:
        continue
    code_str = str(code).strip()
    size_str = str(size).strip()
    if '50*' in size_str or '63*' in size_str:
        price_str = str(price) if price else 'None'
        print(f'{code_str:20s} | {size_str:20s} | price={price_str}')

# Also check if there are any other products in محولات subcategory with these sizes in DB
print("\n=== All rows in Adaptor sheet ===")
for row in ws.iter_rows(min_row=4, values_only=True):
    code, size, pack, box, price = row if len(row) >= 5 else (row[0] if row else None, None, None, None, None)
    if not code:
        continue
    code_str = str(code).strip()
    size_str = str(size).strip() if size else '?'
    price_str = str(price)[:8] if price else '?'
    print(f'{code_str:20s} | {size_str:20s} | price={price_str}')

# Also check if there are multiple column layouts - maybe price was read from wrong column
print("\n\n=== Checking full row for EL52N050F and EL52N063G ===")
for row in ws.iter_rows(min_row=4, values_only=True):
    code = row[0]
    if not code:
        continue
    code_str = str(code).strip()
    if code_str in ['EL52N050F', 'EL52N063G']:
        # Print all columns
        print(f'Code: {code_str}')
        for i, val in enumerate(row):
            print(f'  Column {i}: {val!r}')
